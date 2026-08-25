from typing import List, Optional

from types import SimpleNamespace



from fastapi import APIRouter, UploadFile, File, Query, HTTPException, Header, Form

from sqlalchemy.orm import Session

from sqlalchemy import or_, and_, func

from fastapi.responses import FileResponse, StreamingResponse

from fastapi import Depends

from pydantic import BaseModel, Field

from app.database.database import get_db

import pandas as pd

from datetime import datetime, date, timedelta

import hashlib

import json

import shutil

import os

import re

import uuid

from openpyxl.styles import PatternFill

import subprocess

import tempfile

import xml.etree.ElementTree as ET

from collections import Counter



from app.database.database import SessionLocal



from app.models.sku_master import (

    SKUMaster,

    SKUPiece

)

from app.models.daily_report import DailyReport

from app.models.daily_sales_report import DailySalesReport

from app.models.return_inventory import ReturnInventory

from app.models.stock_inventory import StockInventory

from app.models.sticker_inventory import StickerInventory

from app.models.packing_inventory import PackingInventory, PackingInventoryUsage
from datetime import date

from app.models.sales_upload import SalesUpload

from app.models.inventory_deduction_log import InventoryDeductionLog

from app.models.sales_analytics_detail import SalesAnalyticsDetail

from app.models.flipkart_zone_report import (

    FlipkartZoneBatch,

    FlipkartZoneBatchItem,

    FlipkartZoneReport,

)



from app.services.excel_service import (

    clean_color_name,

    normalize_column_name,

    normalize_sku,



    read_excel_file,

    read_sheet_columns,

    read_csv_columns,

    read_sku_sheet,



    filter_flipkart_orders,

    get_flipkart_target_date,

    filter_amazon_orders,

    filter_ajio_orders,

    filter_meesho_orders,

    filter_flipkart_warehouse_orders,



    aggregate_orders,

    expand_inventory,



    generate_daily_report,

    filter_myntra_orders,

    deduct_return_inventory,

    merge_return_inventory_rows,



)



router = APIRouter()

from app.services.inventory_snapshot_cache import (
    invalidate_inventory_snapshot,
    read_inventory_snapshot,
    write_inventory_snapshot,
)

def _invalidate_sku_master_caches():
    for cache_key in ["sticker_sidebar", "stock", "sticker", "return"]:
        invalidate_inventory_snapshot(cache_key)



BASE_DIR = os.path.abspath(

    os.path.join(

        os.path.dirname(__file__),

        "..",

        "..",

    )

)

UPLOAD_FOLDER = os.path.join(

    BASE_DIR,

    "uploads"

)

PLATFORM_NAMES = [

    "Flipkart",

    "Amazon",

    "Ajio",

    "Meesho",

    "Myntra",

]

REPORTING_PLATFORM_NAMES = [

    *PLATFORM_NAMES,

    "Flipkart Warehouse",

]

PACKING_INVENTORY_TYPES = [
    "shipping_label",
    "shipping_cover",
    "packing_cover",
    "packing_board",
]

PACKING_INVENTORY_LABELS = {
    "shipping_label": "Shipping Label",
    "shipping_cover": "Shipping Cover",
    "packing_cover": "Packing Cover",
    "packing_board": "Packing Board",
}

class ManualSKUPieceCreate(BaseModel):
    color: Optional[str] = ""

class ManualSKUMasterCreate(BaseModel):
    platform: Optional[str] = "Common"
    sku: str = Field(..., min_length=1)
    style: Optional[str] = ""
    size: Optional[str] = ""
    pack_of: Optional[str] = ""
    full_color: Optional[str] = ""
    pieces: List[ManualSKUPieceCreate] = Field(default_factory=list)

class SKUMasterRowUpdate(ManualSKUMasterCreate):
    id: Optional[int] = None

class SKUMasterBulkUpdate(BaseModel):
    items: List[SKUMasterRowUpdate]
    deleted_ids: List[int] = Field(default_factory=list)

class PackingInventoryRowUpdate(BaseModel):
    id: Optional[int] = None
    item_type: str = Field(..., min_length=1)
    platform: Optional[str] = None

    name: str = Field(..., min_length=1)

    qty: int = Field(default=0, ge=0)





class PackingInventoryBulkUpdate(BaseModel):

    items: List[PackingInventoryRowUpdate] = Field(default_factory=list)

    deleted_ids: List[int] = Field(default_factory=list)



os.makedirs(UPLOAD_FOLDER, exist_ok=True)





def _clean_upload_filename(filename: str):

    filename = os.path.basename(

        filename or ""

    ).strip()



    if not filename:

        raise HTTPException(

            status_code=400,

            detail="Uploaded file is missing a filename.",

        )



    return re.sub(

        r"[^A-Za-z0-9._() -]",

        "_",

        filename,

    )





def _save_upload(upload_file: UploadFile):

    filename = _clean_upload_filename(

        upload_file.filename

    )



    file_path = os.path.join(

        UPLOAD_FOLDER,

        filename

    )



    if os.path.exists(file_path):

        name, ext = os.path.splitext(filename)

        file_path = os.path.join(

            UPLOAD_FOLDER,

            f"{name}_{uuid.uuid4().hex[:8]}{ext}"

        )



    try:

        with open(file_path, "wb") as buffer:

            shutil.copyfileobj(

                upload_file.file,

                buffer

            )

    except OSError as e:

        raise HTTPException(

            status_code=500,

            detail=f"Could not save uploaded file: {e}",

        )



    upload_file.saved_path = file_path

    upload_file.saved_filename = os.path.basename(

        file_path

    )



    return file_path






import time

def _cleanup_old_uploads(days: int = 7):
    try:
        now = time.time()
        for filename in os.listdir(UPLOAD_FOLDER):
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(file_path):
                if os.stat(file_path).st_mtime < now - days * 86400:
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"Failed to delete old upload {filename}: {e}")
    except Exception as e:
        print(f"Error during cleanup: {e}")


def _upload_error(label: str, error: Exception):

    if isinstance(error, HTTPException):

        raise error



    raise HTTPException(

        status_code=400,

        detail=f"{label} upload failed: {error}",

    )





def _read_uploaded_orders(label: str, upload_file: UploadFile, reader):

    file_path = _save_upload(upload_file)



    try:

        return reader(file_path)

    except Exception as e:

        raise HTTPException(

            status_code=400,

            detail=f"Could not read {label} file: {e}",

        )





def save_daily_report_rows(

    db: Session,

    report_date: date,

    platform: str,

    report_rows

):

    for row in report_rows:

        style = str(row.get("style", "")).strip()

        color = str(row.get("color", "")).strip()

        size = str(row.get("size", "")).strip()

        qty = _safe_int(row.get("total_order_qty", 0))



        if qty <= 0:

            continue



        existing = db.query(DailyReport).filter(

            DailyReport.report_date == report_date,

            DailyReport.platform == platform,

            DailyReport.style == style,

            DailyReport.color == color,

            DailyReport.size == size,

        ).first()



        if existing:

            existing.total_order_qty = (

                _safe_int(existing.total_order_qty) + qty

            )

            continue



        db.add(

            DailyReport(

                report_date=report_date,

                style=style,

                color=color,

                size=size,

                total_order_qty=qty,

                platform=platform

            )

        )





def _extract_pack_of(sku: str):

    match = re.search(

        r"(?<![A-Z0-9])(\d+)\s*(?:P|PC|PCS)(?![A-Z0-9])",

        str(sku or ""),

        flags=re.IGNORECASE,

    )



    if match:

        return f"{match.group(1)}p"



    return "1p"





def _expand_sales_analytics_details(

    orders,

    platform: str,

    report_date: date,

    db: Session,

):

    sku_master_list = db.query(SKUMaster).all()

    details = []



    for order in orders:

        raw_sku = str(order.get("sku", "")).strip()

        normalized_sku = normalize_sku(raw_sku)



        if not normalized_sku:

            continue



        sku_master = None

        for item in sku_master_list:

            if normalize_sku(item.sku) == normalized_sku:

                sku_master = item

                break



        if not sku_master:

            continue



        order_qty = _safe_int(order.get("quantity", 0))

        if order_qty <= 0:

            continue



        pack_of = sku_master.pack_of or _extract_pack_of(raw_sku)
        
        # fallback to "1p" if it's completely empty
        if not pack_of:
            pack_of = "1p"

        order_id = str(order.get("order_id", "")).strip()

        invoice_amount = _safe_float(order.get("price", 0))



        for piece in sku_master.pieces:

            color = clean_color_name(piece.color or "")

            piece_qty = _safe_int(piece.qty)



            if not color or piece_qty <= 0:

                continue



            details.append(

                SalesAnalyticsDetail(

                    report_date=report_date,

                    platform=platform,

                    order_id=order_id,

                    sku=raw_sku,

                    main_product_type=(

                        sku_master.main_product_type

                        or sku_master.style

                        or "Unknown"

                    ).strip(),

                    style=(sku_master.style or "").strip(),

                    color=color,

                    size=str(sku_master.size or "").strip().upper(),

                    pack_of=pack_of,

                    order_qty=order_qty,

                    piece_qty=order_qty * piece_qty,

                    invoice_amount=invoice_amount,

                )

            )



    return details





def save_sales_analytics_details(

    db: Session,

    report_date: date,

    platform: str,

    orders,

):

    db.query(SalesAnalyticsDetail).filter(

        SalesAnalyticsDetail.report_date == report_date,

        SalesAnalyticsDetail.platform == platform,

    ).delete(synchronize_session=False)



    details = _expand_sales_analytics_details(

        orders,

        platform,

        report_date,

        db,

    )



    for detail in details:

        db.add(detail)



    return len(details)





def _safe_int(value):

    try:

        if pd.isna(value):

            return 0

        return int(value)

    except (TypeError, ValueError):

        return 0





def _safe_float(value):

    try:

        if pd.isna(value):

            return 0

        return float(value)

    except (TypeError, ValueError):

        return 0





def _safe_float(value):

    try:

        if pd.isna(value):

            return 0.0

        return float(value)

    except (TypeError, ValueError):

        return 0.0





def _unique_order_count(orders):

    unique_order_ids = set()

    fallback_order_count = 0



    for row in orders:

        order_id = str(

            row.get("order_id", "")

        ).strip()



        if order_id:

            unique_order_ids.add(order_id)

        else:

            fallback_order_count += 1



    return len(unique_order_ids) + fallback_order_count





def _clean_packing_item_type(item_type):

    clean_type = str(item_type or "").strip().lower()

    if clean_type not in PACKING_INVENTORY_TYPES:

        raise HTTPException(

            status_code=400,

            detail="Packing inventory type is not valid.",

        )

    return clean_type





def _clean_packing_platform(item_type, platform):

    if item_type != "shipping_cover":

        return None



    clean_platform = str(platform or "").strip()

    if clean_platform not in PLATFORM_NAMES:

        raise HTTPException(

            status_code=400,

            detail="Shipping cover platform is not valid.",

        )

    return clean_platform





def seed_packing_inventory_if_empty(db):

    has_rows = db.query(PackingInventory.id).first()

    if has_rows:

        return



    db.add_all([

        PackingInventory(

            item_type="shipping_label",

            platform=None,

            name="Shipping Label",

            qty=0,

        ),

        *[

            PackingInventory(

                item_type="shipping_cover",

                platform=platform,

                name="Shipping Cover",

                qty=0,

            )

            for platform in PLATFORM_NAMES

        ],

        PackingInventory(

            item_type="packing_cover",

            platform=None,

            name="Packing Cover",

            qty=0,

        ),

        PackingInventory(

            item_type="packing_board",

            platform=None,

            name="Packing Board",

            qty=0,

        ),

    ])

    db.flush()





def deduct_packing_inventory_for_platform(db, platform, order_count, piece_count):

    clean_platform = str(platform or "").strip()

    order_qty = _safe_int(order_count)

    piece_qty = _safe_int(piece_count)



    if clean_platform not in PLATFORM_NAMES or (order_qty <= 0 and piece_qty <= 0):

        return {

            "lines_updated": 0,

            "total_qty_deducted": 0,

            "deductions": [],

        }



    seed_packing_inventory_if_empty(db)



    rows = db.query(PackingInventory).filter(

        or_(

            PackingInventory.item_type.in_([

                "shipping_label",

                "packing_cover",

                "packing_board",

            ]),

            and_(

                PackingInventory.item_type == "shipping_cover",

                PackingInventory.platform == clean_platform,

            ),

        )

    ).all()



    lines_updated = 0

    total_qty_deducted = 0

    deductions = []



    for row in rows:

        if row.item_type == "shipping_label":

            requested_qty = order_qty * 2 if clean_platform == "Myntra" else order_qty

        elif row.item_type == "shipping_cover":

            requested_qty = order_qty

        elif row.item_type in {"packing_cover", "packing_board"}:

            requested_qty = piece_qty

        else:

            requested_qty = 0



        if requested_qty <= 0:

            continue



        previous_qty = _safe_int(row.qty)
        deducted_qty = min(previous_qty, requested_qty)
        row.qty = max(previous_qty - requested_qty, 0)

        if previous_qty != row.qty:
            from datetime import date
            today = date.today()
            usage_record = db.query(PackingInventoryUsage).filter_by(
                usage_date=today,
                item_type=row.item_type,
                platform=row.platform,
                name=row.name
            ).first()
            
            if not usage_record:
                usage_record = PackingInventoryUsage(
                    usage_date=today,
                    item_type=row.item_type,
                    platform=row.platform,
                    name=row.name,
                    used_qty=0,
                    restocked_qty=0
                )
                db.add(usage_record)
            
            usage_record.used_qty += (previous_qty - row.qty)

        lines_updated += 1

        total_qty_deducted += deducted_qty

        deductions.append({

            "style": row.item_type,

            "color": row.name,

            "size": row.platform,

            "requested_qty": requested_qty,

            "deducted_qty": deducted_qty,

            "remaining_qty": row.qty,

        })



    return {

        "lines_updated": lines_updated,

        "total_qty_deducted": total_qty_deducted,

        "deductions": deductions,

    }





def _log_inventory_deductions(

    db: Session,

    report_date: date,

    platform: str,

    inventory_type: str,

    deductions,

):

    for deduction in deductions or []:

        deducted_qty = _safe_int(

            deduction.get("deducted_qty", 0)

        )



        if deducted_qty <= 0:

            continue



        raw_size = deduction.get("size")



        db.add(

            InventoryDeductionLog(

                report_date=report_date,

                platform=platform,

                inventory_type=inventory_type,

                style=str(deduction.get("style", "")).strip(),

                color=str(deduction.get("color", "")).strip(),

                size=(

                    str(raw_size).strip()

                    if raw_size is not None and str(raw_size).strip()

                    else None

                ),

                qty=deducted_qty,

            )

        )





def _restore_inventory_deduction_log(

    db: Session,

    report_date: date,

    platform: str = None,

):

    query = db.query(InventoryDeductionLog).filter(

        InventoryDeductionLog.report_date == report_date

    )



    if platform == "All":

        query = query.filter(

            InventoryDeductionLog.platform.in_(PLATFORM_NAMES)

        )

    elif platform:

        query = query.filter(

            InventoryDeductionLog.platform == platform

        )



    logs = query.all()



    for log in logs:

        qty = _safe_int(log.qty)



        if qty <= 0:

            continue



        if log.inventory_type == "stock":

            row = _find_stock_inventory_row(

                db,

                log.style,

                log.color,

                log.size,

            )



            if not row:

                row = StockInventory(

                    style=log.style,

                    color=log.color,

                    size=log.size,

                    qty=0,

                )

                db.add(row)

                db.flush()



            row.qty = _safe_int(row.qty) + qty



        elif log.inventory_type == "return":

            row = db.query(ReturnInventory).filter(

                ReturnInventory.style == log.style,

                ReturnInventory.color == log.color,

                ReturnInventory.size == log.size,

            ).first()



            if not row:

                row = ReturnInventory(

                    style=log.style,

                    color=log.color,

                    size=log.size,

                    qty=0,

                )

                db.add(row)

                db.flush()



            row.qty = _safe_int(row.qty) + qty



        elif log.inventory_type == "sticker":

            row = _find_sticker_inventory_row(

                db,

                log.style,

                log.color,

            )



            if not row:

                row = StickerInventory(

                    style=log.style,

                    color=log.color,

                    qty=0,

                )

                db.add(row)

                db.flush()



            row.qty = _safe_int(row.qty) + qty



        elif log.inventory_type == "packing":

            item_type = str(log.style or "").strip()

            item_name = str(log.color or "").strip()

            item_platform = str(log.size or "").strip() or None



            row = db.query(PackingInventory).filter(

                PackingInventory.item_type == item_type,

                PackingInventory.name == item_name,

                PackingInventory.platform == item_platform,

            ).first()



            if not row:
                row = PackingInventory(
                    item_type=item_type,
                    platform=item_platform,
                    name=item_name,
                    qty=0,
                )
                db.add(row)
                db.flush()

            previous_qty = _safe_int(row.qty)
            row.qty = previous_qty + qty
            
            if previous_qty != row.qty:
                from datetime import date
                today = date.today()
                usage_record = db.query(PackingInventoryUsage).filter_by(
                    usage_date=today,
                    item_type=item_type,
                    platform=item_platform,
                    name=item_name
                ).first()
                if not usage_record:
                    usage_record = PackingInventoryUsage(
                        usage_date=today,
                        item_type=item_type,
                        platform=item_platform,
                        name=item_name,
                        used_qty=0,
                        restocked_qty=0
                    )
                    db.add(usage_record)
                usage_record.restocked_qty += (row.qty - previous_qty)

    restored_count = len(logs)

    restored_qty = sum(

        _safe_int(log.qty)

        for log in logs

    )



    query.delete(synchronize_session=False)



    return {

        "restored_logs": restored_count,

        "restored_qty": restored_qty,

    }





def _delete_saved_upload_files(upload_rows):

    deleted_files = 0



    for upload in upload_rows:

        filename = os.path.basename(

            str(upload.file_name or "").strip()

        )



        if not filename:

            continue



        file_path = os.path.abspath(

            os.path.join(

                UPLOAD_FOLDER,

                filename,

            )

        )

        uploads_root = os.path.abspath(UPLOAD_FOLDER)



        if (

            os.path.commonpath([uploads_root, file_path]) != uploads_root

            or not os.path.isfile(file_path)

        ):

            continue



        os.remove(file_path)

        deleted_files += 1



    return deleted_files





def _rebuild_all_daily_report_rows(

    db: Session,

    report_date: date,

):

    db.query(DailyReport).filter(

        DailyReport.report_date == report_date,

        DailyReport.platform == "All",

    ).delete(synchronize_session=False)



    rows = db.query(DailyReport).filter(

        DailyReport.report_date == report_date,

        DailyReport.platform != "All",

    ).all()



    grouped_rows = {}



    for row in rows:

        key = (

            row.style,

            row.color,

            row.size,

        )



        if key not in grouped_rows:

            grouped_rows[key] = 0



        grouped_rows[key] += _safe_int(row.total_order_qty)



    for (style, color, size), qty in grouped_rows.items():

        if qty <= 0:

            continue



        db.add(

            DailyReport(

                report_date=report_date,

                platform="All",

                style=style,

                color=color,

                size=size,

                total_order_qty=qty,

            )

        )





def save_daily_sales_summary(

    db: Session,

    report_date: date,

    platform: str,

    orders,

    mapped_piece_qty: int

):

    total_invoice_amount = 0



    for row in orders:

        total_invoice_amount += _safe_float(

            row.get("price", 0)

        )



    new_order_count = _unique_order_count(orders)



    existing = db.query(DailySalesReport).filter(

        DailySalesReport.report_date == report_date,

        DailySalesReport.platform == platform

    ).first()



    if existing:

        existing.total_orders = (

            _safe_int(existing.total_orders)

            + new_order_count

        )

        existing.total_piece_qty = (

            _safe_int(existing.total_piece_qty)

            + mapped_piece_qty

        )

        existing.total_invoice_amount = round(

            float(existing.total_invoice_amount or 0)

            + total_invoice_amount,

            2

        )

        return



    db.add(

        DailySalesReport(

            report_date=report_date,

            platform=platform,

            total_orders=new_order_count,

            total_piece_qty=mapped_piece_qty,

            total_invoice_amount=round(

                total_invoice_amount,

                2

            )

        )

    )





def _saved_upload_sha256(upload_file: UploadFile):

    file_path = getattr(

        upload_file,

        "saved_path",

        None

    )



    if not file_path:

        file_path = os.path.join(

            UPLOAD_FOLDER,

            _clean_upload_filename(

                upload_file.filename

            )

        )



    digest = hashlib.sha256()



    with open(file_path, "rb") as saved_file:

        while chunk := saved_file.read(1024 * 1024):

            digest.update(chunk)



    return digest.hexdigest()





def _upload_already_counted(

    db: Session,

    platform: str,

    upload_file: UploadFile,

    report_date: date = None,

):

    if not upload_file:

        return True



    file_hash = _saved_upload_sha256(upload_file)



    query = (

        db.query(SalesUpload)

        .filter(

            SalesUpload.platform == platform,

            SalesUpload.file_hash == file_hash,

        )

    )



    uploads = query.all()

    for upload in uploads:

        has_sales_report = (

            db.query(DailySalesReport)

            .filter(

                DailySalesReport.report_date == upload.report_date,

                DailySalesReport.platform == platform,

            )

            .first()

            is not None

        )

        has_daily_report = (

            db.query(DailyReport)

            .filter(

                DailyReport.report_date == upload.report_date,

                DailyReport.platform == platform,

            )

            .first()

            is not None

        )



        if has_sales_report or has_daily_report:

            return True



        db.delete(upload)



    return False





def _count_sales_upload_once(

    db: Session,

    report_date: date,

    platform: str,

    upload_file: UploadFile

):

    if _upload_already_counted(

        db,

        platform,

        upload_file,

        report_date,

    ):

        return False



    file_hash = _saved_upload_sha256(

        upload_file

    )



    db.add(

        SalesUpload(

            report_date=report_date,

            platform=platform,

            file_name=getattr(

                upload_file,

                "saved_filename",

                upload_file.filename

            ),

            file_hash=file_hash

        )

    )



    return True





def _orders_from_new_platform_uploads(

    db: Session,

    report_date: date,

    platform_orders,

    platform_files,

    flipkart_dispatch_period: Optional[str] = None,

):

    new_orders = []

    counted_platforms = []

    skipped_platforms = []



    for platform_name, orders in platform_orders.items():

        upload_file = platform_files.get(platform_name)

        platform_report_date = _platform_report_date(

            platform_name,

            report_date,

            flipkart_dispatch_period,

        )



        if not upload_file:

            continue



        if _upload_already_counted(

            db,

            platform_name,

            upload_file,

            platform_report_date,

        ):

            skipped_platforms.append(platform_name)

            continue



        new_orders.extend(orders)

        counted_platforms.append(platform_name)



    return new_orders, counted_platforms, skipped_platforms





def _save_new_platform_sales(

    db: Session,

    report_date: date,

    platform_orders,

    platform_files,

    flipkart_dispatch_period: Optional[str] = None,

):

    counted_platforms = []

    skipped_platforms = []



    for platform_name, orders in platform_orders.items():

        upload_file = platform_files.get(platform_name)

        platform_report_date = _platform_report_date(

            platform_name,

            report_date,

            flipkart_dispatch_period,

        )



        if not upload_file or not _count_sales_upload_once(

            db,

            platform_report_date,

            platform_name,

            upload_file

        ):

            skipped_platforms.append(platform_name)

            continue



        platform_aggregated = aggregate_orders(orders)

        platform_expanded = expand_inventory(

            platform_aggregated,

            db

        )

        if platform_name in PLATFORM_NAMES:

            platform_report = generate_daily_report(

                platform_expanded,

                db

            )

            save_daily_report_rows(

                db,

                platform_report_date,

                platform_name,

                platform_report

            )

        save_daily_sales_summary(

            db,

            platform_report_date,

            platform_name,

            orders,

            sum(

                _safe_int(item.get("qty", 0))

                for item in platform_expanded

            )

        )

        save_sales_analytics_details(

            db,

            platform_report_date,

            platform_name,

            orders,

        )

        counted_platforms.append(platform_name)



    return counted_platforms, skipped_platforms





def _platform_report_date(

    platform_name: str,

    default_report_date: date,

    flipkart_dispatch_period: Optional[str] = None,

) -> date:

    if platform_name == "Flipkart":

        return get_flipkart_target_date(flipkart_dispatch_period)

    return default_report_date





def _orders_for_report_date(

    platform_orders: dict,

    report_date: date,

    flipkart_dispatch_period: Optional[str] = None,

) -> list:

    orders = []



    for platform_name, platform_order_list in platform_orders.items():

        if platform_name not in PLATFORM_NAMES:

            continue



        if (

            _platform_report_date(

                platform_name,

                report_date,

                flipkart_dispatch_period,

            )

            != report_date

        ):

            continue

        orders.extend(platform_order_list)



    return orders





def _unknown_platform_skus(db: Session, platform_orders: dict) -> List[dict]:

    master_skus = {

        normalize_sku(row.sku)

        for row in db.query(SKUMaster.sku).all()

        if row.sku

    }



    unknown_by_key = {}



    for platform_name, orders in platform_orders.items():

        for order in orders:

            raw_sku = str(order.get("sku", "")).strip()



            if not raw_sku:

                continue



            normalized = normalize_sku(raw_sku)



            if normalized in master_skus:

                continue



            key = normalized



            if key not in unknown_by_key:

                unknown_by_key[key] = {

                    "platform": platform_name,

                    "platforms": [platform_name],

                    "sku": raw_sku,

                    "normalized_sku": normalized,

                    "quantity": 0,

                }

            elif platform_name not in unknown_by_key[key]["platforms"]:

                unknown_by_key[key]["platforms"].append(platform_name)

                unknown_by_key[key]["platform"] = ", ".join(

                    unknown_by_key[key]["platforms"]

                )



            unknown_by_key[key]["quantity"] += _safe_int(

                order.get("quantity", 0)

            )



    return list(unknown_by_key.values())





def _raise_if_unknown_platform_skus(

    db: Session,

    platform_orders: dict,

):

    unknown_skus = _unknown_platform_skus(

        db,

        platform_orders,

    )



    if unknown_skus:

        raise HTTPException(

            status_code=409,

            detail={

                "code": "UNKNOWN_SKUS",

                "message": (

                    "Some platform SKUs are not available in the SKU master."

                ),

                "skus": unknown_skus,

            },

        )





def _daily_report_query(db: Session, report_date, platform):

    query = db.query(DailyReport)



    if not report_date:

        if platform == "All":

            query = query.filter(

                DailyReport.platform != "All"

            )

        elif platform:

            query = query.filter(

                DailyReport.platform == platform

            )

        return query



    if platform == "All" or not platform:

        conditions = [

            and_(

                DailyReport.platform == platform_name,

                DailyReport.report_date == report_date,

            )

            for platform_name in PLATFORM_NAMES

        ]

        return query.filter(or_(*conditions))



    return query.filter(

        DailyReport.platform == platform,

        DailyReport.report_date == report_date,

    )





def _flipkart_warehouse_daily_report_rows(

    db: Session,

    report_date: date = None,

    from_date: date = None,

    to_date: date = None,

):

    query = db.query(SalesAnalyticsDetail).filter(

        SalesAnalyticsDetail.platform == "Flipkart Warehouse",

    )



    if report_date:

        query = query.filter(

            SalesAnalyticsDetail.report_date == report_date,

        )

    elif from_date and to_date:

        query = query.filter(

            SalesAnalyticsDetail.report_date >= from_date,

            SalesAnalyticsDetail.report_date <= to_date,

        )



    grouped = {}



    for row in query.all():

        key = (

            row.report_date,

            row.style or "Unknown",

            row.color or "Unknown",

            str(row.size or "").strip().upper(),

        )

        grouped[key] = grouped.get(key, 0) + _safe_int(row.piece_qty)



    return [

        {

            "date": str(report_date_key),

            "style": style,

            "color": color,

            "size": size,

            "total_order_qty": qty,

            "platform": "Flipkart Warehouse",

        }

        for (

            report_date_key,

            style,

            color,

            size,

        ), qty in grouped.items()

        if qty > 0

    ]





@router.post("/flipkart-warehouse-report")

def generate_flipkart_warehouse_report(

    file: UploadFile = File(...),

):

    file_path = _save_upload(file)



    try:

        orders = filter_flipkart_warehouse_orders(file_path)

    except ValueError as e:

        raise HTTPException(status_code=400, detail=str(e))

    except Exception as e:

        raise HTTPException(

            status_code=400,

            detail=f"Could not read Flipkart Warehouse file: {e}",

        )



    if not orders:

        raise HTTPException(

            status_code=400,

            detail="No valid Flipkart Warehouse orders found in this file.",

        )



    db: Session = SessionLocal()

    latest_order_date = max(

        order["order_date"]

        for order in orders

        if order.get("order_date")

    )

    report_date = latest_order_date + timedelta(days=1)



    try:

        platform_orders = {

            "Flipkart Warehouse": orders,

        }

        _raise_if_unknown_platform_skus(

            db,

            platform_orders,

        )



        counted_platforms, skipped_platforms = _save_new_platform_sales(

            db,

            report_date,

            platform_orders,

            {

                "Flipkart Warehouse": file,

            },

        )



        if not counted_platforms:

            db.commit()

            return {

                "message": (

                    "This Flipkart Warehouse file was already reported. "

                    "Duplicate files are skipped."

                ),

                "report_date": str(report_date),

                "total_orders": 0,

                "total_piece_qty": 0,

                "total_invoice_amount": 0,

                "counted_platforms": counted_platforms,

                "skipped_duplicate_platforms": skipped_platforms,

            }



        platform_expanded = expand_inventory(

            aggregate_orders(orders),

            db,

        )

        total_piece_qty = sum(

            _safe_int(item.get("qty", 0))

            for item in platform_expanded

        )

        total_invoice_amount = round(

            sum(_safe_float(order.get("price", 0)) for order in orders),

            2,

        )



        db.commit()
        write_inventory_snapshot("return", _serialize_return_inventory_rows(db))

        return {

            "message": "Flipkart Warehouse report generated successfully.",

            "report_date": str(report_date),

            "total_orders": _unique_order_count(orders),

            "total_piece_qty": total_piece_qty,

            "total_invoice_amount": total_invoice_amount,

            "counted_platforms": counted_platforms,

            "skipped_duplicate_platforms": skipped_platforms,

        }

    except Exception:

        db.rollback()

        raise

    finally:

        db.close()





@router.get("/flipkart-warehouse-report")

def list_flipkart_warehouse_report(

    search: str = Query(None),

):

    db: Session = SessionLocal()



    try:

        rows = (

            db.query(SalesAnalyticsDetail)

            .filter(

                SalesAnalyticsDetail.platform == "Flipkart Warehouse",

            )

            .order_by(

                SalesAnalyticsDetail.report_date.desc(),

                SalesAnalyticsDetail.order_id.asc(),

                SalesAnalyticsDetail.sku.asc(),

                SalesAnalyticsDetail.style.asc(),

                SalesAnalyticsDetail.color.asc(),

            )

            .all()

        )



        search_text = str(search or "").strip().lower()

        items = []



        for row in rows:

            item = {

                "date": str(row.report_date),

                "order_id": row.order_id or "",

                "sku": row.sku or "",

                "main_product_type": row.main_product_type or "",

                "style": row.style or "",

                "color": row.color or "",

                "size": row.size or "",

                "pack_of": row.pack_of or "",

                "order_qty": _safe_int(row.order_qty),

                "piece_qty": _safe_int(row.piece_qty),

                "invoice_amount": round(

                    _safe_float(row.invoice_amount),

                    2,

                ),

            }



            if search_text:

                haystack = " ".join(

                    str(value).lower()

                    for value in item.values()

                )

                if search_text not in haystack:

                    continue



            items.append(item)



        totals_by_order = {}

        total_piece_qty = 0

        total_invoice_amount = 0.0



        for item in items:

            total_piece_qty += _safe_int(item["piece_qty"])

            key = (

                item["date"],

                item["order_id"],

                item["sku"],

            )

            totals_by_order.setdefault(

                key,

                _safe_float(item["invoice_amount"]),

            )



        total_invoice_amount = round(

            sum(totals_by_order.values()),

            2,

        )



        return {

            "count": len(items),

            "total_orders": len({

                (item["date"], item["order_id"])

                for item in items

            }),

            "total_piece_qty": total_piece_qty,

            "total_invoice_amount": total_invoice_amount,

            "items": items,

        }

    finally:

        db.close()





@router.get("/flipkart-warehouse-report/export-excel")

def export_flipkart_warehouse_report_excel(

    report_date: str = Query(None),

):

    from io import BytesIO

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter



    db: Session = SessionLocal()



    try:

        parsed_date = None

        if report_date:

            parsed_date = _parse_report_date(report_date)



        query = db.query(SalesAnalyticsDetail).filter(

            SalesAnalyticsDetail.platform == "Flipkart Warehouse",

        )



        if parsed_date:

            query = query.filter(

                SalesAnalyticsDetail.report_date == parsed_date,

            )



        rows = (

            query

            .order_by(

                SalesAnalyticsDetail.report_date.desc(),

                SalesAnalyticsDetail.style.asc(),

                SalesAnalyticsDetail.color.asc(),

                SalesAnalyticsDetail.size.asc(),

            )

            .all()

        )



        summary_rows, size_columns = _pivot_daily_report_rows(rows)



        wb = Workbook()

        ws = wb.active

        ws.title = "Warehouse Report"



        header_fill = PatternFill(

            "solid",

            fgColor="0E7490",

        )

        header_font = Font(

            bold=True,

            color="FFFFFF",

        )

        date_label = str(parsed_date) if parsed_date else "All dates"



        ws.append(["Flipkart Warehouse report"])

        ws.merge_cells(

            start_row=1,

            start_column=1,

            end_row=1,

            end_column=4 + len(size_columns),

        )

        ws["A1"].font = Font(bold=True, size=14)

        ws.append(["Date", date_label])

        ws.append(["Type", "Sales report only"])

        ws.append([])



        headers = [

            "Date",

            "Style",

            "Color",

            *size_columns,

            "Total",

        ]

        ws.append(headers)



        for cell in ws[ws.max_row]:

            cell.font = header_font

            cell.fill = header_fill

            cell.alignment = Alignment(horizontal="center")



        for item in summary_rows:

            ws.append([

                item["date"],

                item["style"],

                item["color"],

                *[

                    item["sizes"].get(size, "")

                    for size in size_columns

                ],

                item["total"],

            ])



        for column_cells in ws.columns:

            column_letter = get_column_letter(

                column_cells[0].column,

            )

            max_length = 0

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(

                        max_length,

                        len(str(cell.value)),

                    )

            ws.column_dimensions[column_letter].width = min(

                max(max_length + 2, 10),

                40,

            )



        buffer = BytesIO()

        wb.save(buffer)

        buffer.seek(0)



        date_part = (

            str(parsed_date) if parsed_date

            else datetime.now().strftime("%d-%m-%Y")

        )

        filename = f"flipkart_warehouse_report_{date_part}.xlsx"



        return StreamingResponse(

            buffer,

            media_type=(

                "application/vnd.openxmlformats-officedocument"

                ".spreadsheetml.sheet"

            ),

            headers={

                "Content-Disposition": (

                    f'attachment; filename="{filename}"'

                )

            },

        )

    finally:

        db.close()





@router.delete("/flipkart-warehouse-report")

def delete_flipkart_warehouse_report(

    report_date: str = Query(...),

    password: str = Query(""),

):

    if password != "Admin":

        raise HTTPException(

            status_code=403,

            detail="Invalid password",

        )



    db: Session = SessionLocal()



    try:

        parsed_date = _parse_report_date(report_date)



        sales_query = db.query(DailySalesReport).filter(

            DailySalesReport.report_date == parsed_date,

            DailySalesReport.platform == "Flipkart Warehouse",

        )

        deleted_sales = sales_query.delete(

            synchronize_session=False

        )



        upload_query = db.query(SalesUpload).filter(

            SalesUpload.report_date == parsed_date,

            SalesUpload.platform == "Flipkart Warehouse",

        )

        upload_rows = upload_query.all()

        deleted_upload_files = _delete_saved_upload_files(upload_rows)

        deleted_uploads = upload_query.delete(

            synchronize_session=False

        )



        analytics_query = db.query(SalesAnalyticsDetail).filter(

            SalesAnalyticsDetail.report_date == parsed_date,

            SalesAnalyticsDetail.platform == "Flipkart Warehouse",

        )

        deleted_analytics_rows = analytics_query.delete(

            synchronize_session=False

        )



        daily_report_query = db.query(DailyReport).filter(

            DailyReport.report_date == parsed_date,

            DailyReport.platform == "Flipkart Warehouse",

        )

        deleted_daily_rows = daily_report_query.delete(

            synchronize_session=False

        )



        db.commit()
        write_inventory_snapshot("return", _serialize_return_inventory_rows(db))



        return {

            "message": "Flipkart Warehouse report deleted successfully",

            "report_date": str(parsed_date),

            "deleted_sales_rows": deleted_sales,

            "deleted_upload_markers": deleted_uploads,

            "deleted_upload_files": deleted_upload_files,

            "deleted_analytics_rows": deleted_analytics_rows,

            "deleted_daily_rows": deleted_daily_rows,

        }

    except Exception:

        db.rollback()

        raise

    finally:

        db.close()





def _sales_row_for_platform(

    db: Session,

    platform_name: str,

    view_date: date,

):

    return (

        db.query(DailySalesReport)

        .filter(

            DailySalesReport.platform == platform_name,

            DailySalesReport.report_date == view_date,

        )

        .first()

    )





# =====================================

# DAILY REPORT QUERY

# =====================================



@router.get("/daily-report")

def get_daily_report(

    report_date: str = Query(None),

    platform: str = Query(None)

):

    db: Session = SessionLocal()



    try:

        parsed_date = None

        if report_date:

            parsed_date = _parse_report_date(report_date)



        rows = (

            _daily_report_query(

                db,

                parsed_date,

                platform or "All",

            )

            .order_by(

                DailyReport.report_date.desc(),

                DailyReport.platform.asc(),

                DailyReport.style.asc(),

                DailyReport.color.asc(),

                DailyReport.size.asc()

            )

            .all()

        )



        response_rows = [

            {

                "date": str(row.report_date),

                "style": row.style,

                "color": row.color,

                "size": row.size,

                "total_order_qty": row.total_order_qty,

                "platform": row.platform

            }

            for row in rows

        ]



        if (platform or "All") in ("All", "Flipkart Warehouse"):

            response_rows.extend(

                _flipkart_warehouse_daily_report_rows(

                    db,

                    parsed_date,

                )

            )

            response_rows.sort(

                key=lambda row: (

                    row["date"],

                    row["platform"],

                    row["style"],

                    row["color"],

                    row["size"],

                ),

                reverse=True,

            )



        return {

            "count": len(response_rows),

            "rows": response_rows

        }

    finally:

        db.close()





@router.get("/daily-report/multi-qty-orders")

def get_daily_report_multi_qty_orders(

    report_date: str = Query(...),

    platform: str = Query("All"),

):

    db: Session = SessionLocal()



    try:

        parsed_date = _parse_report_date(report_date)

        query = db.query(SalesAnalyticsDetail).filter(

            SalesAnalyticsDetail.report_date == parsed_date,

            SalesAnalyticsDetail.order_qty > 1,

        )



        if platform == "All":

            query = query.filter(

                SalesAnalyticsDetail.platform.in_(REPORTING_PLATFORM_NAMES)

            )

        elif platform:

            query = query.filter(

                SalesAnalyticsDetail.platform == platform

            )



        rows = (

            query.order_by(

                SalesAnalyticsDetail.platform.asc(),

                SalesAnalyticsDetail.order_id.asc(),

                SalesAnalyticsDetail.sku.asc(),

            )

            .all()

        )

        unique_rows = {}



        for row in rows:

            key = (

                row.platform or "",

                row.order_id or "",

                row.sku or "",

            )



            if key not in unique_rows:

                unique_rows[key] = {

                    "platform": row.platform or "",

                    "order_id": row.order_id or "",

                    "sku": row.sku or "",

                    "qty": _safe_int(row.order_qty),

                }



        items = sorted(

            unique_rows.values(),

            key=lambda item: (

                item["platform"],

                item["order_id"],

                item["sku"],

            ),

        )



        return {

            "date": str(parsed_date),

            "count": len(items),

            "total_qty": sum(item["qty"] for item in items),

            "items": items,

        }

    finally:

        db.close()





def _parse_report_date(report_date: str):

    for date_format in ("%Y-%m-%d", "%d-%m-%Y"):

        try:

            return datetime.strptime(

                report_date,

                date_format

            ).date()

        except ValueError:

            continue



    raise HTTPException(

        status_code=400,

        detail="Invalid date format. Use YYYY-MM-DD or DD-MM-YYYY."

    )





@router.get("/daily-report/flipkart-zone-summary")

def get_daily_report_flipkart_zone_summary(

    report_date: str = Query(...),

):

    db: Session = SessionLocal()



    try:

        parsed_date = _parse_report_date(report_date)

        return _flipkart_zone_summary_rows(

            db,

            parsed_date,

        )

    finally:

        db.close()





@router.get("/daily-report/flipkart-zone-summary/batches/{batch_id}")
def get_daily_report_flipkart_zone_batch(
    batch_id: int,
):
    db: Session = SessionLocal()

    try:
        batch = (
            db.query(FlipkartZoneBatch)
            .filter(FlipkartZoneBatch.id == batch_id)
            .first()
        )

        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")

        items = (
            db.query(FlipkartZoneBatchItem)
            .filter(FlipkartZoneBatchItem.batch_id == batch_id)
            .all()
        )

        return {
            "id": batch.id,
            "report_date": str(batch.report_date),
            "source_filename": batch.source_filename or "",
            "label_count": batch.label_count,
            "platform": getattr(batch, "platform", "Flipkart") or "Flipkart",
            "created_at": batch.created_at.isoformat() if batch.created_at else "",
            "items": [
                {
                    "zone": item.zone,
                    "label_count": item.label_count,
                    "platform": getattr(item, "platform", "Flipkart") or "Flipkart",
                }
                for item in items
            ],
            "total": sum(item.label_count for item in items),
        }
    finally:
        db.close()

@router.delete("/daily-report/flipkart-zone-summary/batches/{batch_id}")
def delete_daily_report_flipkart_zone_batch(
    batch_id: int,
):
    db: Session = SessionLocal()

    try:
        batch = db.query(FlipkartZoneBatch).filter(FlipkartZoneBatch.id == batch_id).first()
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")
        
        db.query(FlipkartZoneBatchItem).filter(FlipkartZoneBatchItem.batch_id == batch_id).delete()
        db.delete(batch)
        db.flush()
        _sync_flipkart_zone_totals(db, batch.report_date)
        db.commit()
        return {"success": True}
    finally:
        db.close()





@router.get("/daily-report/export")

def export_daily_report(

    report_date: str = Query(None),

    platform: str = Query(None),

):

    import csv

    from io import StringIO



    db: Session = SessionLocal()



    try:

        parsed_date = None

        if report_date:

            parsed_date = _parse_report_date(report_date)



        rows = (

            _daily_report_query(db, parsed_date, platform)

            .order_by(

                DailyReport.platform.asc(),

                DailyReport.style.asc(),

                DailyReport.color.asc(),

                DailyReport.size.asc(),

            )

            .all()

        )



        buffer = StringIO()

        writer = csv.writer(buffer)

        writer.writerow([

            "date",

            "platform",

            "style",

            "color",

            "size",

            "total_order_qty",

        ])

        for row in rows:

            writer.writerow([

                str(row.report_date),

                row.platform,

                row.style,

                row.color,

                row.size,

                row.total_order_qty,

            ])



        buffer.seek(0)

        date_part = (

            str(parsed_date) if parsed_date

            else datetime.now().strftime("%d-%m-%Y")

        )

        plat_part = platform or "all"

        filename = f"daily_report_{date_part}_{plat_part}.csv"



        return StreamingResponse(

            iter([buffer.getvalue()]),

            media_type="text/csv",

            headers={

                "Content-Disposition": (

                    f'attachment; filename="{filename}"'

                )

            },

        )

    finally:

        db.close()





DAILY_REPORT_SIZE_ORDER = [

    "XS", "S", "M", "L", "XL", "2XL",

]





def _pivot_daily_report_rows(rows):

    grouped = {}

    sizes = set()



    for row in rows:

        size = str(row.size or "").upper().strip()

        if size:

            sizes.add(size)



        key = (

            row.platform,

            row.style,

            row.color,

        )



        if key not in grouped:

            grouped[key] = {

                "date": str(row.report_date),

                "platform": row.platform,

                "style": row.style,

                "color": row.color,

                "sizes": {},

                "total": 0,

            }



        qty = int(row.total_order_qty or 0)

        grouped[key]["sizes"][size] = (

            grouped[key]["sizes"].get(size, 0) + qty

        )

        grouped[key]["total"] += qty



    size_columns = [

        size

        for size in DAILY_REPORT_SIZE_ORDER

        if size in sizes

    ]

    size_columns.extend(

        sorted(

            size

            for size in sizes

            if size not in DAILY_REPORT_SIZE_ORDER

        )

    )



    summary = sorted(

        grouped.values(),

        key=lambda item: (

            item["platform"],

            item["style"],

            item["color"],

        ),

    )



    return summary, size_columns





@router.get("/daily-report/export-excel")

def export_daily_report_excel(

    report_date: str = Query(None),

    platform: str = Query(None),

):

    from io import BytesIO

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter



    db: Session = SessionLocal()



    try:

        parsed_date = None

        if report_date:

            parsed_date = _parse_report_date(report_date)



        rows = (

            _daily_report_query(db, parsed_date, platform)

            .order_by(

                DailyReport.platform.asc(),

                DailyReport.style.asc(),

                DailyReport.color.asc(),

                DailyReport.size.asc(),

            )

            .all()

        )



        summary_rows, size_columns = _pivot_daily_report_rows(rows)



        wb = Workbook()

        ws = wb.active

        ws.title = "Daily Report"



        header_fill = PatternFill(

            "solid",

            fgColor="022658",

        )

        header_font = Font(

            bold=True,

            color="FFFFFF",

        )



        plat_label = platform or "All"

        date_label = (

            str(parsed_date) if parsed_date

            else "All dates"

        )



        ws.append(["Daily final order report"])

        ws.merge_cells(

            start_row=1,

            start_column=1,

            end_row=1,

            end_column=4 + len(size_columns),

        )

        ws["A1"].font = Font(bold=True, size=14)

        ws.append(["Date", date_label])

        ws.append(["Platform", plat_label])

        ws.append([])



        headers = [

           

            "Style",

            "Color",

            *size_columns,

            "Total",

        ]

        ws.append(headers)



        for cell in ws[ws.max_row]:

            cell.font = header_font

            cell.fill = header_fill

            cell.alignment = Alignment(horizontal="center")



        for item in summary_rows:

            ws.append([

               

                item["style"],

                item["color"],

                *[

                    item["sizes"].get(size, "")

                    for size in size_columns

                ],

                item["total"],

            ])



        for column_cells in ws.columns:

            column_letter = get_column_letter(

                column_cells[0].column,

            )

            max_length = 0

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(

                        max_length,

                        len(str(cell.value)),

                    )

            ws.column_dimensions[column_letter].width = min(

                max(max_length + 2, 10),

                40,

            )



        buffer = BytesIO()

        wb.save(buffer)

        buffer.seek(0)



        date_part = (

            str(parsed_date) if parsed_date

            else datetime.now().strftime("%d-%m-%Y")

        )

        plat_part = (platform or "all").lower()

        filename = (

            f"daily_report_{date_part}_{plat_part}.xlsx"

        )



        return StreamingResponse(

            buffer,

            media_type=(

                "application/vnd.openxmlformats-officedocument"

                ".spreadsheetml.sheet"

            ),

            headers={

                "Content-Disposition": (

                    f'attachment; filename="{filename}"'

                )

            },

        )

    finally:

        db.close()





@router.delete("/daily-report")

def delete_daily_report(

    report_date: str = Query(...),

    platform: str = Query(None),

    password: str = Query(""),

):

    if password != "Admin":

        raise HTTPException(

            status_code=403,

            detail="Invalid password",

        )



    db: Session = SessionLocal()



    try:

        parsed_date = _parse_report_date(report_date)



        restore_result = _restore_inventory_deduction_log(

            db,

            parsed_date,

            platform,

        )



        query = db.query(DailyReport).filter(

            DailyReport.report_date == parsed_date

        )



        if platform == "All":

            pass

        elif platform:

            query = query.filter(

                DailyReport.platform == platform

            )



        deleted = query.delete(synchronize_session=False)



        sales_query = db.query(DailySalesReport).filter(

            DailySalesReport.report_date == parsed_date

        )



        if platform == "All":

            sales_query = sales_query.filter(

                DailySalesReport.platform.in_(REPORTING_PLATFORM_NAMES)

            )

        elif platform:

            sales_query = sales_query.filter(

                DailySalesReport.platform == platform

            )



        deleted_sales = sales_query.delete(

            synchronize_session=False

        )



        upload_query = db.query(SalesUpload).filter(

            SalesUpload.report_date == parsed_date

        )



        if platform == "All":

            upload_query = upload_query.filter(

                SalesUpload.platform.in_(REPORTING_PLATFORM_NAMES)

            )

        elif platform:

            upload_query = upload_query.filter(

                SalesUpload.platform == platform

            )



        upload_rows = upload_query.all()

        deleted_upload_files = _delete_saved_upload_files(upload_rows)



        deleted_uploads = upload_query.delete(

            synchronize_session=False

        )



        analytics_query = db.query(SalesAnalyticsDetail).filter(

            SalesAnalyticsDetail.report_date == parsed_date

        )



        if platform == "All":

            analytics_query = analytics_query.filter(

                SalesAnalyticsDetail.platform.in_(REPORTING_PLATFORM_NAMES)

            )

        elif platform:

            analytics_query = analytics_query.filter(

                SalesAnalyticsDetail.platform == platform

            )



        deleted_analytics_rows = analytics_query.delete(

            synchronize_session=False

        )



        deleted_zone_rows = 0

        deleted_zone_batch_rows = 0

        deleted_zone_batch_item_rows = 0

        if (platform or "All") in ("All", "Flipkart"):

            deleted_zone_batch_item_rows = (

                db.query(FlipkartZoneBatchItem)

                .filter(FlipkartZoneBatchItem.report_date == parsed_date)

                .delete(synchronize_session=False)

            )

            deleted_zone_batch_rows = (

                db.query(FlipkartZoneBatch)

                .filter(FlipkartZoneBatch.report_date == parsed_date)

                .delete(synchronize_session=False)

            )

            deleted_zone_rows = (

                db.query(FlipkartZoneReport)

                .filter(FlipkartZoneReport.report_date == parsed_date)

                .delete(synchronize_session=False)

            )



        if platform and platform != "All":

            _rebuild_all_daily_report_rows(

                db,

                parsed_date,

            )



        db.commit()
        write_inventory_snapshot("stock", _serialize_stock_inventory_rows(db))
        write_inventory_snapshot("return", _serialize_return_inventory_rows(db))
        write_inventory_snapshot("sticker", _serialize_sticker_inventory_rows(db))



        return {

            "message": "Daily report deleted successfully",

            "deleted_rows": deleted,

            "deleted_sales_rows": deleted_sales,

            "deleted_upload_markers": deleted_uploads,

            "deleted_upload_files": deleted_upload_files,

            "deleted_analytics_rows": deleted_analytics_rows,

            "deleted_zone_rows": deleted_zone_rows,

            "deleted_zone_batch_rows": deleted_zone_batch_rows,

            "deleted_zone_batch_item_rows": deleted_zone_batch_item_rows,

            "restored_inventory_rows": restore_result["restored_logs"],

            "restored_inventory_qty": restore_result["restored_qty"],

            "report_date": str(parsed_date),

            "platform": platform or "all",

        }

    except Exception:

        db.rollback()

        raise

    finally:

        db.close()





MALUR_WAREHOUSE_LABEL_COLUMNS = [

    "size",

    "sleeve",

    "pattern",

    "style_code",

    "type",

    "brand",

    "Net Quantity - 1 U",

    "MRP Rs.XXX.00 (Inclusive of all taxes)",

    "Generic Name",

    "Month & Year of Manufacturing",

    "Manufactured by / Marketed by",

    "SKU ID",

    "Customer Care Details",

    "EAN/FSN/LID Barcode",

]





def _warehouse_label_cell(row, column_name: str):

    value = row.get(column_name, "")



    if pd.isna(value):

        return ""



    return str(value).strip()





def _warehouse_label_mfg_date(row):

    value = row.get("Month & Year of Manufacturing", "")



    if pd.isna(value):

        return ""



    if isinstance(value, (datetime, pd.Timestamp)):

        return value.strftime("%b-%y")



    return str(value).strip()





def _build_flipkart_warehouse_label_pdf(excel_path: str, pdf_path: str):

    from reportlab.graphics.barcode import code128

    from reportlab.lib import colors

    from reportlab.lib.units import mm

    from reportlab.pdfgen import canvas



    try:

        df = pd.read_excel(excel_path)

    except Exception as error:

        raise HTTPException(

            status_code=400,

            detail=f"Could not read Excel file: {error}",

        ) from error



    missing_columns = [

        column

        for column in MALUR_WAREHOUSE_LABEL_COLUMNS

        if column not in df.columns

    ]



    if missing_columns:

        raise HTTPException(

            status_code=400,

            detail=(

                "Missing required columns: "

                + ", ".join(missing_columns)

            ),

        )



    if df.empty:

        raise HTTPException(

            status_code=400,

            detail="Excel file does not contain any label rows.",

        )



    page_width, page_height = 150 * mm, 100 * mm

    pdf = canvas.Canvas(pdf_path, pagesize=(page_width, page_height))



    for _, row in df.iterrows():

        pdf.setLineWidth(2.0)

        pdf.setStrokeColor(colors.black)

        pdf.rect(15, 45, page_width - 30, page_height - 52)



        fields = [

            ("Size", _warehouse_label_cell(row, "size")),

            ("Sleeve", _warehouse_label_cell(row, "sleeve")),

            ("Pattern", _warehouse_label_cell(row, "pattern")),

            ("Style_Code", _warehouse_label_cell(row, "style_code")),

            ("Type", _warehouse_label_cell(row, "type")),

            ("Brand", _warehouse_label_cell(row, "brand")),

            ("Net Quantity - 1U", _warehouse_label_cell(row, "Net Quantity - 1 U")),

            (

                "MRP Rs.XXX.00",

                _warehouse_label_cell(

                    row,

                    "MRP Rs.XXX.00 (Inclusive of all taxes)",

                ),

            ),

            ("(Inclusive of all taxes)", None),

            ("Generic Name", _warehouse_label_cell(row, "Generic Name")),

            ("Month & Year of Manufacturing", _warehouse_label_mfg_date(row)),

            (

                "Manufactured by / Marketed by",

                _warehouse_label_cell(row, "Manufactured by / Marketed by"),

            ),

            ("SKU ID", _warehouse_label_cell(row, "SKU ID")),

            (

                "Customer Care Details",

                _warehouse_label_cell(row, "Customer Care Details"),

            ),

            (

                "EAN/FSN/LID Barcode",

                _warehouse_label_cell(row, "EAN/FSN/LID Barcode"),

            ),

        ]



        pdf.setFont("Times-Bold", 10.5)

        pdf.setFillColor(colors.black)



        col1_x = 25

        hyphen_x = 205

        col2_x = 220

        current_y = page_height - 20

        line_spacing = 15.2



        for label, value in fields:

            pdf.drawString(col1_x, current_y, label)

            if value is not None and value != "":

                pdf.drawString(hyphen_x, current_y, "-")

                pdf.drawString(col2_x, current_y, value)

            current_y -= line_spacing



        barcode_value = _warehouse_label_cell(row, "EAN/FSN/LID Barcode")



        if barcode_value:

            barcode = code128.Code128(

                barcode_value,

                barHeight=20,

                barWidth=1.1,

            )

            barcode_x = (page_width - barcode.width) / 2.0

            barcode_y = 20

            barcode.drawOn(pdf, barcode_x, barcode_y)



            pdf.setFont("Times-Bold", 11)

            text_width = pdf.stringWidth(barcode_value, "Times-Bold", 11.5)

            text_x = (page_width - text_width) / 2.0

            pdf.drawString(text_x, barcode_y - 12, barcode_value)



        pdf.showPage()



    pdf.save()



    return {

        "label_count": len(df),

    }





@router.post("/flipkart-warehouse-labels")

def generate_flipkart_warehouse_labels(

    file: UploadFile = File(...),

):

    filename = _clean_upload_filename(file.filename)

    extension = os.path.splitext(filename)[1].lower()



    if extension not in (".xlsx", ".xls"):

        raise HTTPException(

            status_code=400,

            detail="Upload an Excel file (.xlsx or .xls).",

        )



    input_path = _save_upload(file)

    timestamp = datetime.now().strftime("%d-%m-%Y_%H%M%S")

    output_filename = (

        f"Malur_Warehouse_Label-{datetime.now().strftime('%d-%m-%Y')}.pdf"

    )

    output_path = os.path.join(

        UPLOAD_FOLDER,

        f"flipkart_warehouse_labels_{timestamp}_{output_filename}",

    )



    try:

        _build_flipkart_warehouse_label_pdf(

            input_path,

            output_path,

        )

    except ImportError as error:

        raise HTTPException(

            status_code=500,

            detail="PDF tools are missing. Install reportlab, then try again.",

        ) from error



    return FileResponse(

        output_path,

        media_type="application/pdf",

        filename=output_filename,

    )





@router.get("/sales-reports")

def list_sales_reports():

    db: Session = SessionLocal()



    try:

        rows = (

            db.query(DailySalesReport)

            .order_by(

                DailySalesReport.report_date.desc(),

                DailySalesReport.platform.asc(),

            )

            .all()

        )



        by_date = {}



        for row in rows:

            date_key = str(row.report_date)



            if date_key not in by_date:

                by_date[date_key] = {

                    "report_date": date_key,

                    "platforms": {},

                    "total_orders": 0,

                    "total_piece_qty": 0,

                    "total_invoice_amount": 0.0,

                }



            entry = by_date[date_key]

            platform_name = row.platform

            platform_summary = {

                "platform": platform_name,

                "total_orders": int(row.total_orders or 0),

                "total_piece_qty": int(

                    row.total_piece_qty or 0

                ),

                "total_invoice_amount": round(

                    float(row.total_invoice_amount or 0),

                    2,

                ),

            }



            if platform_name in entry["platforms"]:

                existing = entry["platforms"][platform_name]

                existing["total_orders"] += platform_summary[

                    "total_orders"

                ]

                existing["total_piece_qty"] += platform_summary[

                    "total_piece_qty"

                ]

                existing["total_invoice_amount"] = round(

                    existing["total_invoice_amount"]

                    + platform_summary["total_invoice_amount"],

                    2,

                )

            else:

                entry["platforms"][platform_name] = platform_summary



        reports = []



        for date_key in sorted(by_date.keys(), reverse=True):

            entry = by_date[date_key]

            platforms = sorted(

                entry["platforms"].values(),

                key=lambda item: item["platform"],

            )

            entry["platforms"] = platforms

            entry["total_orders"] = sum(

                item["total_orders"] for item in platforms

            )

            entry["total_piece_qty"] = sum(

                item["total_piece_qty"] for item in platforms

            )

            entry["total_invoice_amount"] = round(

                sum(

                    item["total_invoice_amount"]

                    for item in platforms

                ),

                2,

            )

            entry["platform_count"] = len(platforms)

            reports.append(entry)



        return {

            "count": len(reports),

            "reports": reports,

        }

    finally:

        db.close()





@router.get("/sales-analytics")

def sales_analytics(

    report_date: str = Query(None),

    platform: str = Query("All"),

):

    db: Session = SessionLocal()



    try:

        if report_date:

            parsed_date = _parse_report_date(report_date)

        else:

            parsed_date = datetime.now().date()



        rows = (

            _daily_report_query(

                db,

                parsed_date,

                platform or "All",

            )

            .order_by(

                DailyReport.platform.asc(),

                DailyReport.style.asc(),

                DailyReport.color.asc(),

                DailyReport.size.asc(),

            )

            .all()

        )



        sales_summary = {

            name: {

                "total_orders": 0,

                "total_piece_qty": 0,

                "total_invoice_amount": 0,

            }

            for name in REPORTING_PLATFORM_NAMES

        }



        for platform_name in REPORTING_PLATFORM_NAMES:

            sales_row = _sales_row_for_platform(

                db,

                platform_name,

                parsed_date,

            )



            if not sales_row:

                continue



            sales_summary[platform_name] = {

                "total_orders": int(

                    sales_row.total_orders or 0

                ),

                "total_piece_qty": int(

                    sales_row.total_piece_qty or 0

                ),

                "total_invoice_amount": round(

                    float(sales_row.total_invoice_amount or 0),

                    2

                ),

            }



        platform_totals = {

            name: totals["total_piece_qty"]

            for name, totals in sales_summary.items()

        }



        if platform and platform not in ("", "All"):

            filtered = [

                r for r in rows

                if r.platform == platform

            ]

            if platform == "Flipkart Warehouse":

                filtered.extend(

                    _flipkart_warehouse_daily_report_rows(

                        db,

                        parsed_date,

                    )

                )

            view_platform = platform

        else:

            filtered = list(rows)

            filtered.extend(

                _flipkart_warehouse_daily_report_rows(

                    db,

                    parsed_date,

                )

            )

            view_platform = "All"



        filtered = [

            SimpleNamespace(**row) if isinstance(row, dict) else row

            for row in filtered

        ]



        style_map = {}



        for row in filtered:

            style = row.style or "Unknown"

            size_key = str(row.size or "").upper().strip() or "Î“Ã‡Ã¶"



            if style not in style_map:

                style_map[style] = {

                    "total_qty": 0,

                    "sizes": {},

                }



            qty = int(row.total_order_qty or 0)

            style_map[style]["total_qty"] += qty

            style_map[style]["sizes"][size_key] = (

                style_map[style]["sizes"].get(size_key, 0) + qty

            )



        sorted_styles = sorted(

            style_map.items(),

            key=lambda item: item[1]["total_qty"],

            reverse=True,

        )



        top_products = [

            {

                "style": style,

                "total_qty": data["total_qty"],

                "sizes": data["sizes"],

            }

            for style, data in sorted_styles[:10]

        ]



        style_chart = [

            {

                "style": style,

                "qty": data["total_qty"],

            }

            for style, data in sorted_styles[:10]

        ]



        return {

            "report_date": str(parsed_date),

            "platform": view_platform,

            "platform_totals": platform_totals,

            "sales_summary": sales_summary,

            "total_orders": sum(

                item["total_orders"]

                for item in sales_summary.values()

            ),

            "grand_total": sum(platform_totals.values()),

            "total_invoice_amount": round(

                sum(

                    item["total_invoice_amount"]

                    for item in sales_summary.values()

                ),

                2

            ),

            "top_products": top_products,

            "style_chart": style_chart,

        }

    finally:

        db.close()





@router.get("/sales-analytics/export")

def export_sales_analytics(

    report_date: str = Query(None),

    platform: str = Query("All"),

):

    from io import BytesIO

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter



    data = sales_analytics(

        report_date=report_date,

        platform=platform,

    )



    selected_platform = data.get("platform") or "All"

    sales_summary = data.get("sales_summary") or {}

    top_products = data.get("top_products") or []



    wb = Workbook()

    ws = wb.active

    ws.title = "Sales Summary"



    header_fill = PatternFill(

        "solid",

        fgColor="EDE9FE",

    )

    total_fill = PatternFill(

        "solid",

        fgColor="F5F3FF",

    )



    ws.append([

        "Today's final sale report",

    ])

    ws.merge_cells(

        start_row=1,

        start_column=1,

        end_row=1,

        end_column=4,

    )

    ws["A1"].font = Font(

        bold=True,

        size=14,

    )



    ws.append([

        "Date",

        data.get("report_date", ""),

    ])

    ws.append([

        "Platform",

        selected_platform,

    ])

    ws.append([])

    ws.append([

        "Platform",

        "Total orders",

        "Total piece quantity",

        "Total invoice amount",

    ])



    for cell in ws[5]:

        cell.font = Font(bold=True)

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")



    summary_rows = [

        (name, totals)

        for name, totals in sales_summary.items()

        if selected_platform == "All" or name == selected_platform

    ]



    for name, totals in summary_rows:

        ws.append([

            name,

            int(totals.get("total_orders") or 0),

            int(totals.get("total_piece_qty") or 0),

            float(totals.get("total_invoice_amount") or 0),

        ])



    ws.append([

        "Total",

        int(data.get("total_orders") or 0)

        if selected_platform == "All"

        else sum(

            int(totals.get("total_orders") or 0)

            for _, totals in summary_rows

        ),

        int(data.get("grand_total") or 0)

        if selected_platform == "All"

        else sum(

            int(totals.get("total_piece_qty") or 0)

            for _, totals in summary_rows

        ),

        float(data.get("total_invoice_amount") or 0)

        if selected_platform == "All"

        else sum(

            float(totals.get("total_invoice_amount") or 0)

            for _, totals in summary_rows

        ),

    ])



    for cell in ws[ws.max_row]:

        cell.font = Font(bold=True)

        cell.fill = total_fill



    for row in ws.iter_rows(

        min_row=6,

        min_col=4,

        max_col=4,

    ):

        row[0].number_format = '#,##0.00'



    for column_cells in ws.columns:

        column_letter = get_column_letter(

            column_cells[0].column

        )

        width = max(

            len(str(cell.value or ""))

            for cell in column_cells

        )

        ws.column_dimensions[column_letter].width = min(

            max(width + 2, 12),

            34,

        )



    product_ws = wb.create_sheet("Top Products")

    product_ws.append([

        "Rank",

        "Style",

        "Total quantity",

        "Sizes",

    ])



    for cell in product_ws[1]:

        cell.font = Font(bold=True)

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")



    for index, product in enumerate(top_products, start=1):

        sizes = product.get("sizes") or {}

        size_text = ", ".join(

            f"{size}: {qty}"

            for size, qty in sorted(sizes.items())

        )

        product_ws.append([

            index,

            product.get("style", ""),

            int(product.get("total_qty") or 0),

            size_text,

        ])



    for column_cells in product_ws.columns:

        column_letter = get_column_letter(

            column_cells[0].column

        )

        width = max(

            len(str(cell.value or ""))

            for cell in column_cells

        )

        product_ws.column_dimensions[column_letter].width = min(

            max(width + 2, 12),

            48,

        )



    output = BytesIO()

    wb.save(output)

    output.seek(0)



    safe_platform = re.sub(

        r"[^A-Za-z0-9_-]",

        "_",

        selected_platform.lower(),

    )

    filename = (

        f"final_sale_report_{data.get('report_date')}_{safe_platform}.xlsx"

    )



    return StreamingResponse(

        output,

        media_type=(

            "application/vnd.openxmlformats-officedocument."

            "spreadsheetml.sheet"

        ),

        headers={

            "Content-Disposition": (

                f'attachment; filename="{filename}"'

            )

        },

    )





@router.get("/sales-analytics/export-range")

def export_sales_analytics_range(

    from_date: str = Query(...),

    to_date: str = Query(...),

):

    from io import BytesIO

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter



    parsed_from = _parse_report_date(from_date)

    parsed_to = _parse_report_date(to_date)

    if parsed_from > parsed_to:

        raise HTTPException(

            status_code=400,

            detail="From date cannot be after To date.",

        )



    db: Session = SessionLocal()

    try:

        sales_rows = (

            db.query(DailySalesReport)

            .filter(

                DailySalesReport.report_date >= parsed_from,

                DailySalesReport.report_date <= parsed_to,

            )

            .order_by(

                DailySalesReport.report_date.asc(),

                DailySalesReport.platform.asc(),

            )

            .all()

        )

        if not sales_rows:

            raise HTTPException(

                status_code=404,

                detail="No sales reports found for the selected dates.",

            )



        product_rows = (

            db.query(DailyReport)

            .filter(

                DailyReport.report_date >= parsed_from,

                DailyReport.report_date <= parsed_to,

                DailyReport.platform.in_(PLATFORM_NAMES),

            )

            .all()

        )

        product_rows = list(product_rows)

        product_rows.extend(

            SimpleNamespace(**row)

            for row in _flipkart_warehouse_daily_report_rows(

                db,

                from_date=parsed_from,

                to_date=parsed_to,

            )

        )



        wb = Workbook()

        ws = wb.active

        ws.title = "Sales Summary"

        header_fill = PatternFill("solid", fgColor="EDE9FE")

        total_fill = PatternFill("solid", fgColor="F5F3FF")



        ws.append(["Sales report"])

        ws.merge_cells("A1:E1")

        ws["A1"].font = Font(bold=True, size=14)

        ws.append(["From date", str(parsed_from)])

        ws.append(["To date", str(parsed_to)])

        ws.append([])

        ws.append([

            "Date",

            "Platform",

            "Total orders",

            "Total piece quantity",

            "Total invoice amount",

        ])

        for cell in ws[5]:

            cell.font = Font(bold=True)

            cell.fill = header_fill

            cell.alignment = Alignment(horizontal="center")



        total_orders = 0

        total_pieces = 0

        total_invoice = 0.0

        for row in sales_rows:

            orders = int(row.total_orders or 0)

            pieces = int(row.total_piece_qty or 0)

            invoice = float(row.total_invoice_amount or 0)

            ws.append([

                str(row.report_date),

                row.platform,

                orders,

                pieces,

                invoice,

            ])

            total_orders += orders

            total_pieces += pieces

            total_invoice += invoice



        ws.append([

            "Total",

            "All platforms",

            total_orders,

            total_pieces,

            round(total_invoice, 2),

        ])

        for cell in ws[ws.max_row]:

            cell.font = Font(bold=True)

            cell.fill = total_fill

        for row in ws.iter_rows(min_row=6, min_col=5, max_col=5):

            row[0].number_format = '#,##0.00'



        style_totals = {}

        for row in product_rows:

            style = row.style or "Unknown"

            size = str(row.size or "").strip() or "Î“Ã‡Ã¶"

            entry = style_totals.setdefault(

                style,

                {"total": 0, "sizes": {}},

            )

            quantity = int(row.total_order_qty or 0)

            entry["total"] += quantity

            entry["sizes"][size] = entry["sizes"].get(size, 0) + quantity



        product_ws = wb.create_sheet("Top Products")

        product_ws.append(["Rank", "Style", "Total quantity", "Sizes"])

        for cell in product_ws[1]:

            cell.font = Font(bold=True)

            cell.fill = header_fill

            cell.alignment = Alignment(horizontal="center")

        sorted_products = sorted(

            style_totals.items(),

            key=lambda item: item[1]["total"],

            reverse=True,

        )

        for rank, (style, values) in enumerate(

            sorted_products[:10],

            start=1,

        ):

            sizes = ", ".join(

                f"{size}: {qty}"

                for size, qty in sorted(values["sizes"].items())

            )

            product_ws.append([rank, style, values["total"], sizes])



        for sheet in (ws, product_ws):

            for column_cells in sheet.columns:

                column_letter = get_column_letter(column_cells[0].column)

                width = max(len(str(cell.value or "")) for cell in column_cells)

                sheet.column_dimensions[column_letter].width = min(

                    max(width + 2, 12),

                    48,

                )



        output = BytesIO()

        wb.save(output)

        output.seek(0)

        filename = f"sales_report_{parsed_from}_to_{parsed_to}.xlsx"

        return StreamingResponse(

            output,

            media_type=(

                "application/vnd.openxmlformats-officedocument."

                "spreadsheetml.sheet"

            ),

            headers={

                "Content-Disposition": f'attachment; filename="{filename}"'

            },

        )

    finally:

        db.close()





SALES_ANALYTICS_SIZES = ["S", "M", "L", "XL", "2XL"]





def _parse_optional_report_date(value: Optional[str], fallback: date):

    if isinstance(value, str) and value:

        return _parse_report_date(value)

    return fallback





def _analytics_size(size):

    value = str(size or "").strip().upper()

    if value in ("XXL", "2XL"):

        return "2XL"

    return value or "Unknown"





def _add_size_qty(entry, size, qty):

    size_key = _analytics_size(size)

    entry["sizes"][size_key] = entry["sizes"].get(size_key, 0) + qty

    entry["total"] += qty





def _size_table_rows(grouped, label_key):

    rows = []



    for label, values in grouped.items():

        row = {

            label_key: label,

            "total": values["total"],

            "sizes": {

                size: values["sizes"].get(size, 0)

                for size in SALES_ANALYTICS_SIZES

            },

        }

        rows.append(row)



    return sorted(

        rows,

        key=lambda item: item["total"],

        reverse=True,

    )





def _build_sales_pivot_analytics(

    db: Session,

    from_date: date,

    to_date: date,

    platform: str = "All",

):

    platform_filter = platform if platform and platform != "All" else None



    report_query = db.query(DailyReport).filter(

        DailyReport.report_date >= from_date,

        DailyReport.report_date <= to_date,

        DailyReport.platform.in_(PLATFORM_NAMES),

    )

    sales_query = db.query(DailySalesReport).filter(

        DailySalesReport.report_date >= from_date,

        DailySalesReport.report_date <= to_date,

        DailySalesReport.platform.in_(REPORTING_PLATFORM_NAMES),

    )

    detail_query = db.query(SalesAnalyticsDetail).filter(

        SalesAnalyticsDetail.report_date >= from_date,

        SalesAnalyticsDetail.report_date <= to_date,

    )



    if platform_filter:

        report_query = report_query.filter(

            DailyReport.platform == platform_filter,

        )

        sales_query = sales_query.filter(

            DailySalesReport.platform == platform_filter,

        )

        detail_query = detail_query.filter(

            SalesAnalyticsDetail.platform == platform_filter,

        )



    report_rows = report_query.all()

    sales_rows = sales_query.all()

    detail_rows = detail_query.all()



    platform_style = {}

    style_wise = {}

    color_wise = {}

    main_product_color_wise = {}

    style_color = {}

    sales_by_date = {}

    combo_pack = {}



    for row in report_rows:

        qty = _safe_int(row.total_order_qty)

        style = str(row.style or "Unknown").strip() or "Unknown"

        color = str(row.color or "Unknown").strip() or "Unknown"

        size = _analytics_size(row.size)

        platform_name = row.platform or "Unknown"



        platform_entry = platform_style.setdefault(

            style,

            {

                "style": style,

                "platforms": {},

                "total": 0,

            },

        )

        platform_entry["platforms"][platform_name] = (

            platform_entry["platforms"].get(platform_name, 0) + qty

        )

        platform_entry["total"] += qty



        style_entry = style_wise.setdefault(

            style,

            {"total": 0, "sizes": {}},

        )

        _add_size_qty(style_entry, size, qty)



        color_entry = color_wise.setdefault(

            color,

            {"total": 0, "sizes": {}},

        )

        _add_size_qty(color_entry, size, qty)



        style_color_key = (style, color)

        style_color_entry = style_color.setdefault(

            style_color_key,

            {

                "style": style,

                "color": color,

                "total": 0,

                "sizes": {},

            },

        )

        _add_size_qty(style_color_entry, size, qty)



    for row in sales_rows:

        date_key = str(row.report_date)

        platform_name = row.platform or "Unknown"

        date_entry = sales_by_date.setdefault(

            date_key,

            {

                "date": date_key,

                "platforms": {},

                "total_orders": 0,

                "total_piece_qty": 0,

                "total_invoice_amount": 0.0,

            },

        )

        date_entry["platforms"][platform_name] = {

            "orders": _safe_int(row.total_orders),

            "pieces": _safe_int(row.total_piece_qty),

            "amount": round(_safe_float(row.total_invoice_amount), 2),

        }

        date_entry["total_orders"] += _safe_int(row.total_orders)

        date_entry["total_piece_qty"] += _safe_int(row.total_piece_qty)

        date_entry["total_invoice_amount"] = round(

            date_entry["total_invoice_amount"]

            + _safe_float(row.total_invoice_amount),

            2,

        )



    for row in detail_rows:

        qty = _safe_int(row.piece_qty)

        if qty <= 0:

            continue



        main_product_type = (

            str(row.main_product_type or "").strip()

            or "Unknown"

        )

        detail_color = str(row.color or "Unknown").strip() or "Unknown"



        product_color_key = (main_product_type, detail_color)

        product_color_entry = main_product_color_wise.setdefault(

            product_color_key,

            {

                "main_product_type": main_product_type,

                "color": detail_color,

                "total": 0,

                "sizes": {},

            },

        )

        _add_size_qty(product_color_entry, row.size, qty)



        detail_style = str(row.style or "Unknown").strip() or "Unknown"

        if row.platform == "Flipkart Warehouse":

            warehouse_platform_entry = platform_style.setdefault(

                detail_style,

                {

                    "style": detail_style,

                    "platforms": {},

                    "total": 0,

                },

            )

            warehouse_platform_entry["platforms"][row.platform] = (

                warehouse_platform_entry["platforms"].get(row.platform, 0)

                + qty

            )

            warehouse_platform_entry["total"] += qty



            warehouse_style_entry = style_wise.setdefault(

                detail_style,

                {"total": 0, "sizes": {}},

            )

            _add_size_qty(warehouse_style_entry, row.size, qty)



            warehouse_style_color_key = (

                detail_style,

                detail_color,

            )

            warehouse_style_color_entry = style_color.setdefault(

                warehouse_style_color_key,

                {

                    "style": detail_style,

                    "color": detail_color,

                    "total": 0,

                    "sizes": {},

                },

            )

            _add_size_qty(

                warehouse_style_color_entry,

                row.size,

                qty,

            )



        key = (

            row.pack_of or "1p",

            detail_style,

            row.color or "Unknown",

        )

        entry = combo_pack.setdefault(

            key,

            {

                "pack_of": row.pack_of or "1p",

                "style": detail_style,

                "color": row.color or "Unknown",

                "total": 0,

                "sizes": {},

            },

        )

        _add_size_qty(entry, row.size, qty)



    platform_rows = sorted(

        platform_style.values(),

        key=lambda item: item["total"],

        reverse=True,

    )



    return {

        "from_date": str(from_date),

        "to_date": str(to_date),

        "platform": platform_filter or "All",

        "platforms": REPORTING_PLATFORM_NAMES,

        "sizes": SALES_ANALYTICS_SIZES,

        "totals": {

            "orders": sum(_safe_int(row.total_orders) for row in sales_rows),

            "pieces": sum(_safe_int(row.total_piece_qty) for row in sales_rows),

            "amount": round(

                sum(_safe_float(row.total_invoice_amount) for row in sales_rows),

                2,

            ),

            "report_rows": len(report_rows),

            "combo_rows": len(detail_rows),

        },

        "platform_wise": platform_rows,

        "style_wise": _size_table_rows(style_wise, "style"),

        "color_wise": (

            sorted(

                main_product_color_wise.values(),

                key=lambda item: item["total"],

                reverse=True,

            )

            if main_product_color_wise

            else _size_table_rows(color_wise, "color")

        ),

        "style_color_wise": sorted(

            [

                {

                    "style": values["style"],

                    "color": values["color"],

                    "total": values["total"],

                    "sizes": {

                        size: values["sizes"].get(size, 0)

                        for size in SALES_ANALYTICS_SIZES

                    },

                }

                for values in style_color.values()

            ],

            key=lambda item: item["total"],

            reverse=True,

        ),

        "combo_pack_wise": sorted(

            combo_pack.values(),

            key=lambda item: item["total"],

            reverse=True,

        ),

        "sales_wise": sorted(

            sales_by_date.values(),

            key=lambda item: item["date"],

        ),

    }





@router.get("/sales-pivot-analytics")

def sales_pivot_analytics(

    platform: str = Query("All"),

    period: str = Query(None),

    from_date: str = Query(None),

    to_date: str = Query(None),

    all_dates: bool = Query(False),

):

    db: Session = SessionLocal()



    try:

        latest_sales = (

            db.query(DailySalesReport.report_date)

            .order_by(DailySalesReport.report_date.desc())

            .first()

        )



        earliest_sales = (

            db.query(DailySalesReport.report_date)

            .order_by(DailySalesReport.report_date.asc())

            .first()

        )



        latest_date = (

            latest_sales[0]

            if latest_sales

            else datetime.now().date()

        )



        earliest_date = (

            earliest_sales[0]

            if earliest_sales

            else latest_date

        )



        period = (period or "").lower()



        if period == "weekly":

            parsed_to = latest_date

            parsed_from = latest_date - timedelta(days=6)



        elif period == "monthly":

            parsed_to = latest_date

            parsed_from = latest_date - timedelta(days=29)



        elif period == "custom":

            if not from_date or not to_date:

                raise HTTPException(

                    status_code=400,

                    detail="Please select From and To dates.",

                )



            parsed_from = _parse_report_date(from_date)

            parsed_to = _parse_report_date(to_date)



        else:

            # Backward compatibility

            fallback_to = latest_date

            fallback_from = (

                earliest_date

                if all_dates

                else latest_date.replace(day=1)

            )



            parsed_from = _parse_optional_report_date(

                from_date,

                fallback_from,

            )



            parsed_to = _parse_optional_report_date(

                to_date,

                fallback_to,

            )

        selected_platform = (

            platform

            if isinstance(platform, str) and platform

            else "All"

        )



        if parsed_from > parsed_to:

            raise HTTPException(

                status_code=400,

                detail="From date cannot be after To date.",

            )



        return _build_sales_pivot_analytics(

            db,

            parsed_from,

            parsed_to,

            selected_platform,

        )

    finally:

        db.close()





def _write_size_sheet(ws, rows, primary_headers, sizes):

    header_labels = {

        "main_product_type": "Main product Type",

        "pack_of": "Pack",

    }

    headers = [

        header_labels.get(header, header.replace("_", " ").title())

        for header in primary_headers

    ] + sizes + ["Grand Total"]

    ws.append(headers)



    for row in rows:

        values = [row.get(header, "") for header in primary_headers]

        values.extend(row.get("sizes", {}).get(size, 0) or "" for size in sizes)

        values.append(row.get("total", 0))

        ws.append(values)





def _style_workbook_sheet(wb, title, rows, primary_headers, sizes):

    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter



    ws = wb.create_sheet(title)

    _write_size_sheet(ws, rows, primary_headers, sizes)

    header_fill = PatternFill("solid", fgColor="DBEAFE")



    for cell in ws[1]:

        cell.font = Font(bold=True)

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")



    for column_cells in ws.columns:

        column_letter = get_column_letter(column_cells[0].column)

        width = max(len(str(cell.value or "")) for cell in column_cells)

        ws.column_dimensions[column_letter].width = min(max(width + 2, 12), 34)



    return ws





@router.get("/sales-pivot-analytics/export")

def export_sales_pivot_analytics(

    from_date: str = Query(None),

    to_date: str = Query(None),

    platform: str = Query("All"),

    report_type: str = Query(None),

    all_dates: bool = Query(False),

):

    from io import BytesIO

    from openpyxl import Workbook

    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter



    data = sales_pivot_analytics(

        from_date=from_date,

        to_date=to_date,

        platform=platform,

        all_dates=all_dates,

    )



    report_type = report_type if isinstance(report_type, str) else None

    valid_report_types = {

        "platform-wise",

        "style-wise",

        "color-wise",

        "style-color-wise",

        "combo-pack-wise",

        "sales-wise",

    }

    if report_type and report_type not in valid_report_types:

        raise HTTPException(status_code=400, detail="Unknown analytics report type.")



    wb = Workbook()

    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="DBEAFE")



    if not report_type or report_type == "platform-wise":

        ws = wb.create_sheet("Platform Wise")

        ws.append(["Style", *data["platforms"], "Grand Total"])

        for row in data["platform_wise"]:

            ws.append([

                row["style"],

                *[

                    row.get("platforms", {}).get(platform_name, 0) or ""

                    for platform_name in data["platforms"]

                ],

                row["total"],

            ])



    size_reports = [

        ("style-wise", "Style Wise", "style_wise", ["style"]),

        (

            "color-wise",

            "Color Wise",

            "color_wise",

            ["main_product_type", "color"],

        ),

        (

            "style-color-wise",

            "Style Color Wise",

            "style_color_wise",

            ["style", "color"],

        ),

        (

            "combo-pack-wise",

            "Combo Pack Wise",

            "combo_pack_wise",

            ["pack_of", "style", "color"],

        ),

    ]

    for key, title, data_key, headers in size_reports:

        if not report_type or report_type == key:

            _style_workbook_sheet(

                wb, title, data[data_key], headers, data["sizes"]

            )



    if not report_type or report_type == "sales-wise":

        sales_ws = wb.create_sheet("Sales Wise")

        sales_ws.append(["Date", "Platform", "Orders", "Pieces", "Amount"])

        for item in data["sales_wise"]:

            for platform_name, totals in item.get("platforms", {}).items():

                sales_ws.append([

                    item["date"],

                    platform_name,

                    totals.get("orders", 0),

                    totals.get("pieces", 0),

                    totals.get("amount", 0),

                ])



    for sheet in wb.worksheets:

        for cell in sheet[1]:

            cell.font = Font(bold=True)

            cell.fill = header_fill

            cell.alignment = Alignment(horizontal="center")



        for column_cells in sheet.columns:

            column_letter = get_column_letter(column_cells[0].column)

            width = max(len(str(cell.value or "")) for cell in column_cells)

            sheet.column_dimensions[column_letter].width = min(

                max(width + 2, 12),

                34,

            )



    output = BytesIO()

    wb.save(output)

    output.seek(0)



    export_name = report_type or "sales_pivot_analytics"

    filename = f"{export_name}_{data['from_date']}_to_{data['to_date']}.xlsx"

    return StreamingResponse(

        output,

        media_type=(

            "application/vnd.openxmlformats-officedocument."

            "spreadsheetml.sheet"

        ),

        headers={

            "Content-Disposition": f'attachment; filename="{filename}"'

        },

    )





class ReturnInventoryUpdate(BaseModel):

    qty: int = Field(ge=0)





class StockInventoryUpdate(BaseModel):

    qty: int = Field(ge=0)





class StickerInventoryUpdate(BaseModel):

    qty: int = Field(ge=0)





def _admin_key_valid(x_admin_key: Optional[str]) -> bool:

    expected = os.getenv("ADMIN_API_KEY", "dev-admin")

    return bool(x_admin_key) and x_admin_key == expected







def get_stock_inventory_style(style):

    style_text = str(style).strip().lower()



    if "lsds" in style_text or style_text.startswith("sn"):

        return "lsds"



    if "gv" in style_text:

        return "gv print"



    if "sprn" in style_text:

        return "sprn"



    return None

def normalize_stock_inventory_color(color):

    return clean_color_name(color or "")





STICKER_COLORS = [

    "1 black",

    "2 white",

    "3 grey",

    "4 sandal",

    "5 navy",

    "6 pink",

    "7 brown",

    "8 olive",

]



LSDS_STICKER_STYLES = [

    f"lsds{n:02d}"

    for n in (

        1, 2, 3, 4, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21

    )

]



SN_STICKER_STYLES = [

    "sn450",

    "sn451",

    "sn452",

]



STICKER_STYLES = LSDS_STICKER_STYLES + SN_STICKER_STYLES



STICKER_COLOR_BY_NAME = {

    color.split(" ", 1)[1]: color

    for color in STICKER_COLORS

}





def get_sticker_inventory_style(style):

    style_text = str(style).strip().lower().replace(" ", "")



    lsds_match = re.search(

        r"lsds(\d+)",

        style_text,

    )

    if lsds_match:

        normalized = f"lsds{int(lsds_match.group(1)):02d}"

        if normalized in LSDS_STICKER_STYLES:

            return normalized



    for sticker_style in SN_STICKER_STYLES:

        if sticker_style in style_text:

            return sticker_style



    return None





def normalize_sticker_inventory_color(color):

    return clean_color_name(color or "") or None





def _normalize_inventory_key_part(value):

    return clean_color_name(value or "")





def _same_inventory_value(left, right):

    return _normalize_inventory_key_part(left) == _normalize_inventory_key_part(right)





def _find_stock_inventory_row(db, style, color, size):
    rows = db.query(StockInventory).filter(
        func.lower(StockInventory.style) == str(style or "").strip().lower(),
        func.lower(StockInventory.size) == str(size or "").strip().lower(),
    ).all()
    return next((row for row in rows if _same_inventory_value(row.color, color)), None)





def _find_sticker_inventory_row(db, style, color):

    rows = db.query(StickerInventory).filter(

        func.lower(StickerInventory.style) == str(style or "").strip().lower(),

    ).all()



    return next(

        (

            row

            for row in rows

            if _same_inventory_value(row.color, color)

        ),

        None,

    )





def _sku_master_by_sku(db):

    return {

        normalize_sku(row.sku): row

        for row in db.query(SKUMaster).all()

        if row.sku

    }





def _needed_stock_by_return_key(report_rows):

    needed = {}



    for row in report_rows:

        qty = _safe_int(

            row.get(

                "need_from_stock",

                row.get("stock_inventory", 0),

            )

        )



        if qty <= 0:

            continue



        key = (

            str(row.get("style", "")).strip(),

            clean_color_name(row.get("color", "")),

            str(row.get("size", "")).strip().upper(),

        )

        needed[key] = needed.get(key, 0) + qty



    return needed





def _sku_piece_stock_needs(aggregated_orders, report_rows, db):

    sku_lookup = _sku_master_by_sku(db)

    remaining_needed = _needed_stock_by_return_key(report_rows)



    for order in aggregated_orders:

        sku_master = sku_lookup.get(normalize_sku(order.get("sku", "")))



        if not sku_master:

            continue



        order_qty = _safe_int(

            order.get(

                "total_qty",

                order.get("quantity", 0),

            )

        )



        if order_qty <= 0:

            continue



        sku_style = str(sku_master.style or "").strip()

        sku_size = str(sku_master.size or "").strip().upper()



        for piece in sku_master.pieces:

            piece_qty = _safe_int(piece.qty)

            color = clean_color_name(piece.color or "")



            if not sku_style or not sku_size or not color or piece_qty <= 0:

                continue



            return_key = (

                sku_style,

                color,

                sku_size,

            )

            needed_qty = remaining_needed.get(return_key, 0)



            if needed_qty <= 0:

                continue



            requested_qty = min(order_qty * piece_qty, needed_qty)



            if requested_qty <= 0:

                continue



            remaining_needed[return_key] = needed_qty - requested_qty



            yield sku_master, color, sku_size, requested_qty





def deduct_lsds_stock_inventory(aggregated_orders, report_rows, db):

    lines_updated = 0

    total_qty_deducted = 0

    deductions = []



    for sku_master, color, size, qty in _sku_piece_stock_needs(

        aggregated_orders,

        report_rows,

        db,

    ):

        style = str(sku_master.style or "").strip()



        if not style:

            continue



        stock_row = _find_stock_inventory_row(

            db,

            style,

            color,

            size,

        )



        if not stock_row:

            stock_row = StockInventory(

                style=style,

                color=color,

                size=size,

                qty=0,

            )

            db.add(stock_row)

            db.flush()



        previous_qty = _safe_int(stock_row.qty)

        deducted_qty = min(

            previous_qty,

            qty,

        )

        stock_row.qty = max(

            previous_qty - qty,

            0,

        )



        lines_updated += 1

        total_qty_deducted += deducted_qty



        deductions.append({

            "style": stock_row.style,

            "color": stock_row.color,

            "size": stock_row.size,

            "requested_qty": qty,

            "deducted_qty": deducted_qty,

            "remaining_qty": stock_row.qty,

        })



    return {

        "lines_updated": lines_updated,

        "total_qty_deducted": total_qty_deducted,

        "deductions": deductions,

    }





def deduct_sticker_inventory(aggregated_orders, report_rows, db):

    lines_updated = 0

    total_qty_deducted = 0

    deductions = []



    for sku_master, color, _size, qty in _sku_piece_stock_needs(

        aggregated_orders,

        report_rows,

        db,

    ):

        style = str(sku_master.style or "").strip()

        color = normalize_sticker_inventory_color(color)



        if not style or not color:

            continue



        sticker_row = _find_sticker_inventory_row(

            db,

            style,

            color,

        )



        if not sticker_row:

            sticker_row = StickerInventory(

                style=style,

                color=color,

                qty=0,

            )

            db.add(sticker_row)

            db.flush()



        previous_qty = _safe_int(sticker_row.qty)

        deducted_qty = min(

            previous_qty,

            qty,

        )

        sticker_row.qty = max(

            previous_qty - qty,

            0,

        )



        lines_updated += 1

        total_qty_deducted += deducted_qty



        deductions.append({

            "style": sticker_row.style,

            "color": sticker_row.color,

            "requested_qty": qty,

            "deducted_qty": deducted_qty,

            "remaining_qty": sticker_row.qty,

        })



    return {

        "lines_updated": lines_updated,

        "total_qty_deducted": total_qty_deducted,

        "deductions": deductions,

    }








    sizes = ["M", "L", "XL"]



    lsds_colors = [

        ("1 black", 0, 0, 0),

        ("2 white", 0, 0, 0),

        ("3 grey", 0, 0, 0),

        ("4 sandal", 0, 0, 0),

        ("5 navy", 0, 0, 0),

        ("6 pink", 0, 0, 0),

        ("7 brown", 0, 0, 0),

        ("8 olive", 0, 0, 0),

        ("9 cream", 0, 0, 0),

        ("10 grey melange", 0, 0, 0),

        ("11 charcoal melange", 0, 0, 0),

        ("12 dark grey", 0, 0, 0),

    ]



    gv_print_colors = [

        ("black", 0, 0, 0),

        ("navy", 0, 0, 0),

        ("maroon", 0, 0, 0),

        ('red', 0, 0, 0),

        ("yellow", 0, 0, 0),

        ("sky blue", 0, 0, 0),

        ("light grey", 0, 0, 0),

        ("dark grey", 0, 0, 0),

    ]



    sprn_colors = [

        ("black", 0, 0, 0),

        ("white", 0, 0, 0),

        ("navy", 0, 0, 0),

        ("maroon", 0, 0, 0),

        ("light grey", 0, 0, 0),

        ("dark grey", 0, 0, 0),

    ]



    inventories = [

        ("lsds", lsds_colors),

        ("gv print", gv_print_colors),

        ("sprn", sprn_colors),

    ]



    for style, colors in inventories:

        for color, *qty_values in colors:

            for size, qty in zip(sizes, qty_values):

                db.add(

                    StockInventory(

                        style=style,

                        color=color,

                        size=size,

                        qty=qty,

                    )

                )



    db.commit()





def ensure_sticker_inventory(db):

    legacy_rows = db.query(StickerInventory).filter(

        StickerInventory.style == "lsds",

    ).all()



    if legacy_rows:

        for row in legacy_rows:

            db.delete(row)

        db.flush()



    existing = {

        (row.style, row.color)

        for row in db.query(

            StickerInventory.style,

            StickerInventory.color,

        ).all()

    }



    added = False

    for style in STICKER_STYLES:

        for color in STICKER_COLORS:

            if (style, color) not in existing:

                db.add(

                    StickerInventory(

                        style=style,

                        color=color,

                        qty=0,

                    )

                )

                added = True



    if legacy_rows or added:

        db.commit()





def _serialize_return_inventory_rows(db):
    rows = (
        db.query(ReturnInventory)
        .order_by(
            ReturnInventory.style.asc(),
            ReturnInventory.color.asc(),
            ReturnInventory.size.asc(),
            ReturnInventory.id.asc(),
        )
        .all()
    )
    return [
        {
            "id": row.id,
            "style": row.style or "",
            "color": row.color or "",
            "size": row.size or "",
            "qty": _safe_int(row.qty),
        }
        for row in rows
    ]


def _serialize_stock_inventory_rows(db):
    rows = (
        db.query(StockInventory)
        .order_by(
            StockInventory.style.asc(),
            StockInventory.color.asc(),
            StockInventory.size.asc(),
            StockInventory.id.asc(),
        )
        .all()
    )
    return [
        {
            "id": row.id,
            "style": row.style or "",
            "color": row.color or "",
            "size": row.size or "",
            "qty": _safe_int(row.qty),
        }
        for row in rows
    ]


def _serialize_sticker_inventory_rows(db):
    rows = (
        db.query(StickerInventory)
        .order_by(
            StickerInventory.style.asc(),
            StickerInventory.color.asc(),
            StickerInventory.id.asc(),
        )
        .all()
    )
    return [
        {
            "id": row.id,
            "style": row.style or "",
            "color": row.color or "",
            "qty": _safe_int(row.qty),
        }
        for row in rows
    ]


def _filter_snapshot_rows(rows, search, fields):
    search_text = str(search or "").strip().lower()
    if not search_text:
        return rows

    filtered = []
    for row in rows or []:
        haystack = " ".join(
            str(row.get(field, "") or "").lower()
            for field in fields
        )
        if search_text in haystack:
            filtered.append(row)
    return filtered


# =====================================

@router.get("/return-inventory")

def list_return_inventory(search: Optional[str] = Query(None)):
    snapshot = read_inventory_snapshot("return")
    if snapshot is not None:
        rows = _filter_snapshot_rows(snapshot.get("rows", []), search, ["style", "color", "size"])
        return {"count": len(rows), "rows": rows}

    db: Session = SessionLocal()
    try:
        merged_rows = merge_return_inventory_rows(db)
        if merged_rows:
            db.commit()

        rows = _serialize_return_inventory_rows(db)
        write_inventory_snapshot("return", rows)
        rows = _filter_snapshot_rows(rows, search, ["style", "color", "size"])
        return {"count": len(rows), "rows": rows}
    finally:
        db.close()

@router.patch("/return-inventory/{inventory_id}")

def update_return_inventory_qty(

    inventory_id: int,

    body: ReturnInventoryUpdate,

    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),

):

    if not _admin_key_valid(x_admin_key):

        raise HTTPException(

            status_code=403,

            detail="Invalid or missing admin key. Set X-Admin-Key header "

            "to match ADMIN_API_KEY (default dev key: dev-admin).",

        )

    db: Session = SessionLocal()

    try:

        row = db.query(ReturnInventory).filter(

            ReturnInventory.id == inventory_id

        ).first()

        if not row:

            raise HTTPException(status_code=404, detail="Inventory row not found")

        row.qty = body.qty

        db.commit()

        db.refresh(row)
        write_inventory_snapshot("return", _serialize_return_inventory_rows(db))

        return {

            "id": row.id,

            "style": row.style,

            "color": row.color,

            "size": row.size,

            "qty": row.qty,

        }

    finally:

        db.close()





# =====================================

# FILE UPLOAD

# =====================================




@router.get("/sticker-sidebar-options")
def get_sticker_sidebar_options(db: Session = Depends(get_db)):
    snapshot = read_inventory_snapshot("sticker_sidebar")
    if snapshot and "sidebar_data" in snapshot:
        return snapshot["sidebar_data"]

    sku_masters = db.query(SKUMaster).all()
    
    product_type_map = {}
    all_styles_map = {}
    all_colors_map = {}
    all_sizes_map = {}
    style_options_map = {}
    
    for row in sku_masters:
        pt = str(row.main_product_type or "Unknown").strip()
        pt_lower = pt.lower()
        if pt_lower not in product_type_map:
            product_type_map[pt_lower] = {"name": pt, "styles": {}}
            
        st = str(row.style or "").strip()
        if st:
            st_lower = st.lower()
            if st_lower not in all_styles_map:
                all_styles_map[st_lower] = st
            if st_lower not in product_type_map[pt_lower]["styles"]:
                product_type_map[pt_lower]["styles"][st_lower] = st

            if st_lower not in style_options_map:
                style_options_map[st_lower] = {
                    "style": st,
                    "main_product_type": pt,
                    "colors": {},
                    "sizes": {},
                }
            elif not style_options_map[st_lower]["main_product_type"] and pt:
                style_options_map[st_lower]["main_product_type"] = pt
                
        sz = str(row.size or "").strip()
        if sz:
            sz_lower = sz.lower()
            if sz_lower not in all_sizes_map:
                all_sizes_map[sz_lower] = sz
            if st:
                style_options_map[st_lower]["sizes"][sz_lower] = sz
                
        for piece in row.pieces:
            c = str(piece.color or "").strip()
            if c:
                c_lower = c.lower()
                if c_lower not in all_colors_map:
                    all_colors_map[c_lower] = c
                if st:
                    style_options_map[st_lower]["colors"][c_lower] = c

    product_types = [
        {
            "name": v["name"],
            "styles": sorted(list(v["styles"].values()))
        }
        for v in product_type_map.values()
    ]
    
    product_types.sort(key=lambda x: x["name"].lower())

    style_options = [
        {
            "style": v["style"],
            "main_product_type": v["main_product_type"],
            "colors": sorted(list(v["colors"].values())),
            "sizes": sorted(list(v["sizes"].values())),
        }
        for v in style_options_map.values()
    ]
    style_options.sort(
        key=lambda x: (
            str(x.get("main_product_type") or "").lower(),
            str(x.get("style") or "").lower(),
        )
    )
    
    result = {
        "product_types": product_types,
        "all_styles": sorted(list(all_styles_map.values())),
        "colors": sorted(list(all_colors_map.values())),
        "sizes": sorted(list(all_sizes_map.values())),
        "style_options": style_options,
    }
    
    write_inventory_snapshot("sticker_sidebar", [], extra={"sidebar_data": result})
    return result

@router.get("/sku-master/options")

def get_sku_master_options(db: Session = Depends(get_db)):

    styles = {

        str(row.style).strip()

        for row in db.query(SKUMaster.style).all()

        if row.style and str(row.style).strip()

    }

    sizes = {

        str(row.size).strip()

        for row in db.query(SKUMaster.size).all()

        if row.size and str(row.size).strip()

    }

    colors = {

        clean_color_name(row.color or "")

        for row in db.query(SKUPiece.color).all()

        if row.color and clean_color_name(row.color or "")

    }

    colors_by_position = [set() for _ in range(5)]

    current_sku_master_id = None

    current_position = 0



    for piece in (

        db.query(SKUPiece)

        .order_by(SKUPiece.sku_master_id.asc(), SKUPiece.id.asc())

        .all()

    ):

        if piece.sku_master_id != current_sku_master_id:

            current_sku_master_id = piece.sku_master_id

            current_position = 0



        color_value = clean_color_name(piece.color or "")



        if color_value and current_position < len(colors_by_position):

            colors_by_position[current_position].add(color_value)



        current_position += 1



    return {

        "styles": sorted(styles, key=str.lower),

        "sizes": sorted(sizes, key=str.lower),

        "colors": sorted(colors, key=str.lower),

        "colors_by_position": [

            sorted(position_colors, key=str.lower)

            for position_colors in colors_by_position

        ],

    }





@router.get("/sku-master/rows")

def get_sku_master_rows(db: Session = Depends(get_db)):

    rows = (

        db.query(SKUMaster)

        .order_by(SKUMaster.style.asc(), SKUMaster.sku.asc())

        .all()

    )

    return {

        "count": len(rows),

        "items": [

            {

                "id": row.id,
                "platform": row.platform or "Common",
                "sku": row.sku or "",
                "style": row.style or "",
                "size": row.size or "",
                "pack_of": row.pack_of or "",
                "full_color": row.full_color or "",
                "main_product_type": row.main_product_type or "",
                "pieces": [

                    {

                        "id": piece.id,

                        "color": piece.color or "",

                        "qty": _safe_int(piece.qty),

                    }

                    for piece in row.pieces

                ],

            }

            for row in rows

        ],

    }





@router.put("/sku-master/rows")

def update_sku_master_rows(

    body: SKUMasterBulkUpdate,

    db: Session = Depends(get_db),

):

    try:

        submitted_ids = {item.id for item in body.items if item.id is not None}

        existing_by_id = {

            row.id: row

            for row in db.query(SKUMaster).filter(

                SKUMaster.id.in_(submitted_ids or {-1})

            ).all()

        }

        missing_ids = submitted_ids.difference(existing_by_id)

        if missing_ids:

            raise HTTPException(status_code=404, detail="One or more SKU rows no longer exist.")



        normalized_values = [normalize_sku(item.sku) for item in body.items]

        if any(not value for value in normalized_values):

            raise HTTPException(status_code=400, detail="SKU cannot be empty.")

        if len(normalized_values) != len(set(normalized_values)):

            raise HTTPException(status_code=400, detail="Duplicate SKU values are not allowed.")



        deleted_ids = set(body.deleted_ids)

        untouched_rows = db.query(SKUMaster).filter(

            ~SKUMaster.id.in_(submitted_ids | deleted_ids or {-1})

        ).all()

        untouched_skus = {normalize_sku(row.sku) for row in untouched_rows}

        duplicate = next((item.sku for item in body.items if normalize_sku(item.sku) in untouched_skus), None)

        if duplicate:

            raise HTTPException(status_code=400, detail=f"SKU already exists: {duplicate}")



        for row_id in deleted_ids:

            row = db.query(SKUMaster).filter(SKUMaster.id == row_id).first()

            if row:

                db.delete(row)



        for item in body.items:

            row = existing_by_id.get(item.id) if item.id is not None else None

            if row is None:

                row = SKUMaster()

                db.add(row)

                db.flush()



            row.platform = (item.platform or "Common").strip() or "Common"

            row.sku = item.sku.strip()

            row.style = (item.style or "").strip()

            row.size = (item.size or "").strip().upper()

            row.pack_of = getattr(item, "pack_of", "")

            row.full_color = getattr(item, "full_color", "")

            row.pieces.clear()



            for piece in item.pieces:

                color = clean_color_name(piece.color or "")

                if color and color not in {"-", "nan"}:

                    row.pieces.append(SKUPiece(color=color, qty=1))



        db.commit()
        _invalidate_sku_master_caches()

        return {"message": "SKU master saved", "count": len(body.items)}

    except HTTPException:

        db.rollback()

        raise

    except Exception:

        db.rollback()

        raise





@router.post("/sku-master/manual")
def save_manual_sku_master(
    items: List[ManualSKUMasterCreate],
    db: Session = Depends(get_db),
):
    saved_items = []
    
    # 1. Pre-fetch matching SKUs using an optimized .in_() query to save memory and CPU
    normalized_incoming_skus = [normalize_sku(item.sku) for item in items if item.sku and item.sku.strip()]
    
    existing_skus_dict = {}
    if normalized_incoming_skus:
        # Match the Python normalize_sku logic in SQL: strip(), upper(), replace("-", "_")
        normalized_db_sku = func.replace(func.upper(func.trim(SKUMaster.sku)), "-", "_")
        
        possible_matches = (
            db.query(SKUMaster)
            .filter(normalized_db_sku.in_(normalized_incoming_skus))
            .all()
        )
        for row in possible_matches:
            existing_skus_dict[normalize_sku(row.sku)] = row

    for item in items:
        sku_value = item.sku.strip()

        if not sku_value:
            continue

        normalized_sku = normalize_sku(sku_value)
        existing_sku = existing_skus_dict.get(normalized_sku)

        if existing_sku:

            sku_master = existing_sku

            sku_master.platform = item.platform or "Common"

            sku_master.sku = sku_value

            sku_master.style = (item.style or "").strip()

            sku_master.size = (item.size or "").strip()

            sku_master.pack_of = getattr(item, "pack_of", "").strip()

            sku_master.full_color = getattr(item, "full_color", "").strip()



            db.query(SKUPiece).filter(

                SKUPiece.sku_master_id == sku_master.id

            ).delete()

        else:

            sku_master = SKUMaster(

                platform=item.platform or "Common",

                sku=sku_value,

                style=(item.style or "").strip(),

                size=(item.size or "").strip(),

                pack_of=getattr(item, "pack_of", "").strip(),

                full_color=getattr(item, "full_color", "").strip()

            )

            db.add(sku_master)

            db.flush()



        for piece in item.pieces:

            color_value = clean_color_name(piece.color or "")



            if (

                not color_value

                or color_value == "nan"

                or color_value == "-"

            ):

                continue



            db.add(

                SKUPiece(

                    sku_master_id=sku_master.id,

                    color=color_value,

                    qty=1,

                )

            )



        saved_items.append(

            {

                "sku": sku_master.sku,

                "platform": sku_master.platform,

            }

        )



    if not saved_items:

        raise HTTPException(

            status_code=400,

            detail="Enter at least one SKU.",

        )



    db.commit()
    _invalidate_sku_master_caches()



    return {

        "message": "SKU master updated",

        "count": len(saved_items),

        "items": saved_items,

    }



@router.post("/upload-file")

def upload_file(

    file: UploadFile = File(...)

):



    file_path = _save_upload(file)



    # =====================================

    # AUTO IMPORT SKU MASTER

    # =====================================



    if (

        "sku" in file.filename.lower()

        or

        "master" in file.filename.lower()

    ):



        try:

            with pd.ExcelFile(

                file_path

            ) as excel_file:



                first_sheet = (

                    excel_file.sheet_names[0]

                )



            df = read_sku_sheet(

                file_path,

                first_sheet

            )

        except Exception as e:

            raise HTTPException(

                status_code=400,

                detail=f"Could not read SKU master Excel file: {e}",

            )



        db: Session = SessionLocal()



        try:



            # =====================

            # CLEAR OLD DATA

            # =====================



            db.query(

                SKUPiece

            ).delete()



            db.query(

                SKUMaster

            ).delete()



            db.commit()



            # =====================

            # IMPORT NEW SKU MASTER

            # =====================



            for _, row in df.iterrows():



                sku_value = str(

                    row.get("sku", "")

                ).strip()



                if (

                    not sku_value

                    or sku_value == "nan"

                ):

                    continue



                existing_sku = db.query(

                    SKUMaster

                ).filter(



                    func.lower(SKUMaster.sku) == sku_value.lower()



                ).first()



                if existing_sku:



                    continue



                sku_master = SKUMaster(



                    platform="Common",



                    sku=sku_value,



                    style=str(

                        row.get("style", "")

                    ).strip(),



                    main_product_type=str(

                        row.get("mainproducttype", "")

                    ).strip(),



                    size=str(

                        row.get("size", "")

                    ).strip(),



                    pack_of=str(
                        row.get("packof", "")
                    ).strip(),



                    full_color=str(

                        row.get("fullcolor", "")

                    ).strip()

                )



                db.add(

                    sku_master

                )



                db.flush()



                color_columns = [

                    "color1",

                    "color2",

                    "color3",

                    "color4",

                    "color5",

                ]



                for color_col in color_columns:



                    color_value = clean_color_name(



                        row.get(

                            color_col,

                            ""

                        )

                    )



                    if (

                        not color_value

                        or

                        color_value == "nan"

                        or

                        color_value == "-"

                    ):

                        continue



                    sku_piece = SKUPiece(



                        sku_master_id=

                        sku_master.id,



                        color=color_value,



                        qty=1

                    )



                    db.add(

                        sku_piece

                    )



            db.commit()
            _invalidate_sku_master_caches()



        except Exception as e:



            db.rollback()



            raise HTTPException(

                status_code=400,

                detail=f"Could not import SKU master: {e}",

            )



        finally:



            db.close()



    return {



        "message":

        "File uploaded successfully",



        "filename":

        getattr(

            file,

            "saved_filename",

            file.filename

        )

    }





# =====================================

# READ EXCEL SHEETS

# =====================================



@router.get("/read-master/{filename}")

def read_master_file(filename: str):



    file_path = os.path.join(

        UPLOAD_FOLDER,

        _clean_upload_filename(filename)

    )



    sheet_names = read_excel_file(

        file_path

    )



    return {

        "filename": filename,

        "sheets": sheet_names

    }







# =====================================

# READ SHEET COLUMNS

# =====================================



@router.get(

    "/read-sheet-columns/{filename}/{sheet_name}"

)

def get_sheet_columns(

    filename: str,

    sheet_name: str

):



    file_path = os.path.join(

        UPLOAD_FOLDER,

        _clean_upload_filename(filename)

    )



    return read_sheet_columns(

        file_path,

        sheet_name

    )





# =====================================

# READ CSV COLUMNS

# =====================================



@router.get("/read-csv-columns/{filename}")

def get_csv_columns(filename: str):



    file_path = os.path.join(

        UPLOAD_FOLDER,

        _clean_upload_filename(filename)

    )



    return read_csv_columns(file_path)





# =====================================

# IMPORT SKU MASTER

# =====================================



@router.post(

    "/import-sku-sheet/{filename}/{sheet_name}"

)

def import_sku_sheet(

    filename: str,

    sheet_name: str

):



    file_path = os.path.join(

        UPLOAD_FOLDER,

        _clean_upload_filename(filename)

    )



    df = read_sku_sheet(

        file_path,

        sheet_name

    )



    db: Session = SessionLocal()



    imported_count = 0



    try:



        for _, row in df.iterrows():



            sku_value = str(

                row.get("sku", "")

            ).strip()



            if (

                not sku_value

                or sku_value == "nan"

            ):

                continue



            existing_sku = db.query(

                SKUMaster

            ).filter(

                SKUMaster.sku == sku_value

            ).first()



            if existing_sku:

                continue



            sku_master = SKUMaster(

                platform="Common",

                sku=sku_value,

                style=str(

                    row.get("style", "")

                ).strip(),

                main_product_type=str(

                    row.get("mainproducttype", "")

                ).strip(),

                size=str(

                    row.get("size", "")

                ).strip()

            )



            db.add(sku_master)



            db.flush()



            color_columns = [

                

                ("color1", "color1 Qty"),



                ("color2", "color2 Qty"),



                ("color3", "color3 Qty"),



                ("color4", "color4 Qty"),



                ("color5", "color5 Qty")

            ]



            for color_col, qty_col in color_columns:



                color_value = clean_color_name(

                    row.get(color_col, "")

                )



                qty_value = row.get(

                    normalize_column_name(qty_col),

                    0

                )



                if (

                    not color_value

                    or color_value == "nan"

                    or color_value == "-"

                ):

                    continue



                try:

                    qty_value = int(qty_value)



                except:

                    qty_value = 0



                if qty_value <= 0:

                    continue



                sku_piece = SKUPiece(

                    sku_master_id=sku_master.id,

                    color=color_value,

                    qty=qty_value

                )



                db.add(sku_piece)



            imported_count += 1



        db.commit()



        return {

            "message":

            "SKU sheet imported successfully",



            "imported_count":

            imported_count

        }



    except Exception as e:



        db.rollback()



        return {

            "error": str(e)

        }



    finally:



        db.close()





# =====================================

# FILTER PREVIEW

# =====================================



@router.get(

    "/filter-orders/{platform}/{filename}"

)

def filter_orders(

    platform: str,

    filename: str,

    flipkart_dispatch_period: Optional[str] = Query(None),

):



    file_path = os.path.join(

        UPLOAD_FOLDER,

        _clean_upload_filename(filename)

    )



    platform = platform.lower()



    if platform == "flipkart":



        result = filter_flipkart_orders(

            file_path,

            flipkart_dispatch_period,

        )



    elif platform == "amazon":



        result = filter_amazon_orders(

            file_path

        )



    elif platform == "ajio":



        result = filter_ajio_orders(

            file_path

        )



    elif platform == "meesho":



        result = filter_meesho_orders(

            file_path

        )



    elif platform in ("flipkart-warehouse", "flipkart_warehouse"):



        result = filter_flipkart_warehouse_orders(

            file_path

        )



    else:



        return {

            "error":

            "Invalid platform"

        }



    return {

        "platform": platform,

        "total_orders": len(result),

        "orders": result

    }





# =====================================

# UPLOAD RETURNS

# =====================================



@router.post(

    "/upload-flipkart-returns"

)

@router.post(

    "/upload-returns"

)

@router.post(

    "/upload-return-inventory"

)

def upload_flipkart_returns(
    file: UploadFile = File(...),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
):
    _cleanup_old_uploads(days=7)
    if not _admin_key_valid(x_admin_key):
        raise HTTPException(status_code=403, detail="Invalid or missing admin key.")

    filename = _clean_upload_filename(file.filename)
    if os.path.splitext(filename)[1].lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Upload an Excel file (.xlsx).")

    file_path = _save_upload(file)
    try:
        data = pd.read_excel(file_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    data.columns = [normalize_column_name(c) for c in data.columns]
    
    if not all(col in data.columns for col in ["style", "color", "size", "qty"]):
        raise HTTPException(status_code=400, detail="Excel file must include style, color, size, qty columns.")

    db: Session = SessionLocal()
    try:
        inventory = {}
        from sqlalchemy import func
        unique_styles = {str(r['style']).strip().lower() for _, r in data.iterrows() if pd.notna(r.get('style')) and str(r['style']).strip()}
        if unique_styles:
            for row in db.query(ReturnInventory).filter(func.lower(func.trim(ReturnInventory.style)).in_(unique_styles)).all():
                key = (str(row.style).strip().lower(), str(row.color).strip().lower(), str(row.size).strip().lower())
                inventory[key] = row

        matched_rows = 0
        total_added = 0
        updated_ids = set()

        for index, row in data.iterrows():
            raw_style = str(row["style"]).strip() if pd.notna(row["style"]) else ""
            raw_color = str(row["color"]).strip() if pd.notna(row["color"]) else ""
            raw_size = str(row["size"]).strip().upper() if pd.notna(row["size"]) else ""
            try:
                qty = int(row["qty"])
            except:
                qty = 0
                
            if (not raw_style and not raw_color) or qty == 0:
                continue
                
            matched_rows += 1
            key = (raw_style.lower(), raw_color.lower(), raw_size.lower())
            db_row = inventory.get(key)
            if not db_row:
                db_row = ReturnInventory(style=raw_style, color=raw_color, size=raw_size, qty=0)
                db.add(db_row)
                db.flush()
                inventory[key] = db_row
                
            db_row.qty = _safe_int(db_row.qty) + qty
            total_added += qty
            updated_ids.add(db_row.id)

        db.commit()
        write_inventory_snapshot("return", _serialize_return_inventory_rows(db))
        return {
            "message": "Return inventory updated",
            "matched_rows": matched_rows,
            "updated_cells": len(updated_ids),
            "total_added": total_added
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Could not update return inventory: {e}",
        )
    finally:
        db.close()


@router.post("/upload-stock-inventory")

def upload_stock_inventory(
    file: UploadFile = File(...),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
):
    _cleanup_old_uploads(days=7)
    if not _admin_key_valid(x_admin_key):
        raise HTTPException(status_code=403, detail="Invalid or missing admin key.")

    filename = _clean_upload_filename(file.filename)
    if os.path.splitext(filename)[1].lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Upload an Excel file (.xlsx).")

    file_path = _save_upload(file)
    try:
        data = pd.read_excel(file_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    data.columns = [normalize_column_name(c) for c in data.columns]
    
    if not all(col in data.columns for col in ["style", "color", "size", "qty"]):
        raise HTTPException(status_code=400, detail="Excel file must include style, color, size, qty columns.")

    db: Session = SessionLocal()
    try:
        inventory = {}
        from sqlalchemy import func
        unique_styles = {str(r['style']).strip().lower() for _, r in data.iterrows() if pd.notna(r.get('style')) and str(r['style']).strip()}
        if unique_styles:
            for row in db.query(StockInventory).filter(func.lower(func.trim(StockInventory.style)).in_(unique_styles)).all():
                key = (str(row.style).strip().lower(), str(row.color).strip().lower(), str(row.size).strip().lower())
                inventory[key] = row

        matched_rows = 0
        total_added = 0
        updated_ids = set()

        for index, row in data.iterrows():
            raw_style = str(row["style"]).strip() if pd.notna(row["style"]) else ""
            raw_color = str(row["color"]).strip() if pd.notna(row["color"]) else ""
            raw_size = str(row["size"]).strip().upper() if pd.notna(row["size"]) else ""
            try:
                qty = int(row["qty"])
            except:
                qty = 0
                
            if (not raw_style and not raw_color) or qty == 0:
                continue
                
            matched_rows += 1
            key = (raw_style.lower(), raw_color.lower(), raw_size.lower())
            db_row = inventory.get(key)
            if not db_row:
                db_row = StockInventory(style=raw_style, color=raw_color, size=raw_size, qty=0)
                db.add(db_row)
                db.flush()
                inventory[key] = db_row
                
            db_row.qty = _safe_int(db_row.qty) + qty
            total_added += qty
            updated_ids.add(db_row.id)

        db.commit()
        write_inventory_snapshot("return", _serialize_return_inventory_rows(db))
        return {
            "message": "Stock inventory updated",
            "matched_rows": matched_rows,
            "updated_cells": len(updated_ids),
            "total_added": total_added
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Could not update stock inventory: {e}",
        )
    finally:
        db.close()


@router.get("/stock-inventory")
def list_stock_inventory(search: Optional[str] = Query(None)):
    snapshot = read_inventory_snapshot("stock")
    if snapshot is not None:
        rows = _filter_snapshot_rows(snapshot.get("rows", []), search, ["style", "color", "size"])
        return {"count": len(rows), "rows": rows}

    db: Session = SessionLocal()
    try:
        rows = _serialize_stock_inventory_rows(db)
        write_inventory_snapshot("stock", rows)
        rows = _filter_snapshot_rows(rows, search, ["style", "color", "size"])
        return {"count": len(rows), "rows": rows}
    finally:
        db.close()

@router.patch("/stock-inventory/{inventory_id}")

def update_stock_inventory_qty(

    inventory_id: int,

    body: StockInventoryUpdate,

    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),

):

    if not _admin_key_valid(x_admin_key):

        raise HTTPException(

            status_code=403,

            detail="Invalid or missing admin key. Set X-Admin-Key header "

            "to match ADMIN_API_KEY (default dev key: dev-admin).",

        )

    db: Session = SessionLocal()

    try:

        row = db.query(StockInventory).filter(

            StockInventory.id == inventory_id

        ).first()

        if not row:

            raise HTTPException(status_code=404, detail="Inventory row not found")

        row.qty = body.qty

        db.commit()

        db.refresh(row)
        write_inventory_snapshot("stock", _serialize_stock_inventory_rows(db))

        return {

            "id": row.id,

            "style": row.style,

            "color": row.color,

            "size": row.size,

            "qty": row.qty,

        }

    finally:

        db.close()





@router.post("/upload-sticker-inventory")

def upload_sticker_inventory(
    file: UploadFile = File(...),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
):
    if not _admin_key_valid(x_admin_key):
        raise HTTPException(status_code=403, detail="Invalid or missing admin key.")

    filename = _clean_upload_filename(file.filename)
    if os.path.splitext(filename)[1].lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Upload an Excel file (.xlsx).")

    file_path = _save_upload(file)
    try:
        data = pd.read_excel(file_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    data.columns = [normalize_column_name(c) for c in data.columns]
    
    if not all(col in data.columns for col in ["style", "color", "qty"]):
        raise HTTPException(status_code=400, detail="Excel file must include style, color, qty columns.")

    db: Session = SessionLocal()
    try:
        inventory = {}
        from sqlalchemy import func
        unique_styles = {str(r['style']).strip().lower() for _, r in data.iterrows() if pd.notna(r.get('style')) and str(r['style']).strip()}
        if unique_styles:
            for row in db.query(StickerInventory).filter(func.lower(func.trim(StickerInventory.style)).in_(unique_styles)).all():
                key = (str(row.style).strip().lower(), str(row.color).strip().lower())
                inventory[key] = row

        matched_rows = 0
        total_added = 0
        updated_ids = set()

        for index, row in data.iterrows():
            raw_style = str(row["style"]).strip() if pd.notna(row["style"]) else ""
            raw_color = str(row["color"]).strip() if pd.notna(row["color"]) else ""
            try:
                qty = int(row["qty"])
            except:
                qty = 0
                
            if (not raw_style and not raw_color) or qty == 0:
                continue
                
            matched_rows += 1
            key = (raw_style.lower(), raw_color.lower())
            db_row = inventory.get(key)
            if not db_row:
                db_row = StickerInventory(style=raw_style, color=raw_color, qty=0)
                db.add(db_row)
                db.flush()
                inventory[key] = db_row
                
            db_row.qty = _safe_int(db_row.qty) + qty
            total_added += qty
            updated_ids.add(db_row.id)

        db.commit()
        write_inventory_snapshot("sticker", _serialize_sticker_inventory_rows(db))
        return {
            "message": "Sticker inventory updated",
            "matched_rows": matched_rows,
            "updated_cells": len(updated_ids),
            "total_added": total_added
        }

    except HTTPException:

        db.rollback()

        raise

    except Exception as e:

        db.rollback()

        raise HTTPException(

            status_code=400,

            detail=f"Could not update sticker inventory: {e}",

        )

    finally:

        db.close()





@router.get("/sticker-inventory")
def list_sticker_inventory(search: Optional[str] = Query(None)):
    snapshot = read_inventory_snapshot("sticker")
    if snapshot is not None:
        rows = _filter_snapshot_rows(snapshot.get("rows", []), search, ["style", "color"])
        return {"count": len(rows), "rows": rows}

    db: Session = SessionLocal()
    try:
        ensure_sticker_inventory(db)
        rows = _serialize_sticker_inventory_rows(db)
        write_inventory_snapshot("sticker", rows)
        rows = _filter_snapshot_rows(rows, search, ["style", "color"])
        return {"count": len(rows), "rows": rows}
    finally:
        db.close()

@router.patch("/sticker-inventory/{inventory_id}")

def update_sticker_inventory_qty(

    inventory_id: int,

    body: StickerInventoryUpdate,

    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),

):

    if not _admin_key_valid(x_admin_key):

        raise HTTPException(

            status_code=403,

            detail="Invalid or missing admin key. Set X-Admin-Key header "

            "to match ADMIN_API_KEY (default dev key: dev-admin).",

        )

    db: Session = SessionLocal()

    try:

        row = db.query(StickerInventory).filter(

            StickerInventory.id == inventory_id

        ).first()

        if not row:

            raise HTTPException(status_code=404, detail="Inventory row not found")

        row.qty = body.qty

        db.commit()

        db.refresh(row)
        write_inventory_snapshot("sticker", _serialize_sticker_inventory_rows(db))

        return {

            "id": row.id,

            "style": row.style,

            "color": row.color,

            "qty": row.qty,

        }

    finally:

        db.close()





@router.get("/packing-inventory")

def list_packing_inventory():

    db: Session = SessionLocal()

    try:

        seed_packing_inventory_if_empty(db)

        db.commit()

        rows = (

            db.query(PackingInventory)

            .order_by(

                PackingInventory.item_type.asc(),

                PackingInventory.platform.asc(),

                PackingInventory.name.asc(),

            )

            .all()

        )

        return {

            "types": PACKING_INVENTORY_TYPES,

            "platforms": PLATFORM_NAMES,

            "rows": [

                {

                    "id": row.id,

                    "item_type": row.item_type,

                    "platform": row.platform,

                    "name": row.name,

                    "qty": row.qty,

                }

                for row in rows

            ],

        }

    finally:

        db.close()





@router.put("/packing-inventory")

def save_packing_inventory(

    body: PackingInventoryBulkUpdate,

    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),

):

    if not _admin_key_valid(x_admin_key):

        raise HTTPException(

            status_code=403,

            detail="Invalid or missing admin key.",

        )



    db: Session = SessionLocal()

    try:

        existing_by_id = {

            row.id: row

            for row in db.query(PackingInventory).all()

        }

        deleted_ids = {

            _safe_int(row_id)

            for row_id in body.deleted_ids

            if _safe_int(row_id) > 0

        }



        for row_id in deleted_ids:

            row = existing_by_id.get(row_id)

            if row:

                db.delete(row)



        seen_keys = set()

        saved_rows = []

        for item in body.items:

            item_type = _clean_packing_item_type(item.item_type)

            platform = _clean_packing_platform(item_type, item.platform)

            name = str(item.name or "").strip()



            if not name:

                raise HTTPException(

                    status_code=400,

                    detail="Packing item name cannot be empty.",

                )



            key = (item_type, platform or "", name.lower())

            if key in seen_keys:

                raise HTTPException(

                    status_code=400,

                    detail="Duplicate packing inventory row found.",

                )

            seen_keys.add(key)



            row = None

            if item.id:

                row = existing_by_id.get(item.id)

                if not row:

                    raise HTTPException(

                        status_code=404,

                        detail="Packing inventory row not found.",

                    )



            if not row:
                row = PackingInventory()
                db.add(row)

            old_qty = row.qty if row.id else 0
            new_qty = max(_safe_int(item.qty), 0)

            if old_qty != new_qty:
                from datetime import date
                today = date.today()
                usage_record = db.query(PackingInventoryUsage).filter_by(
                    usage_date=today,
                    item_type=item_type,
                    platform=platform,
                    name=name
                ).first()
                if not usage_record:
                    usage_record = PackingInventoryUsage(
                        usage_date=today,
                        item_type=item_type,
                        platform=platform,
                        name=name,
                        used_qty=0,
                        restocked_qty=0
                    )
                    db.add(usage_record)
                
                if new_qty < old_qty:
                    usage_record.used_qty += (old_qty - new_qty)
                else:
                    usage_record.restocked_qty += (new_qty - old_qty)

            row.item_type = item_type
            row.platform = platform
            row.name = name
            row.qty = new_qty
            saved_rows.append(row)



        db.flush()

        db.commit()



        return {

            "message": "Packing inventory saved",

            "saved_count": len(saved_rows),

            "deleted_count": len(deleted_ids),

        }

    except HTTPException:

        db.rollback()

        raise

    except Exception as e:

        db.rollback()

        raise HTTPException(

            status_code=400,

            detail=f"Could not save packing inventory: {e}",

        )

    finally:

        db.close()





@router.get("/packing-inventory/export/{item_type}")

def export_packing_inventory_excel(item_type: str):

    from io import BytesIO

    from openpyxl import Workbook

    from openpyxl.styles import Alignment, Font, PatternFill

    from openpyxl.utils import get_column_letter



    clean_type = _clean_packing_item_type(item_type)



    db: Session = SessionLocal()

    try:

        seed_packing_inventory_if_empty(db)

        db.commit()

        rows = (

            db.query(PackingInventory)

            .filter(PackingInventory.item_type == clean_type)

            .order_by(

                PackingInventory.platform.asc(),

                PackingInventory.name.asc(),

            )

            .all()

        )



        wb = Workbook()

        ws = wb.active

        title = PACKING_INVENTORY_LABELS[clean_type]

        ws.title = title[:31]



        if clean_type == "shipping_cover":

            row_by_platform = {

                row.platform: _safe_int(row.qty)

                for row in rows

            }

            ws.append(["Item", *PLATFORM_NAMES])

            ws.append([

                title,

                *[row_by_platform.get(platform, 0) for platform in PLATFORM_NAMES],

            ])

        else:

            ws.append(["Item", "Qty"])

            for row in rows:

                ws.append([row.name, _safe_int(row.qty)])



        header_fill = PatternFill("solid", fgColor="1F4E78")

        for cell in ws[1]:

            cell.font = Font(bold=True, color="FFFFFF")

            cell.fill = header_fill

            cell.alignment = Alignment(horizontal="center")



        for row in ws.iter_rows(min_row=2):

            for cell in row:

                if cell.column > 1:

                    cell.number_format = '#,##0'

                    cell.alignment = Alignment(horizontal="right")



        ws.freeze_panes = "A2"

        for column_cells in ws.columns:

            column_letter = get_column_letter(column_cells[0].column)

            width = max(len(str(cell.value or "")) for cell in column_cells)

            ws.column_dimensions[column_letter].width = min(

                max(width + 2, 12),

                32,

            )



        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return StreamingResponse(

            output,

            media_type=(

                "application/vnd.openxmlformats-officedocument."

                "spreadsheetml.sheet"

            ),

            headers={

                "Content-Disposition": (

                    f'attachment; filename="{clean_type}_inventory.xlsx"'

                )

            },

        )

    finally:

        db.close()





@router.get("/packing-inventory/export")
def export_all_packing_inventory_excel():
    from io import BytesIO
    import xlsxwriter

    db: Session = SessionLocal()
    try:
        seed_packing_inventory_if_empty(db)
        db.commit()
        rows = (
            db.query(PackingInventory)
            .order_by(
                PackingInventory.item_type.asc(),
                PackingInventory.platform.asc(),
                PackingInventory.name.asc(),
            )
            .all()
        )

        rows_by_type = {item_type: [] for item_type in PACKING_INVENTORY_LABELS}
        for row in rows:
            if row.item_type in rows_by_type:
                rows_by_type[row.item_type].append(row)

        output = BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet("Packing Inventory")
        
        header_format = wb.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#1F4E78', 'align': 'center'})
        title_format = wb.add_format({'bold': True, 'bg_color': '#D9EAF7', 'align': 'center'})
        num_format = wb.add_format({'num_format': '#,##0', 'align': 'right'})

        current_row = 0
        col_widths = {}
        
        def set_col_width(col, val):
            w = max(len(str(val or "")) + 2, col_widths.get(col, 12))
            col_widths[col] = min(w, 32)
            
        for item_type in PACKING_INVENTORY_LABELS:
            title = PACKING_INVENTORY_LABELS[item_type]
            item_rows = rows_by_type.get(item_type, [])

            if item_type == "shipping_cover":
                row_by_platform = {row.platform: _safe_int(row.qty) for row in item_rows}
                section_width = len(PLATFORM_NAMES) + 2
                ws.merge_range(current_row, 0, current_row, section_width - 1, title, title_format)
                current_row += 1
                
                headers = ["Item", *PLATFORM_NAMES, "Total"]
                for i, h in enumerate(headers):
                    ws.write(current_row, i, h, header_format)
                    set_col_width(i, h)
                current_row += 1
                
                platform_qtys = [row_by_platform.get(platform, 0) for platform in PLATFORM_NAMES]
                values = [title, *platform_qtys, sum(platform_qtys)]
                ws.write(current_row, 0, values[0])
                set_col_width(0, values[0])
                for i, v in enumerate(values[1:], start=1):
                    ws.write_number(current_row, i, v, num_format)
                    set_col_width(i, v)
                current_row += 2
            else:
                section_width = 2
                ws.merge_range(current_row, 0, current_row, section_width - 1, title, title_format)
                current_row += 1
                
                ws.write(current_row, 0, "Item", header_format)
                ws.write(current_row, 1, "Qty", header_format)
                set_col_width(0, "Item")
                set_col_width(1, "Qty")
                current_row += 1
                
                for row in item_rows:
                    ws.write(current_row, 0, row.name)
                    ws.write_number(current_row, 1, _safe_int(row.qty), num_format)
                    set_col_width(0, row.name)
                    current_row += 1

                ws.write(current_row, 0, "Total")
                ws.write_number(current_row, 1, sum(_safe_int(r.qty) for r in item_rows), num_format)
                current_row += 2

        ws.freeze_panes(1, 0)
        for col, width in col_widths.items():
            ws.set_column(col, col, width)

        wb.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="packing_inventory.xlsx"'},
        )


    finally:

        db.close()





@router.get("/inventory-export/{inventory_type}")

def export_inventory_excel(
    inventory_type: str,
    search: Optional[str] = Query(None),
):
    from io import BytesIO
    import xlsxwriter

    inventory_type = inventory_type.strip().lower()
    if inventory_type not in {"return", "stock", "sticker"}:
        raise HTTPException(status_code=404, detail="Inventory type not found.")

    db: Session = SessionLocal()
    try:
        term = f"%{search.strip()}%" if search and search.strip() else None

        if inventory_type == "return":
            if merge_return_inventory_rows(db):
                db.commit()
            query = db.query(ReturnInventory)
            if term:
                query = query.filter(or_(
                    ReturnInventory.style.ilike(term),
                    ReturnInventory.color.ilike(term),
                    ReturnInventory.size.ilike(term),
                ))
            source_rows = query.all()
            preferred_sizes = ["XS", "S", "M", "L", "XL", "2XL"]
            title = "Return Inventory"
            filename = "return_inventory.xlsx"
        elif inventory_type == "stock":
            query = db.query(StockInventory)
            if term:
                query = query.filter(or_(
                    StockInventory.style.ilike(term),
                    StockInventory.color.ilike(term),
                    StockInventory.size.ilike(term),
                ))
            source_rows = query.all()
            preferred_sizes = ["M", "L", "XL"]
            title = "Stock Inventory"
            filename = "stock_inventory.xlsx"
        else:
            ensure_sticker_inventory(db)
            query = db.query(StickerInventory)
            if term:
                query = query.filter(or_(
                    StickerInventory.style.ilike(term),
                    StickerInventory.color.ilike(term),
                ))
            source_rows = query.all()
            preferred_sizes = []
            title = "Sticker Inventory"
            filename = "sticker_inventory.xlsx"

        output = BytesIO()
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet(title)
        
        header_format = wb.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#1F4E78', 'align': 'center'})
        num_format = wb.add_format({'num_format': '#,##0', 'align': 'right'})

        data_rows = []
        if inventory_type == "sticker":
            colors = list(STICKER_COLORS)
            extra_colors = sorted({row.color for row in source_rows if row.color and row.color not in colors})
            columns = colors + extra_colors
            headers = ["Style", *columns, "Total Qty"]
            
            grouped = {}
            for row in source_rows:
                grouped.setdefault(row.style, {})[row.color] = _safe_int(row.qty)

            for style in sorted(grouped):
                quantities = [grouped[style].get(color, 0) for color in columns]
                data_rows.append([style, *quantities, sum(quantities)])
        else:
            found_sizes = {str(r.size or "").upper().strip() for r in source_rows if str(r.size or "").strip()}
            sizes = [s for s in preferred_sizes if s in found_sizes]
            sizes.extend(sorted(found_sizes - set(preferred_sizes)))
            headers = ["Style", "Color", *sizes, "Total Qty"]
            
            grouped = {}
            for row in source_rows:
                key = (str(row.style).strip().lower(), str(row.color).strip().lower())
                size = str(row.size or "").upper().strip()
                grouped.setdefault(key, {})[size] = _safe_int(row.qty)

            for style, color in sorted(grouped):
                quantities = [grouped[(style, color)].get(size, 0) for size in sizes]
                data_rows.append([style, color, *quantities, sum(quantities)])

        col_widths = [11] * len(headers)
        for i, h in enumerate(headers):
            ws.write(0, i, h, header_format)
            col_widths[i] = max(col_widths[i], len(str(h)) + 2)

        start_num_col = 1 if inventory_type == "sticker" else 2
        for r_idx, row_data in enumerate(data_rows, start=1):
            for c_idx, val in enumerate(row_data):
                if c_idx >= start_num_col:
                    ws.write_number(r_idx, c_idx, val, num_format)
                else:
                    ws.write(r_idx, c_idx, val)
                col_widths[c_idx] = min(max(col_widths[c_idx], len(str(val)) + 2), 32)
                
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, len(data_rows), len(headers) - 1)
        for i, w in enumerate(col_widths):
            ws.set_column(i, i, w)

        wb.close()
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    finally:

        db.close()



# =====================================

# FINAL COMBINED REPORT

# =====================================



@router.post("/generate-final-report")

def generate_final_report(



    flipkart_file: UploadFile = File(None),

    flipkart_dispatch_period: Optional[str] = Form(None),



    amazon_file: UploadFile = File(None),



    ajio_file: UploadFile = File(None),



    meesho_file: UploadFile = File(None),



    myntra_file: UploadFile = File(None),



    flipkart_warehouse_file: UploadFile = File(None),

):



    all_orders = []

    platform_orders = {}

    platform_orders = {}



    # =====================

    # FLIPKART

    # =====================



    if flipkart_file:



        flipkart_orders = (

            _read_uploaded_orders(

                "Flipkart",

                flipkart_file,

                lambda file_path: filter_flipkart_orders(

                    file_path,

                    flipkart_dispatch_period,

                )

            )

        )

        platform_orders["Flipkart"] = flipkart_orders



        all_orders.extend(

            flipkart_orders

        )



    # =====================

    # AMAZON

    # =====================



    if amazon_file:



        amazon_orders = (

            _read_uploaded_orders(

                "Amazon",

                amazon_file,

                filter_amazon_orders

            )

        )

        platform_orders["Amazon"] = amazon_orders



        all_orders.extend(

            amazon_orders

        )



    # =====================

    # AJIO

    # =====================



    if ajio_file:



        ajio_orders = (

            _read_uploaded_orders(

                "Ajio",

                ajio_file,

                filter_ajio_orders

            )

        )

        platform_orders["Ajio"] = ajio_orders



        all_orders.extend(

            ajio_orders

        )



    # =====================

    # MEESHO

    # =====================



    if meesho_file:



        meesho_orders = (

            _read_uploaded_orders(

                "Meesho",

                meesho_file,

                filter_meesho_orders

            )

        )

        platform_orders["Meesho"] = meesho_orders



        all_orders.extend(

            meesho_orders

        )



    # =====================

    # MYNTRA

    # =====================



    if myntra_file:



        myntra_orders = (

            _read_uploaded_orders(

                "Myntra",

                myntra_file,

                filter_myntra_orders

            )

        )

        platform_orders["Myntra"] = myntra_orders



        all_orders.extend(

            myntra_orders

        )



    # =====================

    # FLIPKART WAREHOUSE

    # =====================



    if flipkart_warehouse_file:



        flipkart_warehouse_orders = (

            _read_uploaded_orders(

                "Flipkart Warehouse",

                flipkart_warehouse_file,

                filter_flipkart_warehouse_orders

            )

        )

        platform_orders["Flipkart Warehouse"] = flipkart_warehouse_orders



    # =====================

    # AGGREGATE

    # =====================



    aggregated_orders = aggregate_orders(

        all_orders

    )



    db: Session = SessionLocal()



    try:

        _raise_if_unknown_platform_skus(

            db,

            platform_orders,

        )



        expanded_inventory = expand_inventory(

            aggregated_orders,

            db

        )



        final_report = generate_daily_report(

            expanded_inventory,

            db

        )



        report_date = datetime.now().date()



        all_platform_orders = _orders_for_report_date(

            platform_orders,

            report_date,

            flipkart_dispatch_period,

        )

        all_platform_aggregated = aggregate_orders(

            all_platform_orders

        )

        all_platform_expanded = expand_inventory(

            all_platform_aggregated,

            db

        )

        all_platform_report = generate_daily_report(

            all_platform_expanded,

            db

        )



        save_daily_report_rows(

            db,

            report_date,

            "All",

            all_platform_report

        )



        counted_platforms, skipped_platforms = _save_new_platform_sales(

            db,

            report_date,

            platform_orders,

            {

                "Flipkart": flipkart_file,

                "Amazon": amazon_file,

                "Ajio": ajio_file,

                "Meesho": meesho_file,

                "Myntra": myntra_file,

                "Flipkart Warehouse": flipkart_warehouse_file,

            },

            flipkart_dispatch_period,

        )



        db.commit()



        return {

            "total_platform_orders":

            len(all_orders),



            "total_inventory_items":

            len(final_report),



            "report":

            final_report,



            "sales_counted_platforms":

            counted_platforms,



            "sales_skipped_duplicate_platforms":

            skipped_platforms

        }

    finally:

        db.close()



def _collect_marketplace_orders(

    flipkart_file: UploadFile = None,

    flipkart_dispatch_period: Optional[str] = None,

    amazon_file: UploadFile = None,

    ajio_file: UploadFile = None,

    meesho_file: UploadFile = None,

    myntra_file: UploadFile = None,

    flipkart_warehouse_file: UploadFile = None,

):

    all_orders = []

    platform_orders = {}



    # =====================

    # FLIPKART

    # =====================



    if flipkart_file:



        flipkart_orders = (

            _read_uploaded_orders(

                "Flipkart",

                flipkart_file,

                lambda file_path: filter_flipkart_orders(

                    file_path,

                    flipkart_dispatch_period,

                )

            )

        )

        platform_orders["Flipkart"] = flipkart_orders



        all_orders.extend(

            flipkart_orders

        )



    # =====================

    # AMAZON

    # =====================



    if amazon_file:



        amazon_orders = (

            _read_uploaded_orders(

                "Amazon",

                amazon_file,

                filter_amazon_orders

            )

        )

        platform_orders["Amazon"] = amazon_orders



        all_orders.extend(

            amazon_orders

        )



    # =====================

    # AJIO

    # =====================



    if ajio_file:



        ajio_orders = (

            _read_uploaded_orders(

                "Ajio",

                ajio_file,

                filter_ajio_orders

            )

        )

        platform_orders["Ajio"] = ajio_orders



        all_orders.extend(

            ajio_orders

        )



    # =====================

    # MEESHO

    # =====================



    if meesho_file:



        meesho_orders = (

            _read_uploaded_orders(

                "Meesho",

                meesho_file,

                filter_meesho_orders

            )

        )

        print("MEESHO SAMPLE")

        print(meesho_orders[:5])

        platform_orders["Meesho"] = meesho_orders



        all_orders.extend(

            meesho_orders

        )



            # =====================

    # MYNTRA

    # =====================



    if myntra_file:



        myntra_orders = (

            _read_uploaded_orders(

                "Myntra",

                myntra_file,

                filter_myntra_orders

            )

        )

        platform_orders["Myntra"] = myntra_orders



        all_orders.extend(

            myntra_orders

        )



    # =====================

    # FLIPKART WAREHOUSE

    # =====================



    if flipkart_warehouse_file:



        flipkart_warehouse_orders = (

            _read_uploaded_orders(

                "Flipkart Warehouse",

                flipkart_warehouse_file,

                filter_flipkart_warehouse_orders

            )

        )

        platform_orders["Flipkart Warehouse"] = flipkart_warehouse_orders



    return all_orders, platform_orders





@router.post("/export-final-report")

def export_final_report(



    flipkart_file: UploadFile = File(None),

    flipkart_dispatch_period: Optional[str] = Form(None),



    amazon_file: UploadFile = File(None),



    ajio_file: UploadFile = File(None),



    meesho_file: UploadFile = File(None),



    myntra_file: UploadFile = File(None),



    flipkart_warehouse_file: UploadFile = File(None),



    include_detail_columns: bool = Form(True),

    include_order_summary: bool = Form(True),

):



    all_orders, platform_orders = _collect_marketplace_orders(

        flipkart_file,

        flipkart_dispatch_period,

        amazon_file,

        ajio_file,

        meesho_file,

        myntra_file,

        flipkart_warehouse_file,

    )



    if not platform_orders:

        raise HTTPException(

            status_code=400,

            detail="Upload at least one marketplace order file.",

        )



    aggregated_orders = aggregate_orders(

        all_orders

    )



    db: Session = SessionLocal()

    order_summary_rows = []



    try:

        _raise_if_unknown_platform_skus(

            db,

            platform_orders,

        )



        expanded_inventory = expand_inventory(

            aggregated_orders,

            db

        )



        final_report = generate_daily_report(

            expanded_inventory,

            db

        )



        report_date = datetime.now().date()



        all_platform_orders = _orders_for_report_date(

            platform_orders,

            report_date,

            flipkart_dispatch_period,

        )

        all_platform_aggregated = aggregate_orders(

            all_platform_orders

        )

        all_platform_expanded = expand_inventory(

            all_platform_aggregated,

            db

        )

        all_platform_report = generate_daily_report(

            all_platform_expanded,

            db

        )



        save_daily_report_rows(

            db,

            report_date,

            "All",

            all_platform_report

        )



        counted_platforms, _ = _save_new_platform_sales(

            db,

            report_date,

            platform_orders,

            {

                "Flipkart": flipkart_file,

                "Amazon": amazon_file,

                "Ajio": ajio_file,

                "Meesho": meesho_file,

                "Myntra": myntra_file,

                "Flipkart Warehouse": flipkart_warehouse_file,

            },

            flipkart_dispatch_period,

        )



        if include_order_summary:

            for platform_name, orders in platform_orders.items():

                if platform_name not in PLATFORM_NAMES:

                    continue



                platform_aggregated = aggregate_orders(orders)

                platform_expanded = expand_inventory(

                    platform_aggregated,

                    db

                )

                order_summary_rows.append(

                    {

                        "platform": platform_name,

                        "total_orders": _unique_order_count(orders),

                        "piece_qty": sum(

                            _safe_int(item.get("qty", 0))

                            for item in platform_expanded

                        ),

                    }

                )



        for platform in counted_platforms:

            if platform not in PLATFORM_NAMES:

                continue



            platform_deduction_orders = aggregate_orders(

                platform_orders.get(platform, [])

            )



            if not platform_deduction_orders:

                continue



            deduction_inventory = expand_inventory(

                platform_deduction_orders,

                db

            )

            stock_deduction_report = generate_daily_report(

                deduction_inventory,

                db

            )



            stock_result = deduct_lsds_stock_inventory(

                platform_deduction_orders,

                stock_deduction_report,

                db

            )

            _log_inventory_deductions(

                db,

                report_date,

                platform,

                "stock",

                stock_result.get("deductions", []),

            )



            sticker_result = deduct_sticker_inventory(

                platform_deduction_orders,

                stock_deduction_report,

                db

            )

            _log_inventory_deductions(

                db,

                report_date,

                platform,

                "sticker",

                sticker_result.get("deductions", []),

            )



            return_result = deduct_return_inventory(

                deduction_inventory,

                db

            )

            _log_inventory_deductions(

                db,

                report_date,

                platform,

                "return",

                return_result.get("deductions", []),

            )



            platform_order_count = _unique_order_count(

                platform_orders.get(platform, [])

            )

            platform_piece_count = sum(

                _safe_int(item.get("qty", 0))

                for item in deduction_inventory

            )

            packing_result = deduct_packing_inventory_for_platform(

                db,

                platform,

                platform_order_count,

                platform_piece_count,

            )

            _log_inventory_deductions(

                db,

                report_date,

                platform,

                "packing",

                packing_result.get("deductions", []),

            )



        db.commit()
        write_inventory_snapshot("stock", _serialize_stock_inventory_rows(db))
        write_inventory_snapshot("return", _serialize_return_inventory_rows(db))
        write_inventory_snapshot("sticker", _serialize_sticker_inventory_rows(db))

    except Exception:

        db.rollback()

        raise

    finally:

        db.close()



       # =====================

    # EXPORT EXCEL

    # =====================



    from openpyxl import Workbook

    from openpyxl.utils import get_column_letter

    from openpyxl.styles import (

        Font,

        Alignment,

        PatternFill,

        Border,

        Side

    )



    now = datetime.now()

    timestamp = now.strftime(

        "%d-%m-%Y_%H%M"

    )

    today_date = now.strftime("%d-%m-%Y")

    generated_at = now.strftime(

        "%d-%m-%Y %I:%M %p"

    )



    output_file = (

        os.path.join(

            UPLOAD_FOLDER,

            f"final_report_{timestamp}.xlsx"

        )

    )



    wb = Workbook()



    ws = wb.active



    final_report = sorted(



    final_report,



    key=lambda x: (



        x.get("style", ""),



        x.get("color", "")

    )

    )



    size_order = ["XS", "S", "M", "L", "XL", "2XL"]

    available_sizes = []

    seen_sizes = set()



    for item in final_report:

        size_value = str(item.get("size", "")).upper().strip()

        if size_value and size_value in size_order and size_value not in seen_sizes:

            available_sizes.append(size_value)

            seen_sizes.add(size_value)



    available_sizes = [

        size for size in size_order

        if size in available_sizes

    ]

    columns_per_size = 3 if include_detail_columns else 1

    total_columns = 2 + (len(available_sizes) * columns_per_size)

    summary_columns = 3 if include_order_summary else 0

    sheet_columns = max(total_columns, summary_columns, 2)

    last_column = get_column_letter(max(total_columns, 2))

    header_start_row = 2



    if include_order_summary:

        header_start_row = 6 + len(order_summary_rows)



    header_end_row = (

        header_start_row + 1

        if include_detail_columns

        else header_start_row

    )

    data_start_row = header_end_row + 1



    ws.title = "Final Report"

    ws.merge_cells(f"A1:{get_column_letter(sheet_columns)}1")

    ws["A1"] = f"Generated At: {generated_at}"



    if include_order_summary:

        ws.merge_cells(f"A2:{get_column_letter(sheet_columns)}2")

        ws["A2"] = "Order Summary"

        ws.append([

            "Platform",

            "Total Orders",

            "Total Piece Qty",

        ])



        for summary in order_summary_rows:

            ws.append([

                summary["platform"],

                summary["total_orders"],

                summary["piece_qty"],

            ])



        ws.append([

            "Total",

            sum(item["total_orders"] for item in order_summary_rows),

            sum(item["piece_qty"] for item in order_summary_rows),

        ])

        ws.append([""] * sheet_columns)



    # =====================

    # HEADER DESIGN

    # =====================



    headers = ["Style", "Color"]

    sub_headers = ["", ""]



    for size in available_sizes:

        if include_detail_columns:

            headers.extend([size, "", ""])

            sub_headers.extend([

                "Total Order",

                "Return Stock",

                "Need to Print"

            ])

        else:

            headers.append(size)



    ws.append(headers)



    if include_detail_columns:

        ws.append(sub_headers)



        ws.merge_cells(f"A{header_start_row}:A{header_end_row}")

        ws.merge_cells(f"B{header_start_row}:B{header_end_row}")



        for index, _ in enumerate(available_sizes):

            start_col = 3 + (index * columns_per_size)

            end_col = start_col + columns_per_size - 1

            ws.merge_cells(

                (

                    f"{get_column_letter(start_col)}{header_start_row}:"

                    f"{get_column_letter(end_col)}{header_start_row}"

                )

            )



    # =====================

    # STYLES

    # =====================



    header_fill = PatternFill(

        start_color="213A69",

        end_color="213A69",

        fill_type="solid"

    )



    header_font = Font(

        bold=True,

        color="FFFFFF",

        size=14,

    )



    data_font = Font(size=12 )



    center_align = Alignment(

        horizontal="center",

        vertical="center",

        wrap_text=True,

    )



    thin_border = Border(

        left=Side(style="thin", color="FFFFFF"),

        right=Side(style="thin", color="FFFFFF"),

        bottom=Side(style="thin", color="FFFFFF")

    )



    timestamp_fill = PatternFill(

        start_color="E5E7EB",

        end_color="E5E7EB",

        fill_type="solid"

    )



    subtotal_fill = PatternFill(

        start_color="213A69",

        end_color="213A69",

        fill_type="solid"

    )



    table_light_fill = PatternFill(

        start_color="7A9BD6",

        end_color="7A9BD6",

        fill_type="solid"

    )



    table_lighter_fill = PatternFill(

        start_color="ABC4EA",

        end_color="ABC4EA",

        fill_type="solid"

    )



    subtotal_font = Font(

        bold=True,

        color="FFFFFF",

        size=12,

    )



    summary_title_fill = PatternFill(

        start_color="FDE68A",

        end_color="FDE68A",

        fill_type="solid"

    )



    summary_header_fill = PatternFill(

        start_color="FEF3C7",

        end_color="FEF3C7",

        fill_type="solid"

    )



    ws["A1"].fill = timestamp_fill

    ws["A1"].font = Font(bold=True, color="111827", size=12)

    ws["A1"].alignment = Alignment(

        horizontal="center",

        vertical="center",

    )

    ws.row_dimensions[1].height = 18



    if include_order_summary:

        ws["A2"].fill = summary_title_fill

        ws["A2"].font = Font(bold=True, color="92400E", size=13)

        ws["A2"].alignment = center_align



        summary_end_row = 3 + len(order_summary_rows) + 1



        for row in ws.iter_rows(

            min_row=3,

            max_row=summary_end_row,

            min_col=1,

            max_col=3

        ):

            for cell in row:

                cell.alignment = center_align

                cell.border = thin_border



                if cell.row == 3:

                    cell.fill = summary_header_fill

                    cell.font = Font(bold=True, color="78350F", size=11)

                elif cell.row == summary_end_row:

                    cell.fill = subtotal_fill

                    cell.font = subtotal_font

                else:

                    cell.font = Font(size=11)



    for row in ws.iter_rows(

        min_row=header_start_row,

        max_row=header_end_row,

        min_col=1,

        max_col=max(total_columns, 2)

    ):



        for cell in row:



            cell.fill = header_fill



            cell.font = header_font



            cell.alignment = center_align





    # =====================

    # GROUP DATA

    # =====================



    grouped_data = {}



    for item in final_report:



        key = (

            item["style"],

            item["color"]

        )



        if key not in grouped_data:



            grouped_data[key] = {}



        grouped_data[key][

            item["size"].upper()

        ] = item



    # =====================

    # WRITE DATA

    # =====================



    subtotal_rows = set()

    blank_rows = set()

    current_style = None

    style_totals = None



    def add_style_subtotal_row(style_name):

        subtotal_row = [

            "Total",

            ""

        ]



        for size in available_sizes:

            totals = style_totals.get(

                size,

                {

                    "total_order_qty": 0,

                    "return_qty": 0,

                    "stock_qty": 0,

                }

            )



            subtotal_row.extend([

                (

                    " "

                    if totals["total_order_qty"] == 0

                    else totals["total_order_qty"]

                )

            ])



            if include_detail_columns:

                subtotal_row.extend([

                    (

                        " "

                        if totals["return_qty"] == 0

                        else totals["return_qty"]

                    ),

                    (

                        " "

                        if totals["stock_qty"] == 0

                        else totals["stock_qty"]

                    ),

                ])



        ws.append(subtotal_row)



        subtotal_row_number = ws.max_row

        subtotal_rows.add(subtotal_row_number)



        for cell in ws[subtotal_row_number]:

            cell.fill = subtotal_fill

            cell.font = subtotal_font

            cell.alignment = center_align



    for (style, color), sizes in grouped_data.items():



        if current_style is None:

            current_style = style

            style_totals = {

                size: {

                    "total_order_qty": 0,

                    "return_qty": 0,

                    "stock_qty": 0,

                }

                for size in available_sizes

            }



        elif style != current_style:

            add_style_subtotal_row(current_style)        



            current_style = style

            style_totals = {

                size: {

                    "total_order_qty": 0,

                    "return_qty": 0,

                    "stock_qty": 0,

                }

                for size in available_sizes

            }



        row = [

            style,

            color

        ]



        for size in available_sizes:



            if size in sizes:



                data = sizes[size]



                total_qty = data.get("total_order_qty", 0)

                return_qty = data.get(

                    "used_return_qty",

                    data.get("return_inventory", 0)

                )

                stock_qty = data.get(

                    "need_from_stock",

                    data.get("stock_inventory", 0)

                )



                style_totals[size]["total_order_qty"] += _safe_int(

                    total_qty or 0

                )

                style_totals[size]["return_qty"] += _safe_int(

                    return_qty or 0

                )

                style_totals[size]["stock_qty"] += _safe_int(

                    stock_qty or 0

                )



                row.extend([

                    (

                        "-"

                        if total_qty == 0

                        else total_qty

                    )

                ])



                if include_detail_columns:

                    row.extend([

                        (

                            "-"

                            if return_qty == 0

                            else return_qty

                        ),



                        (

                            "-"

                            if stock_qty == 0

                            else stock_qty

                        )

                    ])



            else:

                row.extend(["-"] * columns_per_size)



        ws.append(row)



    if current_style is not None:

        add_style_subtotal_row(current_style)



    # =====================

    # COLUMN WIDTH (A4 portrait)

    # =====================



    if total_columns <= 11:



        style_w, color_w, data_w = (

            18,

            16,

            8

        )



    elif total_columns <= 14:



        style_w, color_w, data_w = (

            16,

            14,

            8

        )



    else:



        style_w, color_w, data_w = (

            14,

            12,

            6

        )



    ws.column_dimensions["A"].width = style_w

    ws.column_dimensions["B"].width = color_w



    for col in range(

        3,

        sheet_columns + 1

    ):



        ws.column_dimensions[

            get_column_letter(col)

        ].width = data_w



    # =====================

    # PAGE SETUP

    # =====================



    ws.sheet_properties.pageSetUpPr.fitToPage = True



    ws.page_setup.orientation = "portrait"



    ws.page_setup.paperSize = 9



    ws.page_setup.fitToWidth = 1



    ws.page_setup.fitToHeight = False



    # =====================

    # DATA CELL STYLE

    # =====================



    for row in ws.iter_rows(



        min_row=data_start_row,



        max_row=ws.max_row,



        min_col=1,



        max_col=ws.max_column

    ):



        for cell in row:



            if cell.row in blank_rows:

                continue



            cell.alignment = center_align

            cell.border = thin_border



            if cell.row in subtotal_rows:

                cell.font = subtotal_font

                cell.fill = subtotal_fill

                continue



            cell.font = data_font

            if (cell.row - data_start_row) % 2 == 0:

                cell.fill = table_light_fill

            else:

                cell.fill = table_lighter_fill



    wb.save(output_file)



    return FileResponse(

        output_file,

        media_type=(

            "application/vnd.openxmlformats-"

            "officedocument.spreadsheetml.sheet"

        ),

        filename=(

            f"final_report_{today_date}.xlsx"

        )

    )





@router.post("/confirm-final-report")

def confirm_final_report(



    flipkart_file: UploadFile = File(None),

    flipkart_dispatch_period: Optional[str] = Form(None),



    amazon_file: UploadFile = File(None),



    ajio_file: UploadFile = File(None),



    meesho_file: UploadFile = File(None),



    myntra_file: UploadFile = File(None),



    flipkart_warehouse_file: UploadFile = File(None),

):



    all_orders, platform_orders = _collect_marketplace_orders(

        flipkart_file,

        flipkart_dispatch_period,

        amazon_file,

        ajio_file,

        meesho_file,

        myntra_file,

        flipkart_warehouse_file,

    )



    if not platform_orders:

        raise HTTPException(

            status_code=400,

            detail="Upload at least one marketplace order file.",

        )



    db: Session = SessionLocal()

    report_date = datetime.now().date()

    platform_files = {

        "Flipkart": flipkart_file,

        "Amazon": amazon_file,

        "Ajio": ajio_file,

        "Meesho": meesho_file,

        "Myntra": myntra_file,

        "Flipkart Warehouse": flipkart_warehouse_file,

    }



    try:

        _raise_if_unknown_platform_skus(

            db,

            platform_orders,

        )



        new_orders, counted_platforms, skipped_platforms = (

            _orders_from_new_platform_uploads(

                db,

                report_date,

                platform_orders,

                platform_files,

                flipkart_dispatch_period,

            )

        )



        if not new_orders:

            return {

                "message": (

                    "No new uploads to confirm. "

                    "Duplicate files are skipped."

                ),

                "lines_updated": 0,

                "total_qty_deducted": 0,

                "deductions": [],

                "counted_platforms": counted_platforms,

                "skipped_duplicate_platforms": skipped_platforms,

            }



        inventory_orders = [

            order

            for platform_name, orders in platform_orders.items()

            if platform_name in counted_platforms

            and platform_name in PLATFORM_NAMES

            for order in orders

        ]



        counted_reporting_platforms, skipped_reporting_platforms = (

            _save_new_platform_sales(

                db,

                report_date,

                {

                    platform_name: orders

                    for platform_name, orders in platform_orders.items()

                    if platform_name in counted_platforms

                },

                platform_files,

                flipkart_dispatch_period,

            )

        )



        if not inventory_orders:

            db.commit()

            return {

                "message": (

                    "Reporting-only uploads were saved. "

                    "No inventory quantities were deducted."

                ),

                "lines_updated": 0,

                "total_qty_deducted": 0,

                "deductions": [],

                "counted_platforms": counted_reporting_platforms,

                "skipped_duplicate_platforms": (

                    skipped_platforms + skipped_reporting_platforms

                ),

            }



        aggregated_orders = aggregate_orders(inventory_orders)



        expanded_inventory = expand_inventory(

            aggregated_orders,

            db,

        )



        result = deduct_return_inventory(

            expanded_inventory,

            db,

        )

        db.commit()

        result["counted_platforms"] = counted_platforms

        result["skipped_duplicate_platforms"] = skipped_platforms

        return result

    finally:

        db.close()





def _clean_extracted_sku(raw_value: str):

    value = str(raw_value or "").strip()

    value = re.sub(r"\s+", " ", value)

    value = re.split(

        r"\s{2,}|\b(?:qty|quantity|hsn|asin|fnsku|tax|description)\b",

        value,

        flags=re.IGNORECASE,

    )[0].strip(" :-#|")

    value = re.sub(r"\s+", "", value)

    return value[:80]





def _looks_like_amazon_seller_sku(value: str):

    value = str(value or "").strip()



    if len(value) < 5:

        return False



    if not re.search(r"[-_]", value):

        return False



    if not re.search(r"[A-Za-z]", value) or not re.search(r"\d", value):

        return False



    return True





def _extract_amazon_invoice_sku(text: str):

    combined_text = re.sub(r"\s+", " ", str(text or ""))

    wrapped_sku_match = re.search(

        (

            r"\bB0[A-Z0-9]{8,}\s*\(\s*"

            r"([A-Z0-9][A-Z0-9._/\-()\s]{2,80}?)"

            r"\s*\)\s*(?:HSN\b|Î“Ã©â•£|Rs\.?|₹|$)"

        ),

        combined_text,

    )



    if wrapped_sku_match:

        sku = _clean_extracted_sku(wrapped_sku_match.group(1))

        if _looks_like_amazon_seller_sku(sku):

            return sku



    lines = [

        re.sub(r"\s+", " ", line).strip()

        for line in str(text or "").splitlines()

        if line.strip()

    ]



    sku_patterns = [

        r"\(\s*([A-Z0-9][A-Z0-9._/\-()]{2,80})\s*\)",

        r"\bMerchant\s+SKU\b\s*(?:No\.?|#|:|-)?\s*([A-Za-z0-9][A-Za-z0-9._/\- ]{1,80})",

        r"\bSeller\s+SKU\b\s*(?:No\.?|#|:|-)?\s*([A-Za-z0-9][A-Za-z0-9._/\- ]{1,80})",

        r"\bSKU\b\s*(?:No\.?|#|:|-)?\s*([A-Za-z0-9][A-Za-z0-9._/\- ]{1,80})",

    ]



    for index, line in enumerate(lines):

        has_sku_label = re.search(r"\bSKU\b", line, flags=re.IGNORECASE)

        has_parenthesized_code = re.search(

            r"\(\s*[A-Z0-9][A-Z0-9._/\-]{2,80}\s*\)",

            line,

        )



        if not has_sku_label and not has_parenthesized_code:

            continue



        for pattern in sku_patterns:

            match = re.search(pattern, line, flags=re.IGNORECASE)

            if match:

                sku = _clean_extracted_sku(match.group(1))

                if (

                    sku

                    and not re.fullmatch(r"sku", sku, flags=re.IGNORECASE)

                    and _looks_like_amazon_seller_sku(sku)

                ):

                    return sku



        if re.fullmatch(

            r"(?:Merchant\s+|Seller\s+)?SKU\s*:?",

            line,

            flags=re.IGNORECASE,

        ) and index + 1 < len(lines):

            sku = _clean_extracted_sku(lines[index + 1])

            if sku and _looks_like_amazon_seller_sku(sku):

                return sku



    return None





def _extract_amazon_invoice_qty(text: str):

    text = str(text or "")

    patterns = [

        r"\bQty\s*:?\s*(\d+)",

        r"\bQuantity\s*:?\s*(\d+)",

        r"\bQty\b\s+(\d+)",

    ]



    for pattern in patterns:

        match = re.search(pattern, text, flags=re.IGNORECASE)

        if match:

            return match.group(1)



    return "1"





def _extract_amazon_invoice_item_qty(text: str):

    text = str(text or "")

    patterns = [

        r"\bQty\s*:?\s*(\d+)",

        r"\bQuantity\s*:?\s*(\d+)",

        r"\bHSN\s*:?\s*\d+.*?(?:Î“Ã©â•£|Rs\.?|₹)\s*[\d,.]+\s+(\d+)\b",

        r"(?:Î“Ã©â•£|Rs\.?|₹)\s*[\d,.]+\s+(\d+)\s+(?:Î“Ã©â•£|Rs\.?|₹)",

    ]



    for pattern in patterns:

        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)

        if match:

            return match.group(1)



    return "1"





def _extract_amazon_invoice_items(text: str):

    lines = [

        re.sub(r"\s+", " ", line).strip()

        for line in str(text or "").splitlines()

        if line.strip()

    ]

    items = []

    seen_skus = set()



    sku_patterns = [

        r"\(\s*([A-Z0-9][A-Z0-9._/()\s-]{2,80})\s*\)",

        r"^([A-Za-z0-9][A-Za-z0-9._/()-]{2,80})\s*\)$",

        r"\bMerchant\s+SKU\b\s*(?:No\.?|#|:|-)?\s*([A-Za-z0-9][A-Za-z0-9._/\- ]{1,80})",

        r"\bSeller\s+SKU\b\s*(?:No\.?|#|:|-)?\s*([A-Za-z0-9][A-Za-z0-9._/\- ]{1,80})",

        r"\bSKU\b\s*(?:No\.?|#|:|-)?\s*([A-Za-z0-9][A-Za-z0-9._/\- ]{1,80})",

    ]



    for index, line in enumerate(lines):

        candidate_skus = []



        for pattern in sku_patterns:

            for match in re.finditer(pattern, line, flags=re.IGNORECASE):

                candidate_skus.append(match.group(1))



        if re.fullmatch(

            r"(?:Merchant\s+|Seller\s+)?SKU\s*:?",

            line,

            flags=re.IGNORECASE,

        ) and index + 1 < len(lines):

            candidate_skus.append(lines[index + 1])



        for raw_sku in candidate_skus:

            sku = _clean_extracted_sku(raw_sku)

            normalized_sku = sku.upper()



            if (

                not sku

                or normalized_sku in seen_skus

                or re.fullmatch(r"sku", sku, flags=re.IGNORECASE)

                or not _looks_like_amazon_seller_sku(sku)

            ):

                continue



            item_text = "\n".join(lines[index:index + 8])

            items.append(

                {

                    "sku": sku,

                    "qty": _extract_amazon_invoice_item_qty(item_text),

                }

            )

            seen_skus.add(normalized_sku)



    if items:

        return items



    sku = _extract_amazon_invoice_sku(text)

    if sku:

        return [

            {

                "sku": sku,

                "qty": _extract_amazon_invoice_qty(text),

            }

        ]



    return []





def _extract_amazon_invoice_order_number(text: str):

    match = re.search(

        r"Order\s+Number\s*:?\s*([0-9-]+)",

        str(text or ""),

        flags=re.IGNORECASE,

    )



    return match.group(1) if match else ""





def _extract_amazon_invoice_description(text: str):

    lines = [

        re.sub(r"\s+", " ", line).strip()

        for line in str(text or "").splitlines()

        if line.strip()

    ]

    description_lines = []

    collecting = False



    for line in lines:

        if re.match(r"^1\s+", line):

            collecting = True

            line = re.sub(r"^1\s+", "", line).strip()



        if collecting:

            if re.search(r"\bHSN\s*:", line, flags=re.IGNORECASE):

                break



            if line:

                description_lines.append(line)



    if not description_lines:

        for line in lines:

            if "B0" in line or "(" in line:

                description_lines.append(line)



    description = " ".join(description_lines)

    return description[:700]





def _is_amazon_invoice_page(text: str):

    normalized_text = re.sub(r"\s+", " ", str(text or "")).strip()



    if not normalized_text:

        return False



    return bool(

        re.search(

            r"\bTax\s+Invoice\b|\bBill\s+of\s+Supply\b|\bCash\s+Memo\b",

            normalized_text,

            flags=re.IGNORECASE,

        )

    )





def _amazon_invoice_page_display(invoice_indices):

    page_numbers = [

        index + 1

        for index in invoice_indices

    ]



    if not page_numbers:

        return ""



    if len(page_numbers) == 1:

        return str(page_numbers[0])



    return f"{page_numbers[0]}-{page_numbers[-1]}"





def _iter_amazon_label_invoice_groups(reader):

    current_label_index = None

    current_invoice_indices = []



    for page_index, page in enumerate(reader.pages):

        text = page.extract_text() or ""



        if _is_amazon_invoice_page(text):

            if current_label_index is not None:

                current_invoice_indices.append(page_index)

            continue



        if current_label_index is not None and current_invoice_indices:

            yield current_label_index, current_invoice_indices



        current_label_index = page_index

        current_invoice_indices = []



    if current_label_index is not None and current_invoice_indices:

        yield current_label_index, current_invoice_indices





def _make_sku_overlay(width: float, height: float, sku_items):

    from io import BytesIO

    from reportlab.lib.colors import black

    from reportlab.pdfgen import canvas



    packet = BytesIO()

    overlay = canvas.Canvas(packet, pagesize=(float(width), float(height)))

    labels = [

        f"( {item.get('sku') or 'NOT FOUND'} ) (Qty: {item.get('qty') or '1'})"

        for item in sku_items

    ] or ["( NOT FOUND ) (Qty: 1)"]

    x = 50

    line_height = 20

    box_height = max(25.3, (line_height * len(labels)) + 6)

    y = 200 - box_height

    box_width = 520

    font_size = 17 if len(labels) == 1 else 14



    overlay.setStrokeColor(black)

    overlay.setLineWidth(1)

    overlay.rect(x, y, box_width, box_height, stroke=1, fill=0)

    overlay.setFillColor(black)

    overlay.setFont("Times-Bold", font_size)



    for index, label in enumerate(labels):

        text_y = y + box_height - 16 - (index * line_height)

        overlay.drawString(x + 5, text_y, label[:72])



    overlay.save()

    packet.seek(0)

    return packet





def _ajio_sku_layout(content_width: float, sku_items, label_text_left: float = 22.0):

    from reportlab.pdfbase.pdfmetrics import stringWidth



    MM = 2.835

    side_margin = 5 * MM

    bottom_margin = 5 * MM

    line_gap = 2

    font_name = "Times-Bold"

    font_size = 10

    min_font_size = 6

    text_x = side_margin + float(label_text_left)



    labels = [

        f"( {item.get('sku') or 'NOT FOUND'} ) (Qty: {item.get('qty') or '1'})"

        for item in (sku_items or [])

    ] or ["( NOT FOUND ) (Qty: 1)"]



    max_text_width = float(content_width) - float(label_text_left)



    for label in labels:

        size = font_size

        while (

            size > min_font_size

            and stringWidth(label, font_name, size) > max_text_width

        ):

            size -= 0.5

        font_size = min(font_size, size)



    line_height = font_size + line_gap

    label_count = len(labels)

    text_block_height = (

        ((label_count - 1) * line_height) + font_size

        if label_count > 1

        else font_size

    )

    strip_height = text_block_height + bottom_margin



    return {

        "labels": labels,

        "font_name": font_name,

        "font_size": font_size,

        "line_height": line_height,

        "side_margin": side_margin,

        "bottom_margin": bottom_margin,

        "strip_height": strip_height,

        "text_x": text_x,

        "label_text_left": float(label_text_left),

    }





def _make_ajio_sku_overlay(

    width: float,

    height: float,

    sku_items,

    layout=None,

):

    from io import BytesIO

    from reportlab.lib.colors import black

    from reportlab.pdfgen import canvas



    packet = BytesIO()

    page_width = float(width)

    page_height = float(height)

    overlay = canvas.Canvas(packet, pagesize=(page_width, page_height))



    layout = layout or _ajio_sku_layout(

        page_width - (2 * (5 * 2.835)),

        sku_items,

    )

    labels = layout["labels"]

    text_x = layout["text_x"]

    bottom_margin = layout["bottom_margin"]

    line_height = layout["line_height"]



    overlay.setFillColor(black)

    overlay.setFont(layout["font_name"], layout["font_size"])



    if len(labels) == 1:

        overlay.drawString(text_x, bottom_margin, labels[0])

    else:

        y = bottom_margin

        for label in reversed(labels):

            overlay.drawString(text_x, y, label)

            y += line_height



    overlay.save()

    packet.seek(0)



    return packet



def _build_amazon_label_cropper_pdf(

    input_path: str,

    output_path: str,

    manual_entries=None,

):

    from pypdf import PdfReader, PdfWriter



    reader = PdfReader(input_path)

    writer = PdfWriter()

    total_pages = len(reader.pages)



    if total_pages < 2:

        raise HTTPException(

            status_code=400,

            detail="Amazon PDF must contain label and invoice page pairs.",

        )



    missing_skus = []

    manual_entries = manual_entries or {}

    page_items = []



    for group_index, (label_index, invoice_indices) in enumerate(

        _iter_amazon_label_invoice_groups(reader),

        start=1,

    ):

        label_page = reader.pages[label_index]

        invoice_text = "\n".join(

            reader.pages[invoice_index].extract_text() or ""

            for invoice_index in invoice_indices

        )

        label_number = group_index

        invoice_items = _extract_amazon_invoice_items(invoice_text)

        qty = _extract_amazon_invoice_qty(invoice_text)



        if not invoice_items:

            manual_entry = manual_entries.get(str(label_number), {})

            manual_sku = str(manual_entry.get("sku", "")).strip()

            manual_qty = str(manual_entry.get("qty", qty or "1")).strip() or "1"



            if manual_sku:

                invoice_items = [

                    {

                        "sku": manual_sku,

                        "qty": manual_qty,

                    }

                ]



        if not invoice_items:

            missing_skus.append(

                {

                    "label_number": label_number,

                    "label_page": label_index + 1,

                    "invoice_page": _amazon_invoice_page_display(

                        invoice_indices

                    ),

                    "invoice_pages": [

                        invoice_index + 1

                        for invoice_index in invoice_indices

                    ],

                    "order_number": _extract_amazon_invoice_order_number(

                        invoice_text

                    ),

                    "description": _extract_amazon_invoice_description(

                        invoice_text

                    ),

                    "qty": qty or "1",

                }

            )

            continue



        page_items.append((label_page, invoice_items))



    if missing_skus:

        raise HTTPException(

            status_code=409,

            detail={

                "code": "LABEL_SKUS_MISSING",

                "message": "Some Amazon label SKUs could not be found.",

                "items": missing_skus,

            },

        )



    for label_page, invoice_items in page_items:

        width = float(label_page.mediabox.width)

        height = float(label_page.mediabox.height)

        overlay_reader = PdfReader(

            _make_sku_overlay(

                width,

                height,

                invoice_items,

            )

        )

        label_page.merge_page(overlay_reader.pages[0])

        writer.add_page(label_page)



    if len(writer.pages) == 0:

        raise HTTPException(

            status_code=400,

            detail="No shipping label pages were found in the uploaded PDF.",

        )



    with open(output_path, "wb") as output_file:

        writer.write(output_file)



    return {

        "label_count": len(writer.pages),

        "missing_skus": [],

    }





def _normalize_ajio_customer_name(value: str):

    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())





def _extract_ajio_invoice_consignment_no(text: str):

    match = re.search(r"Consignment\s+No\s*:\s*([A-Z0-9]+)", str(text or ""), flags=re.IGNORECASE)

    return match.group(1).strip() if match else ""



def _extract_ajio_label_shipment_no(text: str):

    match = re.search(r"Shipment#\s*:(.*?)(?:Return Address|Total)", str(text or ""), flags=re.IGNORECASE | re.DOTALL)

    if match:

        tokens = re.findall(r"[A-Z0-9]{8,25}", match.group(1), flags=re.IGNORECASE)

        if tokens:

            return max(tokens, key=len).upper()

    return ""



def _extract_ajio_label_customer_name(text: str):

    match = re.search(

        r"Ship\s+To\s*:\s*[\r\n]*([^\r\n]+)",

        str(text or ""),

        flags=re.IGNORECASE,

    )

    return match.group(1).strip() if match else ""





def _extract_ajio_invoice_customer_name(text: str):

    match = re.search(

        r"(?:Recipient\s+Address|BILL\s+TO\s*/\s*SHIP\s+TO)\s*:\s*[\r\n]*([A-Za-z0-9 ]+?)(?=\s*Mobile\s+No|[\r\n]|$)",

        str(text or ""),

        flags=re.IGNORECASE,

    )



    return match.group(1).strip() if match else ""


def _extract_ajio_invoice_carrier_name(text: str):
    match = re.search(r"Carrier Name:\s*([A-Za-z0-9_]+)", str(text or ""), flags=re.IGNORECASE)
    if match:
        carrier = match.group(1).strip().upper()
        if carrier in ["XPRESSBEES", "DELHIVERY", "SHADOWFAX"]:
            return carrier
    return ""





def _normalize_ajio_invoice_sku_source(text: str):

    source = str(text or "").replace("\r", "\n")

    return re.sub(

        r"([A-Z0-9][A-Z0-9_.-]{2,})\s*\n\s*([A-Z0-9][A-Z0-9_.-]{2,})",

        r"\1\2",

        source,

        flags=re.IGNORECASE,

    )





def _looks_like_ajio_sku(value: str):

    value = str(value or "").strip()



    if len(value) < 5:

        return False



    if not re.search(r"[A-Za-z]", value) or not re.search(r"\d", value):

        return False



    if not re.fullmatch(r"[A-Z0-9][A-Z0-9_.-]*\d+P(?:C)?", value.upper()):

        return False



    return True





def _extract_ajio_sku_candidates(sku_source: str):

    pattern = re.compile(

        r"\b[A-Z0-9][A-Z0-9_.-]{3,}(?:_[A-Z0-9_.-]+)*\d+P(?:C)?\b",

        flags=re.IGNORECASE,

    )

    candidates = pattern.findall(str(sku_source or "").upper())

    return [

        candidate

        for candidate in dict.fromkeys(candidates)

        if _looks_like_ajio_sku(candidate)

    ]





def _pick_best_ajio_sku(candidates):

    if not candidates:

        return None



    return max(candidates, key=len)





def _extract_ajio_item_skus_from_block(block: str):

    normalized_block = _normalize_ajio_invoice_sku_source(block)

    return _extract_ajio_sku_candidates(normalized_block)





def _detect_ajio_label_text_left(label_path: str, page_index: int) -> float:

    import fitz



    doc = fitz.open(label_path)

    try:

        page = doc[page_index]

        positions = []



        for block in page.get_text("dict").get("blocks", []):

            if block.get("type") != 0:

                continue



            for line in block.get("lines", []):

                for span in line.get("spans", []):

                    text = str(span.get("text", "")).strip()

                    if text == "Ship To:":

                        positions.append(float(span["bbox"][0]))



        return min(positions) if positions else 22.0

    finally:

        doc.close()





def _extract_ajio_invoice_sku(text: str):

    item_match = re.search(

        r"Item\s+Details.*?Brand\s*:",

        str(text or ""),

        flags=re.IGNORECASE | re.DOTALL,

    )

    sku_source = item_match.group(0) if item_match else str(text or "")

    candidates = _extract_ajio_item_skus_from_block(sku_source)



    return _pick_best_ajio_sku(candidates)





def _extract_ajio_invoice_items(text: str):

    # B2B format check

    if "EAN-" in str(text or ""):

        eans_raw = re.findall(r"EAN-\s*([A-Z0-9_.\-\s]+?)(?=\bQty\b|\bTotal\b|\bProduct\b|\bHSN\b)", text, flags=re.IGNORECASE)

        eans = [re.sub(r"\s+", "", e).upper() for e in eans_raw]



        qty_match = re.search(r"Qty\s*\(In\s*Pcs\.?\)(.*?)(?:Unit\s*Price|Total\s*Price)", text, flags=re.IGNORECASE | re.DOTALL)

        qtys = []

        if qty_match:

            qtys = re.findall(r"\b(\d+)\b", qty_match.group(1))

        

        items = []

        seen_skus = set()

        for i, ean in enumerate(eans):

            sku = ean.upper()

            if sku in seen_skus:

                continue

            qty = qtys[i] if i < len(qtys) else "1"

            items.append({"sku": sku, "qty": qty})

            seen_skus.add(sku)

        if items:

            return items



    item_match = re.search(

        r"Item\s+Details.*?(?:Tax\s+Summary|Total\s+in\s+Words|E\.\s*&\s*O\.E\.)",

        str(text or ""),

        flags=re.IGNORECASE | re.DOTALL,

    )

    sku_source = item_match.group(0) if item_match else str(text or "")

    

    items = []

    seen_skus = set()



    if re.search(r"Brand\s*:", sku_source, flags=re.IGNORECASE):

        item_blocks = re.split(r"Brand\s*:", sku_source, flags=re.IGNORECASE)

        for block in item_blocks[:-1]:

            candidates = _extract_ajio_item_skus_from_block(block)

            sku = _pick_best_ajio_sku(candidates)



            if not sku or sku in seen_skus:

                continue



            items.append(

                {

                    "sku": sku,

                    "qty": "1",

                }

            )

            seen_skus.add(sku)

    else:

        # New format logic

        pattern = re.compile(r"\b([A-Z0-9][A-Z0-9_.-]{3,}(?:_[A-Z0-9_.-]+)*\d+P(?:C)?)\b", flags=re.IGNORECASE)

        lines = str(text or "").splitlines()

        for i, line in enumerate(lines):

            for candidate in pattern.findall(line):

                if _looks_like_ajio_sku(candidate):

                    sku = candidate.upper()

                    if sku in seen_skus:

                        continue

                    

                    qty = "1"

                    if i + 1 < len(lines):

                        next_line = lines[i+1]

                        qty_match = re.search(r"\b\d{8}\b\s+(\d+(?:\.\d+)?)\s+", next_line)

                        if qty_match:

                            qty = str(int(float(qty_match.group(1))))

                    

                    items.append({"sku": sku, "qty": qty})

                    seen_skus.add(sku)



    if items:

        return items



    sku = _extract_ajio_invoice_sku(text)

    if sku:

        return [

            {

                "sku": sku,

                "qty": _extract_ajio_invoice_qty(text),

            }

        ]



    return []





def _extract_ajio_invoice_qty(text: str):

    match = re.search(

        r"\b61091000\b\s+\b999799\b\s+(\d+)\b",

        str(text or ""),

        flags=re.IGNORECASE,

    )



    return match.group(1) if match else "1"





def _render_ajio_cropped_label_page(

    label_path: str,

    page_index: int,

    sku_items,

    layout=None,

) -> bytes:

    from io import BytesIO



    import fitz



    source = fitz.open(label_path)

    try:

        source_page = source[page_index]

        original_width = float(source_page.rect.width)

        label_text_left = _detect_ajio_label_text_left(label_path, page_index)

        layout = layout or _ajio_sku_layout(

            original_width,

            sku_items,

            label_text_left=label_text_left,

        )

        side_margin = layout["side_margin"]

        strip_height = layout["strip_height"]

        original_height = float(source_page.rect.height)



        output = fitz.open()

        try:

            new_page = output.new_page(

                width=original_width + (2 * side_margin),

                height=original_height + strip_height,

            )

            label_rect = fitz.Rect(

                side_margin,

                0,

                side_margin + original_width,

                original_height,

            )

            new_page.show_pdf_page(label_rect, source, page_index)



            overlay_packet = _make_ajio_sku_overlay(

                float(new_page.rect.width),

                float(new_page.rect.height),

                sku_items,

                layout=layout,

            )

            overlay_doc = fitz.open("pdf", overlay_packet.read())

            try:

                new_page.show_pdf_page(new_page.rect, overlay_doc, 0)

            finally:

                overlay_doc.close()



            return output.tobytes()

        finally:

            output.close()

    finally:

        source.close()



def _build_ajio_label_cropper_pdf(

    label_input_paths,

    invoice_input_paths,

    output_path: str,

):

    from io import BytesIO



    from pypdf import PdfReader, PdfWriter



    import fitz
    fitz_writer = fitz.open()
    invoice_items_by_customer = {}
    missing_invoices = []
    zone_counts = Counter()



    for invoice_path in invoice_input_paths:

        invoice_reader = PdfReader(invoice_path)



        for invoice_page_number, invoice_page in enumerate(

            invoice_reader.pages,

            start=1,

        ):

            invoice_text = invoice_page.extract_text() or ""

            consignment_no = _extract_ajio_invoice_consignment_no(invoice_text)

            customer_name = _extract_ajio_invoice_customer_name(invoice_text)

            

            if consignment_no:

                customer_key = consignment_no

            else:

                customer_key = _normalize_ajio_customer_name(customer_name)



            invoice_items = _extract_ajio_invoice_items(invoice_text)
            carrier_name = _extract_ajio_invoice_carrier_name(invoice_text)

            if not customer_key or not invoice_items:
                continue

            invoice_items_by_customer.setdefault(customer_key, []).append(
                {
                    "items": invoice_items,
                    "customer_name": customer_name or consignment_no,
                    "carrier_name": carrier_name,

                    "invoice_page": invoice_page_number,

                    "invoice_file": os.path.basename(invoice_path),

                }

            )



    for label_path in label_input_paths:

        label_reader = PdfReader(label_path)



        for label_page_number, label_page in enumerate(

            label_reader.pages,

            start=1,

        ):

            label_text = label_page.extract_text() or ""

            

            shipment_no = _extract_ajio_label_shipment_no(label_text)

            customer_name = _extract_ajio_label_customer_name(label_text)



            if shipment_no and shipment_no in invoice_items_by_customer:

                customer_key = shipment_no

            else:

                customer_key = _normalize_ajio_customer_name(customer_name)



            invoice_matches = invoice_items_by_customer.get(customer_key, [])



            if not invoice_matches:

                missing_invoices.append(

                    {

                        "label_file": os.path.basename(label_path),

                        "label_page": label_page_number,

                        "customer_name": shipment_no or customer_name or "Unknown",

                    }

                )

                continue



            invoice_item = invoice_matches.pop(0)
            invoice_items = invoice_item["items"]
            
            carrier = invoice_item.get("carrier_name", "")
            if carrier in ["XPRESSBEES", "DELHIVERY", "SHADOWFAX"]:
                zone_counts[carrier] += 1
                
            width = float(label_page.mediabox.width)

            label_text_left = _detect_ajio_label_text_left(

                label_path,

                label_page_number - 1,

            )

            layout = _ajio_sku_layout(

                width,

                invoice_items,

                label_text_left=label_text_left,

            )

            cropped_page_bytes = _render_ajio_cropped_label_page(

                label_path,

                label_page_number - 1,

                invoice_items,

                layout=layout,

            )

            cropped_doc = fitz.open("pdf", cropped_page_bytes)

            fitz_writer.insert_pdf(cropped_doc)



    if missing_invoices:

        raise HTTPException(

            status_code=400,

            detail={

                "code": "AJIO_INVOICES_MISSING",

                "message": (

                    "Some AJIO shipping labels could not be matched to "

                    "customer invoices."

                ),

                "items": missing_invoices,

            },

        )



    if len(fitz_writer) == 0:

        raise HTTPException(

            status_code=400,

            detail="No AJIO shipping label pages were found.",

        )



    fitz_writer.save(output_path)

    return {
        "label_count": len(fitz_writer),
        "zone_counts": zone_counts,
        "missing_pages": [],
        "report_date": datetime.now().date(),
    }





def _transform_point(matrix, x_value, y_value):

    a, b, c, d, e, f = matrix

    return (

        a * x_value + c * y_value + e,

        b * x_value + d * y_value + f,

    )





def _multiply_pdf_matrix(left, right):

    a, b, c, d, e, f = left

    g, h, i, j, k, l = right

    return [

        a * g + b * i,

        a * h + b * j,

        c * g + d * i,

        c * h + d * j,

        e * g + f * i + k,

        e * h + f * j + l,

    ]





def _flipkart_page_horizontal_segments(page):

    from pypdf.generic import ContentStream



    content = page.get_contents()

    if content is None:

        return []



    stream = ContentStream(content, page.pdf)

    matrix = [1, 0, 0, 1, 0, 0]

    stack = []

    dash = []

    current = None

    segments = []



    for operands, operator in stream.operations:

        if operator == b"q":

            stack.append((matrix[:], dash[:]))

            continue



        if operator == b"Q":

            if stack:

                matrix, dash = stack.pop()

            continue



        if operator == b"cm":

            matrix = _multiply_pdf_matrix(

                matrix,

                [float(value) for value in operands],

            )

            continue



        if operator == b"d":

            dash = list(operands[0]) if operands else []

            continue



        if operator == b"m":

            current = _transform_point(

                matrix,

                float(operands[0]),

                float(operands[1]),

            )

            continue



        if operator != b"l" or current is None:

            if operator in (b"h", b"n", b"S", b"s", b"f", b"F", b"f*"):

                current = None

            continue



        next_point = _transform_point(

            matrix,

            float(operands[0]),

            float(operands[1]),

        )

        x1, y1 = current

        x2, y2 = next_point



        if abs(y1 - y2) < 1 and abs(x2 - x1) > 20:

            segments.append(

                {

                    "x1": min(x1, x2),

                    "x2": max(x1, x2),

                    "y": (y1 + y2) / 2,

                    "dash": dash[:],

                }

            )



        current = next_point



    return segments





def _detect_flipkart_separator_y(page):

    width = float(page.mediabox.width)

    height = float(page.mediabox.height)

    segments = _flipkart_page_horizontal_segments(page)

    dashed_segments = [

        segment

        for segment in segments

        if segment["dash"]

        and (segment["x2"] - segment["x1"]) >= width * 0.65

        and height * 0.35 <= segment["y"] <= height * 0.65

    ]



    if dashed_segments:

        return max(

            dashed_segments,

            key=lambda segment: segment["x2"] - segment["x1"],

        )["y"]



    full_width_segments = [

        segment

        for segment in segments

        if (segment["x2"] - segment["x1"]) >= width * 0.75

        and height * 0.35 <= segment["y"] <= height * 0.65

    ]



    if full_width_segments:

        return max(

            full_width_segments,

            key=lambda segment: segment["y"],

        )["y"]



    raise HTTPException(

        status_code=400,

        detail="Could not detect the Flipkart label separator line.",

    )





def _detect_flipkart_label_bounds(page, separator_y: float):

    from collections import Counter



    width = float(page.mediabox.width)

    height = float(page.mediabox.height)

    segments = _flipkart_page_horizontal_segments(page)

    label_segments = [

        segment

        for segment in segments

        if separator_y + 2 <= segment["y"] <= height - 20

        and 160 <= (segment["x2"] - segment["x1"]) <= 260

    ]



    if not label_segments:

        return 0, separator_y, width, height



    endpoint_pairs = Counter(

        (

            round(segment["x1"] * 4) / 4,

            round(segment["x2"] * 4) / 4,

        )

        for segment in label_segments

    )

    left, right = endpoint_pairs.most_common(1)[0][0]

    top = max(

        segment["y"]

        for segment in label_segments

        if abs(segment["x1"] - left) < 1

        and abs(segment["x2"] - right) < 1

    )



    return (

        max(0, left - 26.25),

        max(0, separator_y + 3.5),

        min(width, right + 26.5),

        min(height, top + 7.25),

    )





def _detect_flipkart_crop_box_fast(page):

    from collections import Counter

    from pypdf.generic import ContentStream



    width = float(page.mediabox.width)

    height = float(page.mediabox.height)

    content = page.get_contents()



    if content is None:

        raise HTTPException(

            status_code=400,

            detail="Could not read the Flipkart label page content.",

        )



    stream = ContentStream(content, page.pdf)

    matrix = [1, 0, 0, 1, 0, 0]

    stack = []

    dash = []

    current = None

    endpoint_pairs = Counter()

    endpoint_tops = {}

    dashed_separator = None

    full_width_separators = []



    for index, (operands, operator) in enumerate(stream.operations):

        if operator == b"q":

            stack.append((matrix[:], dash[:]))

            continue



        if operator == b"Q":

            if stack:

                matrix, dash = stack.pop()

            continue



        if operator == b"cm":

            matrix = _multiply_pdf_matrix(

                matrix,

                [float(value) for value in operands],

            )

            continue



        if operator == b"d":

            dash = list(operands[0]) if operands else []

            continue



        if operator == b"m":

            current = _transform_point(

                matrix,

                float(operands[0]),

                float(operands[1]),

            )

            continue



        if operator != b"l" or current is None:

            if operator in (b"h", b"n", b"S", b"s", b"f", b"F", b"f*"):

                current = None

            continue



        next_point = _transform_point(

            matrix,

            float(operands[0]),

            float(operands[1]),

        )

        x1, y1 = current

        x2, y2 = next_point

        current = next_point



        if abs(y1 - y2) >= 1:

            continue



        left = min(x1, x2)

        right = max(x1, x2)

        y_position = (y1 + y2) / 2

        line_width = right - left



        if height * 0.35 <= y_position <= height * 0.65:

            if dash and line_width >= width * 0.65:

                dashed_separator = y_position

            elif line_width >= width * 0.75:

                full_width_separators.append(y_position)



        if y_position >= height * 0.5 and 160 <= line_width <= 260:

            endpoint_pair = (

                round(left * 4) / 4,

                round(right * 4) / 4,

            )

            endpoint_pairs[endpoint_pair] += 1

            endpoint_tops[endpoint_pair] = max(

                endpoint_tops.get(endpoint_pair, -1),

                y_position,

            )



        if (

            index > 3500

            and endpoint_pairs

            and (

                dashed_separator is not None

                or full_width_separators

            )

        ):

            break



    if dashed_separator is not None:

        separator_y = dashed_separator

    elif full_width_separators:

        separator_y = max(full_width_separators)

    else:

        raise HTTPException(

            status_code=400,

            detail="Could not detect the Flipkart label separator line.",

        )



    if not endpoint_pairs:

        raise HTTPException(

            status_code=400,

            detail="Could not detect the Flipkart label border.",

        )



    left, right = endpoint_pairs.most_common(1)[0][0]

    top = endpoint_tops[(left, right)]



    return (

        max(0, left - 26.25),

        max(0, separator_y + 3.5),

        min(width, right + 26.5),

        min(height, top + 7.25),

    )





def _build_flipkart_label_cropper_pdf(

    input_path: str,

    output_path: str,

    manual_zones=None,

):

    from pypdf import PdfReader, PdfWriter

    from pypdf.generic import RectangleObject



    reader = PdfReader(input_path)

    writer = PdfWriter()



    if len(reader.pages) == 0:

        raise HTTPException(

            status_code=400,

            detail="Flipkart PDF does not contain any pages.",

        )



    first_page = reader.pages[0]

    reusable_crop_box = _detect_flipkart_crop_box_fast(first_page)

    reusable_page_size = (

        round(float(first_page.mediabox.width), 2),

        round(float(first_page.mediabox.height), 2),

    )



    import fitz
    fitz_doc = fitz.open(input_path)
    for i, page in enumerate(reader.pages):
        page_size = (round(float(page.mediabox.width), 2), round(float(page.mediabox.height), 2))
        if page_size == reusable_page_size:
            crop_box = reusable_crop_box
        else:
            try:
                crop_box = _detect_flipkart_crop_box_fast(page)
            except HTTPException:
                separator_y = _detect_flipkart_separator_y(page)
                crop_box = _detect_flipkart_label_bounds(page, separator_y)
        cb = crop_box
        page_height = float(page.mediabox.height)
        fitz_page = fitz_doc[i]
        fitz_rect = fitz.Rect(cb[0], page_height - cb[3], cb[2], page_height - cb[1])
        fitz_page.set_cropbox(fitz_rect)
    fitz_doc.save(output_path)

    xml_path = _convert_pdf_to_xml(output_path)

    zone_result = _extract_zones_from_xml(

        xml_path,

        manual_zones=manual_zones,

    )



    if zone_result["label_count"] == 0:

        raise HTTPException(

            status_code=400,

            detail=(

                "Could not read Flipkart zones from this PDF. "

                "No 'Not for resale.' label markers were found."

            ),

        )



    if zone_result["missing_pages"]:

        return {

            "label_count": len(writer.pages),

            "cropped_pdf": output_path,

            "zone_counts": dict(zone_result["zone_counts"]),

            "missing_pages": zone_result["missing_pages"],

            "report_date": zone_result["report_date"],

        }

    import tempfile



    excel_path = os.path.join(

        tempfile.gettempdir(),

        f"zone_summary_{uuid.uuid4().hex}.xlsx"

    )



    _generate_zone_summary_excel(

        zone_result["zone_counts"],

        excel_path,

    )



    return {

        "label_count": len(writer.pages),

        "cropped_pdf": output_path,

        "zone_summary_excel": excel_path,

        "zone_counts": dict(zone_result["zone_counts"]),

        "missing_pages": zone_result["missing_pages"],

        "report_date": zone_result["report_date"],

    }



def _convert_pdf_to_xml(pdf_path: str) -> str:

    """

    Convert a PDF into XML using Poppler's pdftohtml.



    Returns:

        Path to the generated XML file.

    """



    temp_dir = tempfile.mkdtemp()



    output_base = os.path.join(

        temp_dir,

        "flipkart_labels"

    )

    hardcoded_poppler = r"F:\Inventory-Management\poppler-26.02.0\Library\bin\pdftohtml.exe"
    if os.path.exists(hardcoded_poppler):
        POPPLER_EXE = hardcoded_poppler
    else:
        import shutil
        if shutil.which("pdftohtml"):
            POPPLER_EXE = "pdftohtml"
        else:
            POPPLER_EXE = "pdftohtml" # Will fallback to FileNotFoundError

    try:

        subprocess.run(

            [

                POPPLER_EXE,

                "-xml",

                "-nodrm",

                "-hidden",

                "-enc",

                "UTF-8",

                pdf_path,

                output_base,

            ],

            check=True,

            stdout=subprocess.DEVNULL,

            stderr=subprocess.DEVNULL,

        )



    except FileNotFoundError:

        raise HTTPException(

            status_code=500,

            detail=(

                "Poppler (pdftohtml) is not installed "

                "or is not available in PATH."

            ),

        )



    except subprocess.CalledProcessError:

        raise HTTPException(

            status_code=500,

            detail="Failed to convert PDF to XML.",

        )



    xml_path = output_base + ".xml"



    if not os.path.exists(xml_path):

        raise HTTPException(

            status_code=500,

            detail="XML file was not generated.",

        )



    return xml_path



INVALID_FLIPKART_ZONE_TOKENS = {
    "AT",
    "AWB",
    "COD",
    "CPD",
    "CUSTOMER",
    "DATE",
    "DESCRIPTION",
    "GST",
    "GSTIN",
    "HBD",
    "HRS",
    "IGST",
    "INVOICE",
    "NAME",
    "ORDERED",
    "PACKAGING",
    "PREPAID",
    "PRINTED",
    "QTY",
    "SHIPPING",
    "SKU",
    "SOLD",
    "STD",
    "TAX",
    "TRANSPARENT",
    "USE",
    "ZONE",
}





def _normalize_flipkart_zone(value: str):
    value = re.sub(r"\s+", " ", str(value or "").strip().upper())
    match = re.match(r"^([A-Z0-9]{1,10})\b", value)

    if not match:
        return ""

    zone = match.group(1)

    if zone.isdigit():
        return ""

    if len(zone) == 1 and not zone.isalpha():
        return ""

    if zone in INVALID_FLIPKART_ZONE_TOKENS:
        return ""

    return zone





def _parse_flipkart_hbd_date(value: str):

    match = re.search(

        r"\bHBD\s*:\s*(\d{1,2})\s*[-/]\s*(\d{1,2})\b",

        value or "",

        flags=re.IGNORECASE,

    )



    if not match:

        return None



    day, month = match.groups()



    try:

        return date(

            datetime.now().year,

            int(month),

            int(day),

        )

    except ValueError:

        return None





def _extract_flipkart_zone_after_marker(lines, marker_index: int):
    search_range = 10
    
    for i in range(marker_index + 1, min(len(lines), marker_index + search_range)):
        line = lines[i].strip()
        packaging_prefix = "use transparent packaging"
        if line.lower().startswith(packaging_prefix):
            zone_text = line[len("Use Transparent Packaging"):].strip()
            zone = _normalize_flipkart_zone(zone_text)
            if zone: return zone
        else:
            zone = _normalize_flipkart_zone(line)
            if zone: return zone
            
    for i in range(marker_index - 1, max(-1, marker_index - search_range), -1):
        line = lines[i].strip()
        packaging_prefix = "use transparent packaging"
        if line.lower().startswith(packaging_prefix):
            zone_text = line[len("Use Transparent Packaging"):].strip()
            zone = _normalize_flipkart_zone(zone_text)
            if zone: return zone
        else:
            zone = _normalize_flipkart_zone(line)
            if zone: return zone

    return ""





def _flipkart_xml_page_values(page):

    line_nodes = page.findall("line")



    if line_nodes:

        return [

            (line.text or "").strip()

            for line in line_nodes

        ]



    text_items = []



    for node in page.findall("text"):

        text = "".join(node.itertext()).strip()



        if not text:

            continue



        try:

            top = float(node.attrib.get("top", 0))

            left = float(node.attrib.get("left", 0))

        except ValueError:

            top = 0

            left = 0



        text_items.append((top, left, text))



    return [

        text

        for _, _, text in sorted(text_items)

    ]





def _extract_flipkart_first_label_hbd_date(root):

    for page in root.findall("page"):

        lines = _flipkart_xml_page_values(page)



        for text in lines:

            if text.strip().lower() == "not for resale.":

                return None



            hbd_date = _parse_flipkart_hbd_date(text)

            if hbd_date:

                return hbd_date



    return None





def _extract_flipkart_zone_from_text_segment(segment):

    packaging_index = None



    for index, text in enumerate(segment):

        if text.strip().lower().startswith("use transparent packaging"):

            packaging_index = index

            same_item_zone = _normalize_flipkart_zone(

                text[len("Use Transparent Packaging"):]

            )



            if same_item_zone:

                return same_item_zone



            break



    if packaging_index is not None:
        for text in segment[packaging_index + 1:]:
            zone = _normalize_flipkart_zone(text)
            if zone:
                return zone
        
        for text in reversed(segment[:packaging_index]):
            zone = _normalize_flipkart_zone(text)
            if zone:
                return zone



    marker_index = next(

        (

            index

            for index, text in enumerate(segment)

            if text.strip().lower() == "not for resale."

        ),

        None,

    )



    if marker_index is None:

        return ""



    for text in segment[marker_index + 1:]:

        zone = _normalize_flipkart_zone(text)



        if zone:

            return zone



    for text in reversed(segment[:marker_index]):

        zone = _normalize_flipkart_zone(text)



        if zone:

            return zone



    return ""





def _extract_zones_from_xml(xml_path: str, manual_zones=None):

    """

    Reads the XML generated from the cropped PDF and extracts zones.



    Returns:

        {

            "zone_counts": Counter,

            "missing_pages": [

                {"page": 3},

                {"page": 17},

                ...

            ]

        }

    """



    manual_zones = manual_zones or {}

    tree = ET.parse(xml_path)

    root = tree.getroot()



    zone_counts = Counter()

    missing_pages = []

    label_number = 0

    report_date = _extract_flipkart_first_label_hbd_date(root)



    for page in root.findall("page"):



        page_no = int(page.attrib.get("number", 0))

        line_nodes = page.findall("line")



        lines = _flipkart_xml_page_values(page)



        if not line_nodes:

            marker_indices = [

                index

                for index, text in enumerate(lines)

                if text.strip().lower() == "not for resale."

            ]



            for marker_position, marker_index in enumerate(marker_indices):

                label_number += 1

                segment_start = (

                    marker_indices[marker_position - 1] + 1

                    if marker_position > 0

                    else 0

                )

                segment_end = (

                    marker_indices[marker_position + 1]

                    if marker_position + 1 < len(marker_indices)

                    else len(lines)

                )

                zone = _extract_flipkart_zone_from_text_segment(

                    lines[segment_start:segment_end]

                )

                manual_zone = _normalize_flipkart_zone(

                    manual_zones.get(str(label_number), "")

                )



                if manual_zone:

                    zone = manual_zone



                if not zone:

                    missing_pages.append({

                        "page": page_no,

                        "label_number": label_number,

                    })

                    continue



                zone_counts[zone] += 1



            continue



        for i, text in enumerate(lines):

            if text.strip().lower() != "not for resale.":

                continue



            label_number += 1

            zone = _extract_flipkart_zone_after_marker(lines, i)

            manual_zone = _normalize_flipkart_zone(

                manual_zones.get(str(label_number), "")

            )



            if manual_zone:

                zone = manual_zone



            if not zone:

                missing_pages.append({

                    "page": page_no,

                    "label_number": label_number,

                })

                continue



            zone_counts[zone] += 1



    return {

        "zone_counts": zone_counts,

        "missing_pages": missing_pages,

        "report_date": report_date or datetime.now().date(),

        "label_count": label_number,

    }



from openpyxl import Workbook

from openpyxl.styles import Font





def _generate_zone_summary_excel(zone_counts, output_path):

    """

    Generate Zone Summary.xlsx

    """



    wb = Workbook()

    ws = wb.active

    ws.title = "Zone Summary"



    ws.append(["Zone", "Count"])



    for cell in ws[1]:

        cell.font = Font(bold=True)



    total = 0



    for zone in sorted(zone_counts.keys()):

        count = int(zone_counts[zone])

        ws.append([zone, count])

        total += count



    ws.append([])

    ws.append(["Total Labels Processed", total])



    wb.save(output_path)



    return output_path





def _sync_flipkart_zone_totals(db: Session, report_date):

    db.query(FlipkartZoneReport).filter(

        FlipkartZoneReport.report_date == report_date,

    ).delete(synchronize_session=False)



    totals = Counter()

    batch_rows = (

        db.query(FlipkartZoneBatchItem)

        .filter(FlipkartZoneBatchItem.report_date == report_date)

        .all()

    )



    for row in batch_rows:

        totals[row.zone] += int(row.label_count or 0)



    for zone, label_count in sorted(totals.items()):

        db.add(

            FlipkartZoneReport(

                report_date=report_date,

                zone=zone,

                label_count=int(label_count or 0),

            )

        )





def _save_flipkart_zone_report(
    report_date,
    zone_counts,
    label_count=None,
    source_filename=None,
    platform="Flipkart",
):
    db: Session = SessionLocal()

    try:
        if source_filename:
            existing_batch = (
                db.query(FlipkartZoneBatch)
                .filter(
                    FlipkartZoneBatch.report_date == report_date,
                    FlipkartZoneBatch.source_filename == source_filename,
                    FlipkartZoneBatch.platform == platform
                )
                .first()
            )
            if existing_batch:
                db.query(FlipkartZoneBatchItem).filter(FlipkartZoneBatchItem.batch_id == existing_batch.id).delete()
                db.delete(existing_batch)
                db.flush()

        batch_label_count = (
            int(label_count)
            if label_count is not None
            else sum(int(value or 0) for value in zone_counts.values())
        )
        batch = FlipkartZoneBatch(
            report_date=report_date,
            source_filename=source_filename,
            platform=platform,
            label_count=batch_label_count,
        )
        db.add(batch)
        db.flush()

        for zone, label_count in sorted(zone_counts.items()):
            db.add(
                FlipkartZoneBatchItem(
                    batch_id=batch.id,
                    report_date=report_date,
                    zone=zone,
                    platform=platform,
                    label_count=int(label_count or 0),
                )
            )



        _sync_flipkart_zone_totals(

            db,

            report_date,

        )

        db.commit()

        return batch.id

    finally:

        db.close()





def _flipkart_zone_summary_rows(db: Session, report_date):
    batch_items = (
        db.query(FlipkartZoneBatchItem)
        .filter(FlipkartZoneBatchItem.report_date == report_date)
        .all()
    )

    platform_totals = {}
    for row in batch_items:
        platform = getattr(row, "platform", "Flipkart") or "Flipkart"
        if platform not in platform_totals:
            platform_totals[platform] = Counter()
        platform_totals[platform][row.zone] += int(row.label_count or 0)

    if not platform_totals:
        legacy_rows = (
            db.query(FlipkartZoneReport)
            .filter(FlipkartZoneReport.report_date == report_date)
            .order_by(FlipkartZoneReport.zone.asc())
            .all()
        )
        if legacy_rows:
            platform_totals["Flipkart"] = Counter()
            for row in legacy_rows:
                platform_totals["Flipkart"][row.zone] += int(row.label_count or 0)

    items = []
    for platform, totals in platform_totals.items():
        for zone, label_count in sorted(totals.items()):
            items.append({
                "platform": platform,
                "zone": zone,
                "label_count": int(label_count or 0),
            })
    
    # sort items by platform, then zone
    items.sort(key=lambda x: (x["platform"], x["zone"]))

    batches = [

        {
            "id": row.id,
            "label_count": int(row.label_count or 0),
            "source_filename": row.source_filename or "",
            "platform": getattr(row, "platform", "Flipkart") or "Flipkart",
            "created_at": row.created_at.isoformat() if row.created_at else "",
        }

        for row in (

            db.query(FlipkartZoneBatch)

            .filter(FlipkartZoneBatch.report_date == report_date)

            .order_by(FlipkartZoneBatch.created_at.asc())

            .all()

        )

    ]



    return {

        "date": str(report_date),

        "items": items,

        "total": sum(item["label_count"] for item in items),

        "batches": batches,

    }



MEESHO_LABEL_SIZE_TOKENS = {

    "XS",

    "S",

    "M",

    "L",

    "XL",

    "XXL",

    "2XL",

    "3XL",

    "4XL",

    "5XL",

    "6XL",

    "FREE",

    "FREESIZE",

    "FREE SIZE",

}



MEESHO_LABEL_SIZE_ORDER = [

    "XS",

    "S",

    "M",

    "L",

    "XL",

    "XXL",

    "3XL",

    "4XL",

]





def _natural_sort_key(value: str):

    return [

        int(part) if part.isdigit() else part.lower()

        for part in re.split(r"(\d+)", str(value or ""))

    ]





def _normalize_meesho_size(value: str):

    size = re.sub(r"\s+", " ", str(value or "").strip().upper())



    if size == "FREE SIZE":

        return size



    return size.replace(" ", "")





def _meesho_size_sort_key(value: str):

    size = _normalize_meesho_size(value)



    try:

        return (

            MEESHO_LABEL_SIZE_ORDER.index(size),

            "",

        )

    except ValueError:

        return (

            len(MEESHO_LABEL_SIZE_ORDER),

            size.lower(),

        )





def _meesho_sku_family_sort_key(value: str):

    sku = str(value or "").strip().upper()



    known_prefixes = [

        "I&D-VEST-POLY-PRNT",

    ]



    for prefix in known_prefixes:

        if sku.startswith(prefix):

            return prefix



    match = re.match(r"^(LSDSPLN|LSDS\d+|SN\d+)(?=[-_]|$)", sku)



    if match:

        return match.group(1)



    return sku





def _meesho_line_start_size(words):

    if not words:

        return None



    first_word = _normalize_meesho_size(words[0])



    if len(words) >= 2:

        first_two_words = _normalize_meesho_size(

            f"{words[0]} {words[1]}"

        )



        if first_two_words in MEESHO_LABEL_SIZE_TOKENS:

            return first_two_words



    if first_word in MEESHO_LABEL_SIZE_TOKENS:

        return first_word



    return None





def _find_meesho_size_in_words(words, start_index=0):

    for index in range(start_index, len(words)):

        size = _meesho_line_start_size(words[index:index + 2])



        if size:

            return index, size



    return None, None





def _extract_meesho_label_product(text: str):

    product_match = re.search(

        r"Product\s+Details",

        text or "",

        flags=re.IGNORECASE,

    )



    if not product_match:

        return None



    section = text[product_match.end():]

    header_match = re.search(

        r"SKU\s+Size\s+Qty\s+Color\s+Order\s+No\.?",

        section,

        flags=re.IGNORECASE,

    )



    if header_match:

        section = section[header_match.end():]



    section = re.split(

        r"TAX\s+INVOICE|BILL\s+TO|SHIP\s+TO|Purchase\s+Order",

        section,

        maxsplit=1,

        flags=re.IGNORECASE,

    )[0]



    lines = [

        line.strip()

        for line in section.splitlines()

        if line.strip()

    ]

    sku_parts = []

    size = None



    for line in lines[:8]:

        words = line.split()



        if not words:

            continue



        line_start_size = _meesho_line_start_size(words)

        if sku_parts and (

            line_start_size

            or re.fullmatch(r"\d+", words[0])

        ):

            size = size or line_start_size

            break



        size_index, line_size = _find_meesho_size_in_words(

            words,

            start_index=1,

        )



        if size_index is not None:

            sku_parts.append(

                " ".join(words[:size_index])

            )

            size = size or line_size

            break



        sku_parts.append(line)



    sku = "".join(sku_parts).replace(" ", "").strip(":-")

    return {

        "sku": sku or None,

        "size": size,

    }











def _extract_meesho_label_carrier(text: str):
    match = re.search(r"(Valmo|Xpress Bees|Delhivery|Shadowfax|Ecom Express)", str(text or ""), flags=re.IGNORECASE)
    if match:
        return match.group(1).strip().upper()
    return "UNKNOWN"

def _build_meesho_label_sorted_pdf(input_paths, output_path: str):

    from pypdf import PdfReader, PdfWriter



    writer = PdfWriter()
    page_entries = []
    zone_counts = Counter()
    input_paths = input_paths if isinstance(input_paths, list) else [input_paths]

    for file_index, input_path in enumerate(input_paths):
        reader = PdfReader(input_path)

        for page_index, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            product = _extract_meesho_label_product(text)
            carrier_name = _extract_meesho_label_carrier(text)
            zone_counts[carrier_name] += 1

            page_entries.append({
                "index": len(page_entries),
                "file_index": file_index,
                "page_index": page_index,
                "page": page,
                "sku": product["sku"],
                "size": product["size"],
                "carrier_name": carrier_name,
            })



    if not page_entries:

        raise HTTPException(

            status_code=400,

            detail="Meesho PDF does not contain any pages.",

        )



    sorted_entries = sorted(

        page_entries,

        key=lambda item: (

            item["sku"] is None,

            _natural_sort_key(

                _meesho_sku_family_sort_key(item["sku"] or "")

            ),

            item["size"] is None,

            _meesho_size_sort_key(item["size"] or ""),

            _natural_sort_key(item["sku"] or ""),

            item["index"],

        ),

    )



    for entry in sorted_entries:

        writer.add_page(entry["page"])



    if reader.metadata:

        writer.add_metadata(

            {

                key: value

                for key, value in reader.metadata.items()

                if value is not None

            }

        )



    with open(output_path, "wb") as output_file:

        writer.write(output_file)



    return {
        "label_count": len(writer.pages),
        "zone_counts": zone_counts,
        "missing_pages": [],
        "report_date": datetime.now().date(),
        "readable_sku_count": sum(
            1
            for entry in page_entries
            if entry["sku"]
        ),
        "readable_size_count": sum(
            1
            for entry in page_entries
            if entry["size"]
        ),
    }





def _cropped_output_filename(upload_file: UploadFile):

    clean_name = _clean_upload_filename(

        upload_file.filename

    )

    name_without_extension, _ = os.path.splitext(clean_name)

    return f"{name_without_extension} - cropped.pdf"





def _meesho_sorted_output_filename(upload_files):

    clean_names = [

        os.path.splitext(_clean_upload_filename(file.filename))[0]

        for file in upload_files

        if file and file.filename

    ]

    base_name = " + ".join(clean_names) or "meesho-labels"



    if len(base_name) > 140:

        base_name = f"{clean_names[0]} + {len(clean_names) - 1} more"



    return f"{base_name} sorted.pdf"





@router.post("/label-cropper")

def label_cropper(

    flipkart_file: UploadFile = File(None),

    amazon_file: UploadFile = File(None),

    ajio_label_file: UploadFile = File(None),

    ajio_label_files: Optional[List[UploadFile]] = File(None),

    ajio_invoice_file: UploadFile = File(None),

    ajio_invoice_files: Optional[List[UploadFile]] = File(None),

    meesho_file: UploadFile = File(None),

    meesho_files: Optional[List[UploadFile]] = File(None),

    amazon_manual_entries: str = Form(None),

    flipkart_manual_zones: str = Form(None),

):

    ajio_label_uploads = [

        file

        for file in (

            ([ajio_label_file] if ajio_label_file else [])

            + (ajio_label_files or [])

        )

        if file

    ]

    ajio_invoice_uploads = [

        file

        for file in (

            ([ajio_invoice_file] if ajio_invoice_file else [])

            + (ajio_invoice_files or [])

        )

        if file

    ]

    meesho_uploads = [

        file

        for file in ([meesho_file] if meesho_file else []) + (meesho_files or [])

        if file

    ]

    uploaded_files = [

        file

        for file in (

            amazon_file,

            flipkart_file,

            *ajio_label_uploads,

            *ajio_invoice_uploads,

            *meesho_uploads,

        )

        if file

    ]



    if not uploaded_files:

        raise HTTPException(

            status_code=400,

            detail="Upload a Flipkart, Amazon, AJIO, or Meesho label PDF.",

        )



    selected_platform_count = sum(

        [

            bool(amazon_file),

            bool(flipkart_file),

            bool(ajio_label_uploads or ajio_invoice_uploads),

            bool(meesho_uploads),

        ]

    )



    if selected_platform_count > 1:

        raise HTTPException(

            status_code=400,

            detail="Upload one platform label type per request.",

        )



    if (ajio_label_uploads and not ajio_invoice_uploads) or (

        ajio_invoice_uploads and not ajio_label_uploads

    ):

        raise HTTPException(

            status_code=400,

            detail="Upload AJIO shipping labels and customer invoices together.",

        )



    if amazon_file and not amazon_file.filename.lower().endswith(".pdf"):

        raise HTTPException(

            status_code=400,

            detail="Amazon label cropper accepts PDF files only.",

        )



    if flipkart_file and not flipkart_file.filename.lower().endswith(".pdf"):

        raise HTTPException(

            status_code=400,

            detail="Flipkart label cropper accepts PDF files only.",

        )



    for file in ajio_label_uploads + ajio_invoice_uploads:

        if not file.filename.lower().endswith(".pdf"):

            raise HTTPException(

                status_code=400,

                detail="AJIO label cropper accepts PDF files only.",

            )



    for file in meesho_uploads:

        if not file.filename.lower().endswith(".pdf"):

            raise HTTPException(

                status_code=400,

                detail="Meesho label sorter accepts PDF files only.",

            )



    upload_file = (

        amazon_file

        or flipkart_file

        or (ajio_label_uploads[0] if ajio_label_uploads else None)

        or (meesho_uploads[0] if meesho_uploads else None)

    )

    input_path = _save_upload(upload_file) if (amazon_file or flipkart_file) else None

    ajio_label_input_paths = [

        _save_upload(file)

        for file in ajio_label_uploads

    ]

    ajio_invoice_input_paths = [

        _save_upload(file)

        for file in ajio_invoice_uploads

    ]

    meesho_input_paths = [

        _save_upload(file)

        for file in meesho_uploads

    ]

    timestamp = datetime.now().strftime("%d-%m-%Y_%H%M%S")

    platform = (

        "amazon" if amazon_file

        else "flipkart" if flipkart_file

        else "ajio" if ajio_label_uploads

        else "meesho"

    )

    output_filename = (

        _meesho_sorted_output_filename(meesho_uploads)

        if meesho_uploads

        else _cropped_output_filename(upload_file)

    )

    output_path = os.path.join(

        UPLOAD_FOLDER,

        f"{platform}_{timestamp}_{output_filename}",

    )



    try:

        if amazon_file:

            manual_entries = {}



            if amazon_manual_entries:

                try:

                    manual_entries = json.loads(amazon_manual_entries)

                except json.JSONDecodeError as error:

                    raise HTTPException(

                        status_code=400,

                        detail="Invalid manual SKU entries.",

                    ) from error



            _build_amazon_label_cropper_pdf(

                input_path,

                output_path,

                manual_entries=manual_entries,

            )

        elif flipkart_file:

            manual_zones = {}



            if flipkart_manual_zones:

                try:

                    manual_zones = json.loads(flipkart_manual_zones)

                except json.JSONDecodeError as error:

                    raise HTTPException(

                        status_code=400,

                        detail="Invalid manual Flipkart zones.",

                    ) from error



            flipkart_result = _build_flipkart_label_cropper_pdf(

                input_path,

                output_path,

                manual_zones=manual_zones,

            )

        elif ajio_label_uploads:
            ajio_result = _build_ajio_label_cropper_pdf(
                ajio_label_input_paths,
                ajio_invoice_input_paths,
                output_path,
            )
        else:
            meesho_result = _build_meesho_label_sorted_pdf(meesho_input_paths, output_path)

    except ImportError as error:
        import traceback
        traceback.print_exc()   # prints the real error in terminal
        raise HTTPException(
            status_code=500,
            detail=(
                "PDF tools are missing. Install pypdf and reportlab, then try again."
            ),
        ) from error

    if flipkart_file or ajio_label_uploads or meesho_uploads:
        result = flipkart_result if flipkart_file else (ajio_result if ajio_label_uploads else meesho_result)
        source_file = flipkart_file if flipkart_file else (ajio_label_uploads[0] if ajio_label_uploads else meesho_uploads[0])
        if result.get("missing_pages"):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "FLIPKART_ZONES_MISSING",
                    "items": result["missing_pages"],
                    "zone_counts": result["zone_counts"],
                    "label_count": result["label_count"],
                },
            )

        platform_name = "Flipkart" if flipkart_file else ("Ajio" if ajio_label_uploads else "Meesho")
        _save_flipkart_zone_report(
            result["report_date"],
            result["zone_counts"],
            label_count=result["label_count"],
            source_filename=_clean_upload_filename(source_file.filename),
            platform=platform_name,
        )
        zone_summary_header = json.dumps(
            {
                "report_date": str(result["report_date"]),
                "label_count": sum(
                    int(value or 0)
                    for value in result["zone_counts"].values()
                ),
                "zone_counts": result["zone_counts"],
            }
        )

        return FileResponse(
            output_path,
            media_type="application/pdf",
            filename=output_filename,
            headers={
                "X-Flipkart-Zone-Summary": zone_summary_header,
            },
        )

    return FileResponse(

        output_path,

        media_type="application/pdf",

        filename=output_filename,

    )





@router.get(

    "/current-sku-master"

)

def current_sku_master():



    upload_folder = UPLOAD_FOLDER



    if not os.path.exists(

        upload_folder

    ):



        return {

            "filename": None

        }



    files = os.listdir(

        upload_folder

    )



    sku_files = [



        file for file in files



        if (

            "sku" in file.lower()

            or

            "master" in file.lower()

        )

    ]



    if not sku_files:



        return {

            "filename": None

        }



    latest_file = max(



        sku_files,



        key=lambda f: os.path.getctime(

            os.path.join(

                upload_folder,

                f

            )

        )

    )



    return {

        "filename": latest_file

    }



@router.get("/stock-alerts")

def stock_alerts(

    threshold: int = Query(250)

):

    db: Session = SessionLocal()



    try:

        

        ensure_sticker_inventory(db)



        stock_rows = (

            db.query(StockInventory)

            .join(

                SKUMaster, 

                (StockInventory.style == SKUMaster.style) & (StockInventory.size == SKUMaster.size)

            )

            .join(

                SKUPiece, 

                (SKUMaster.id == SKUPiece.sku_master_id) & (StockInventory.color == SKUPiece.color)

            )

            .filter(

                StockInventory.qty < threshold

            )

            .order_by(

                StockInventory.qty.asc(),

                StockInventory.style.asc(),

                StockInventory.color.asc(),

                StockInventory.size.asc(),

            )

            .all()

        )



        sticker_rows = (

            db.query(StickerInventory)

            .filter(

                StickerInventory.qty < threshold

            )

            .order_by(

                StickerInventory.qty.asc(),

                StickerInventory.style.asc(),

                StickerInventory.color.asc(),

            )

            .all()

        )



        stock_items = [

            {

                "id": row.id,

                "style": row.style,

                "color": row.color,

                "size": row.size,

                "qty": row.qty,

                "type": "piece",

            }

            for row in stock_rows

        ]



        sticker_items = [

            {

                "id": row.id,

                "style": row.style,

                "color": row.color,

                "qty": row.qty,

                "type": "sticker",

            }

            for row in sticker_rows

        ]



        return {

            "count": len(stock_items) + len(sticker_items),

            "stock_count": len(stock_items),

            "sticker_count": len(sticker_items),

            "threshold": threshold,

            "items": stock_items + sticker_items,

            "stock_items": stock_items,

            "sticker_items": sticker_items,

        }



    finally:

        db.close()



@router.delete(

    "/delete-sku-master"

)

def delete_sku_master(db: Session = Depends(get_db)):



    upload_folder = UPLOAD_FOLDER



    if os.path.exists(

        upload_folder

    ):



        files = os.listdir(

            upload_folder

        )



        for file in files:



            if (

                "sku" in file.lower()

                or

                "master" in file.lower()

            ):



                file_path = os.path.join(

                    upload_folder,

                    file

                )



                if os.path.exists(

                    file_path

                ):



                    try:



                        os.remove(

                            file_path

                        )



                    except PermissionError:



                        raise HTTPException(

                            status_code=409,

                            detail=(

                                "Cannot delete SKU master because the uploaded "

                                f"file is still open or locked: {file}. Close "

                                "Excel if it is open, restart the backend if "

                                "needed, then try again."

                            ),

                        )



                    except OSError as e:



                        raise HTTPException(

                            status_code=500,

                            detail=f"Could not delete SKU master file {file}: {e}",

                        )



    db.query(

        SKUPiece

    ).delete()



    db.query(

        SKUMaster

    ).delete()



    db.commit()



    return {



        "message":

        "SKU master deleted successfully"

    }





@router.delete("/clear-cache")
def api_clear_cache():
    import tempfile
    import glob
    snapshot_dir = os.path.join(tempfile.gettempdir(), "inventory_snapshot_cache")
    count = 0
    if os.path.exists(snapshot_dir):
        for f in glob.glob(os.path.join(snapshot_dir, "*.json")):
            try:
                os.remove(f)
                count += 1
            except:
                pass
    return {"message": f"Successfully cleared {count} JSON cache files."}



@router.get("/dashboard-stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    from app.models.daily_sales_report import DailySalesReport
    from sqlalchemy import func
    
    total_orders = db.query(func.sum(DailySalesReport.total_orders)).scalar() or 0
    total_pieces = db.query(func.sum(DailySalesReport.total_piece_qty)).scalar() or 0
    
    latest_date = db.query(func.max(DailySalesReport.report_date)).scalar()
    daily_orders = 0
    if latest_date:
        daily_orders = db.query(func.sum(DailySalesReport.total_orders)).filter(DailySalesReport.report_date == latest_date).scalar() or 0
        
    return {
        "total_orders": total_orders,
        "total_pieces": total_pieces,
        "daily_orders": daily_orders,
        "latest_date": latest_date.isoformat() if latest_date else None
    }


@router.get("/packing-inventory/usage-history")
def get_packing_usage_history(item_type: str = Query(None)):
    db: Session = SessionLocal()
    try:
        from datetime import timedelta, date
        start_date = date.today() - timedelta(days=30)
        
        query = db.query(PackingInventoryUsage).filter(PackingInventoryUsage.usage_date >= start_date)
        if item_type and item_type != "all":
            query = query.filter(PackingInventoryUsage.item_type == item_type)
            
        rows = query.order_by(PackingInventoryUsage.usage_date.asc()).all()
        
        # Group by date for the chart
        grouped = {}
        for r in rows:
            d_str = r.usage_date.strftime("%Y-%m-%d")
            if d_str not in grouped:
                grouped[d_str] = {"date": d_str, "used": 0, "restocked": 0}
            grouped[d_str]["used"] += (r.used_qty or 0)
            grouped[d_str]["restocked"] += (r.restocked_qty or 0)
            
        history = list(grouped.values())
        history.sort(key=lambda x: x["date"])
        
        return {"history": history}
    finally:
        db.close()
